from __future__ import annotations

import datetime
from typing import Iterable

from django.contrib.auth import get_user_model
from django.db.models import Prefetch, Q
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from journal.academic_year_context import filter_temporary_credentials_for_year
from journal.models import (
    AcademicYear,
    AssessmentGroup,
    AssessmentItem,
    AssessmentResult,
    CourseApplication,
    FinalGradeRule,
    Grade,
    GroupSubject,
    Instrument,
    StudentAssessmentGroup,
    StudentEnrollment,
    StudentSubject,
    Subject,
    SubjectResult,
    TeacherEnrollment,
    TemporaryCredential,
)


HEADER_FILL = PatternFill('solid', fgColor='D9EAF7')
HEADER_FONT = Font(bold=True)
DEFAULT_COLUMN_WIDTH = 18
MAX_COLUMN_WIDTH = 60
TEXT_NUMBER_FORMAT = '@'
EXPORT_VALUE_LABELS = {
    Subject.FINAL_GRADE_TYPE_NUMERIC: 'Оценка (1–5)',
    Subject.FINAL_GRADE_TYPE_PASS_FAIL: 'Зачёт / Незачёт',
}


def localized_export_value(value, fallback=''):
    return EXPORT_VALUE_LABELS.get(value, fallback or value or '')


def build_full_export_workbook(academic_year: AcademicYear | None = None) -> Workbook:
    """Build a user-facing export for the administrator-selected academic year."""
    academic_year = academic_year or AcademicYear.get_active()
    workbook = Workbook()
    workbook.remove(workbook.active)

    write_readme_sheet(workbook, academic_year)
    write_users_sheet(workbook, academic_year)
    write_reference_sheets(workbook, academic_year)
    write_students_sheet(workbook, academic_year)
    write_teachers_sheet(workbook, academic_year)
    write_assignments_sheets(workbook, academic_year)
    write_grades_sheet(workbook, academic_year)
    write_results_sheet(workbook, academic_year)
    write_assessment_sheets(workbook, academic_year)
    write_applications_sheet(workbook, academic_year)
    write_credentials_sheet(workbook, academic_year)

    return workbook


def write_readme_sheet(workbook: Workbook, academic_year: AcademicYear | None) -> None:
    worksheet = workbook.create_sheet('Описание')
    now = timezone.localtime()
    rows = [
        ['Файл', 'Выгрузка данных электронного журнала'],
        ['Учебный год', academic_year.name if academic_year else 'Не выбран'],
        ['Дата выгрузки', now.strftime('%d.%m.%Y %H:%M:%S')],
        ['Формат', 'Пользовательские данные без внутренних идентификаторов и служебных полей'],
        ['Важно', 'Файл содержит персональные данные. Храните и передавайте его безопасно.'],
    ]
    for row in rows:
        worksheet.append(row)
    format_sheet(worksheet)


def write_users_sheet(workbook: Workbook, academic_year: AcademicYear | None) -> None:
    User = get_user_model()
    user_ids: set[int] = set()
    if academic_year is not None:
        user_ids.update(
            StudentEnrollment.objects.filter(academic_year=academic_year)
            .exclude(student__user__isnull=True)
            .values_list('student__user_id', flat=True)
        )
        user_ids.update(
            TeacherEnrollment.objects.filter(academic_year=academic_year)
            .exclude(teacher__user__isnull=True)
            .values_list('teacher__user_id', flat=True)
        )
    user_filter = Q(pk__in=user_ids)
    if academic_year is not None and academic_year.is_active:
        user_filter |= Q(is_staff=True)
    users = (
        User.objects.filter(user_filter)
        .select_related('teacher_profile', 'student_profile')
        .prefetch_related('groups')
        .order_by('username')
    )

    def role(user):
        names = set(user.groups.values_list('name', flat=True))
        if user.is_superuser or user.is_staff:
            return 'Администратор'
        if 'Преподаватель' in names or hasattr(user, 'teacher_profile'):
            return 'Преподаватель'
        if 'Ученик' in names or hasattr(user, 'student_profile'):
            return 'Ученик'
        return 'Пользователь'

    write_custom_sheet(
        workbook,
        'Пользователи',
        users,
        [
            ('Логин', lambda user: user.username),
            ('ФИО', lambda user: user.get_full_name()),
            ('Email', lambda user: user.email),
            ('Роль', role),
        ],
        text_headers={'Логин'},
    )


def write_reference_sheets(workbook: Workbook, academic_year: AcademicYear | None) -> None:
    subject_ids: set[int] = set()
    if academic_year is not None:
        subject_ids.update(
            GroupSubject.objects.filter(group__academic_year=academic_year)
            .values_list('subject_id', flat=True)
        )
        subject_ids.update(
            StudentSubject.objects.filter(academic_year=academic_year)
            .values_list('subject_id', flat=True)
        )
        subject_ids.update(
            Grade.objects.filter(academic_year=academic_year)
            .values_list('subject_id', flat=True)
        )
        subject_ids.update(
            SubjectResult.objects.filter(academic_year=academic_year)
            .values_list('subject_id', flat=True)
        )
        subject_ids.update(
            AssessmentGroup.objects.filter(academic_year=academic_year)
            .values_list('subject_id', flat=True)
        )
    write_custom_sheet(
        workbook,
        'Учебные годы',
        AcademicYear.objects.filter(pk=getattr(academic_year, 'pk', None)),
        [
            ('Учебный год', lambda item: item.name),
            ('Начало', lambda item: item.starts_on),
            ('Окончание', lambda item: item.ends_on),
        ],
    )
    write_custom_sheet(
        workbook,
        'Инструменты',
        Instrument.objects.order_by('name'),
        [('Название инструмента', lambda item: item.name)],
    )
    write_custom_sheet(
        workbook,
        'Предметы',
        Subject.objects.filter(pk__in=subject_ids).order_by('name'),
        [
            ('Название предмета', lambda item: item.name),
            ('Тип предмета', lambda item: 'Индивидуальный' if item.is_specialty else 'Групповой'),
            ('Режим ведения', lambda item: item.get_assessment_mode_display()),
            (
                'Тип итоговой оценки',
                lambda item: localized_export_value(
                    item.final_grade_type,
                    item.get_final_grade_type_display(),
                ),
            ),
        ],
    )


def write_students_sheet(workbook: Workbook, academic_year: AcademicYear | None) -> None:
    if academic_year is None:
        enrollments = StudentEnrollment.objects.none()
    else:
        enrollments = (
            StudentEnrollment.objects.filter(academic_year=academic_year)
            .select_related('student', 'group', 'academic_year')
            .prefetch_related(
                Prefetch(
                    'student__course_applications',
                    queryset=CourseApplication.objects.filter(academic_year=academic_year).order_by('-registration_date'),
                    to_attr='export_course_applications',
                )
            )
            .order_by('full_name')
        )

    def registration_date(enrollment):
        applications = getattr(enrollment.student, 'export_course_applications', ())
        return applications[0].registration_date if applications else None

    write_custom_sheet(
        workbook,
        'Ученики',
        enrollments,
        [
            ('Ученик', lambda item: item.full_name),
            ('Пол', lambda item: item.get_gender_display()),
            ('Дата рождения', lambda item: item.birth_date),
            ('Группа', lambda item: item.group.name if item.group_id else ''),
            ('Инструмент', lambda item: item.instrument_name),
            ('Партия в оркестре', lambda item: item.orchestra_part),
            ('Музыкальное образование', lambda item: item.get_music_education_display()),
            ('Телефон ученика', lambda item: item.student_phone),
            ('Телефон родителей', lambda item: item.parent_contacts),
            ('Город / Церковь', lambda item: item.city_church),
            ('Комментарий', lambda item: item.comments),
            ('Активен', lambda item: item.is_active),
            ('Дата регистрации заявки', registration_date),
            ('Учебный год', lambda item: item.academic_year.name),
        ],
        text_headers={'Телефон ученика', 'Телефон родителей'},
    )


def write_teachers_sheet(workbook: Workbook, academic_year: AcademicYear | None) -> None:
    memberships = (
        TeacherEnrollment.objects.filter(academic_year=academic_year)
        .select_related('teacher', 'academic_year')
        .order_by('teacher__full_name')
        if academic_year else TeacherEnrollment.objects.none()
    )
    write_custom_sheet(
        workbook,
        'Преподаватели',
        memberships,
        [
            ('Преподаватель', lambda item: item.teacher.full_name),
            ('Дата рождения', lambda item: item.teacher.birth_date),
            ('Телефон', lambda item: item.teacher.phone),
            ('Email', lambda item: item.teacher.email),
            ('Комментарий', lambda item: item.teacher.comments),
            ('Активен', lambda item: item.is_active),
            ('Учебный год', lambda item: item.academic_year.name),
        ],
        text_headers={'Телефон'},
    )


def write_assignments_sheets(workbook: Workbook, academic_year: AcademicYear | None) -> None:
    groups = (
        academic_year.study_groups.all().order_by('name')
        if academic_year else AcademicYear.objects.none()
    )
    write_custom_sheet(
        workbook,
        'Группы',
        groups,
        [
            ('Название группы', lambda item: item.name),
            ('Активна', lambda item: item.is_active),
            ('Учебный год', lambda item: item.academic_year.name),
        ],
    )
    group_subjects = (
        GroupSubject.objects.filter(group__academic_year=academic_year)
        .select_related('group', 'subject', 'teacher', 'group__academic_year')
        .order_by('group__name', 'subject__name')
        if academic_year else GroupSubject.objects.none()
    )
    write_custom_sheet(
        workbook,
        'Предметы групп',
        group_subjects,
        [
            ('Группа', lambda item: item.group.name),
            ('Предмет', lambda item: item.subject_name_snapshot or item.subject.name),
            ('Преподаватель', lambda item: item.teacher_name_snapshot or item.teacher.full_name),
            ('Активен', lambda item: item.is_active),
            ('Учебный год', lambda item: item.group.academic_year.name),
        ],
    )
    individual = (
        StudentSubject.objects.filter(academic_year=academic_year)
        .select_related('student', 'subject', 'teacher', 'academic_year')
        .order_by('student__full_name', 'subject__name')
        if academic_year else StudentSubject.objects.none()
    )
    write_custom_sheet(
        workbook,
        'Индивидуальные предметы',
        individual,
        [
            ('Ученик', lambda item: item.student.full_name),
            ('Индивидуальный предмет', lambda item: item.subject_name_snapshot or item.subject.name),
            ('Преподаватель', lambda item: item.teacher_name_snapshot or item.teacher.full_name),
            ('Активен', lambda item: item.is_active),
            ('Учебный год', lambda item: item.academic_year.name),
        ],
    )


def write_grades_sheet(workbook: Workbook, academic_year: AcademicYear | None) -> None:
    grades = (
        Grade.objects.filter(academic_year=academic_year)
        .select_related('student', 'enrollment', 'enrollment__group', 'subject', 'teacher', 'academic_year')
        .order_by('student__full_name', 'subject__name', 'date')
        if academic_year else Grade.objects.none()
    )
    write_custom_sheet(
        workbook,
        'Оценки',
        grades,
        [
            ('Ученик', lambda item: item.student_name_snapshot or item.student.full_name),
            ('Группа ученика', lambda item: item.group_name_snapshot or (item.enrollment.group.name if item.enrollment_id and item.enrollment.group_id else '')),
            ('Предмет', lambda item: item.subject_name_snapshot or item.subject.name),
            ('Преподаватель', lambda item: item.teacher_name_snapshot or item.teacher.full_name),
            ('Оценка', lambda item: item.value),
            ('Дата оценки', lambda item: item.date),
            ('Комментарий', lambda item: item.comment),
            ('Учебный год', lambda item: item.academic_year.name),
        ],
        text_headers={'Оценка'},
    )


def write_results_sheet(workbook: Workbook, academic_year: AcademicYear | None) -> None:
    results = (
        SubjectResult.objects.filter(academic_year=academic_year)
        .select_related('student', 'subject', 'academic_year', 'enrollment', 'enrollment__group')
        .order_by('student__full_name', 'subject__name')
        if academic_year else SubjectResult.objects.none()
    )
    write_custom_sheet(
        workbook,
        'Итоги',
        results,
        [
            ('Ученик', lambda item: item.student_name_snapshot or item.student.full_name),
            ('Предмет', lambda item: item.subject_name_snapshot or item.subject.name),
            ('Экзамен', lambda item: item.exam_grade),
            ('Итоговая оценка', lambda item: item.final_grade),
            ('Группа', lambda item: item.group_name_snapshot or (item.enrollment.group.name if item.enrollment_id and item.enrollment.group_id else '')),
            ('Учебный год', lambda item: item.academic_year.name),
        ],
        text_headers={'Экзамен', 'Итоговая оценка'},
    )


def write_assessment_sheets(workbook: Workbook, academic_year: AcademicYear | None) -> None:
    groups = (
        AssessmentGroup.objects.filter(academic_year=academic_year)
        .select_related('subject', 'academic_year')
        .order_by('subject__name', 'sort_order', 'name')
        if academic_year else AssessmentGroup.objects.none()
    )
    write_custom_sheet(
        workbook,
        'Группы произведений',
        groups,
        [
            ('Группа произведений', lambda item: item.name),
            ('Предмет', lambda item: item.subject.name),
            ('Описание', lambda item: item.description),
            ('Активна', lambda item: item.is_active),
            ('Учебный год', lambda item: item.academic_year.name),
        ],
    )
    items = (
        AssessmentItem.objects.filter(academic_year=academic_year)
        .select_related('group', 'subject', 'responsible_teacher', 'academic_year')
        .order_by('group__sort_order', 'sort_order', 'title')
        if academic_year else AssessmentItem.objects.none()
    )
    write_custom_sheet(
        workbook,
        'Произведения',
        items,
        [
            ('Произведение', lambda item: item.title),
            ('Группа произведений', lambda item: item.group.name),
            ('Предмет', lambda item: item.subject.name),
            ('Преподаватель-дирижёр', lambda item: item.responsible_teacher.full_name if item.responsible_teacher_id else ''),
            ('Обязательное', lambda item: yes_no(item.is_required)),
            ('Описание', lambda item: item.description),
            ('Активно', lambda item: item.is_active),
            ('Учебный год', lambda item: item.academic_year.name),
        ],
    )
    assignments = (
        StudentAssessmentGroup.objects.filter(academic_year=academic_year)
        .select_related('student', 'assessment_group', 'assessment_group__subject', 'academic_year')
        .order_by('student__full_name', 'assessment_group__subject__name', 'assessment_group__name')
        if academic_year else StudentAssessmentGroup.objects.none()
    )
    write_custom_sheet(
        workbook,
        'Назначения групп произведений',
        assignments,
        [
            ('Ученик', lambda item: item.student.full_name),
            ('Группа произведений', lambda item: item.assessment_group.name),
            ('Предмет', lambda item: item.assessment_group.subject.name),
            ('Активно', lambda item: item.is_active),
            ('Учебный год', lambda item: item.academic_year.name),
        ],
    )
    assessment_results = (
        AssessmentResult.objects.filter(item__academic_year=academic_year)
        .select_related('enrollment', 'item', 'item__group', 'item__subject', 'item__academic_year', 'assessed_by')
        .order_by('enrollment__full_name', 'item__title')
        if academic_year else AssessmentResult.objects.none()
    )
    write_custom_sheet(
        workbook,
        'Результаты произведений',
        assessment_results,
        [
            ('Ученик', lambda item: item.enrollment.full_name),
            ('Произведение', lambda item: item.item.title),
            ('Группа произведений', lambda item: item.item.group.name),
            ('Предмет', lambda item: item.item.subject.name),
            ('Результат', lambda item: item.get_status_display()),
            ('Преподаватель', lambda item: item.assessed_by.full_name),
            ('Дата результата', lambda item: item.assessed_at),
            ('Комментарий', lambda item: item.comment),
            ('Учебный год', lambda item: item.item.academic_year.name),
        ],
    )
    rules = (
        FinalGradeRule.objects.filter(academic_year=academic_year)
        .select_related('subject', 'assessment_group', 'academic_year')
        .order_by('subject__name', 'priority')
        if academic_year else FinalGradeRule.objects.none()
    )
    write_custom_sheet(
        workbook,
        'Правила итоговых оценок',
        rules,
        [
            ('Предмет', lambda item: item.subject.name),
            ('Группа произведений', lambda item: item.assessment_group.name if item.assessment_group_id else 'Общее правило'),
            ('Тип правила', lambda item: item.get_rule_type_display()),
            ('Количество зачётов', lambda item: item.passed_count),
            ('Условие выполнено', lambda item: '' if item.condition_value is None else yes_no(item.condition_value)),
            ('Итоговая оценка', lambda item: item.grade),
            ('Учебный год', lambda item: item.academic_year.name),
        ],
        text_headers={'Итоговая оценка'},
    )


def write_applications_sheet(workbook: Workbook, academic_year: AcademicYear | None) -> None:
    applications = (
        CourseApplication.objects.filter(academic_year=academic_year)
        .select_related('academic_year', 'instrument_reference')
        .order_by('last_name', 'first_name', 'middle_name')
        if academic_year else CourseApplication.objects.none()
    )
    write_custom_sheet(
        workbook,
        'Заявки',
        applications,
        [
            ('Ученик', lambda item: item.full_name),
            ('Пол', lambda item: item.get_gender_display()),
            ('Дата рождения', lambda item: item.birth_date),
            ('Город / Церковь', lambda item: item.city_church),
            ('Инструмент', lambda item: item.instrument),
            ('Партия в оркестре', lambda item: item.orchestra_part),
            ('Музыкальное образование', lambda item: item.get_music_education_display()),
            ('Телефон ученика', lambda item: item.student_phone),
            ('Телефон родителей', lambda item: item.parent_contacts),
            ('Комментарий', lambda item: item.comments),
            ('Статус заявки', lambda item: item.get_status_display()),
            ('Дата регистрации заявки', lambda item: item.registration_date),
            ('Учебный год', lambda item: item.academic_year.name),
        ],
        text_headers={'Телефон ученика', 'Телефон родителей'},
    )


def write_credentials_sheet(workbook: Workbook, academic_year: AcademicYear | None) -> None:
    credentials = TemporaryCredential.objects.all().select_related(
        'course_application',
        'course_application__academic_year',
        'user',
        'user__student_profile',
        'user__teacher_profile',
    )
    credentials = filter_temporary_credentials_for_year(credentials, academic_year)
    credentials = credentials.order_by('login')

    def owner_name(item):
        if item.course_application_id:
            return item.course_application.full_name
        if item.user_id:
            student = getattr(item.user, 'student_profile', None)
            if student is not None:
                return student.full_name
            teacher = getattr(item.user, 'teacher_profile', None)
            if teacher is not None:
                return teacher.full_name
            return item.user.get_full_name() or item.user.username
        return ''

    def owner_role(item):
        if item.user_id and (item.user.is_superuser or item.user.is_staff):
            return 'Администратор'
        if item.user_id and getattr(item.user, 'teacher_profile', None) is not None:
            return 'Преподаватель'
        if item.course_application_id or (
            item.user_id and getattr(item.user, 'student_profile', None) is not None
        ):
            return 'Ученик'
        return 'Пользователь'

    write_custom_sheet(
        workbook,
        'Временные доступы',
        credentials,
        [
            ('ФИО', owner_name),
            ('Роль', owner_role),
            ('Логин', lambda item: item.login),
            ('Временный пароль', lambda item: item.temporary_password),
            ('Телефон ученика', lambda item: item.student_phone),
            ('Дата выдачи', lambda item: item.created_at),
            ('Учебный год', lambda _item: academic_year.name if academic_year else ''),
        ],
        text_headers={'Логин', 'Временный пароль', 'Телефон ученика'},
    )


def write_custom_sheet(
    workbook: Workbook,
    title: str,
    queryset: Iterable,
    columns: list[tuple[str, callable]],
    *,
    text_headers: set[str] | None = None,
) -> None:
    worksheet = workbook.create_sheet(safe_sheet_title(title))
    headers = [header for header, _getter in columns]
    worksheet.append(headers)
    text_headers = text_headers or set()
    text_columns = {index + 1 for index, header in enumerate(headers) if header in text_headers}

    for obj in queryset:
        values = [format_value(getter(obj)) for _header, getter in columns]
        worksheet.append(values)
        row_number = worksheet.max_row
        for column_number in text_columns:
            cell = worksheet.cell(row=row_number, column=column_number)
            if cell.value is not None:
                cell.value = str(cell.value)
                cell.number_format = TEXT_NUMBER_FORMAT

    format_sheet(worksheet)


def format_value(value):
    if value is None:
        return ''
    if isinstance(value, bool):
        return yes_no(value)
    if isinstance(value, datetime.datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.replace(tzinfo=None)
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, (int, float)):
        return value
    return clean_excel_text(str(value))


def clean_excel_text(value: str) -> str:
    value = ILLEGAL_CHARACTERS_RE.sub('', value)
    if value.lstrip().startswith(('=', '+', '-', '@')):
        return f"'{value}"
    return value


def yes_no(value: bool) -> str:
    return 'Да' if value else 'Нет'


def safe_sheet_title(title: str) -> str:
    for char in ['\\', '/', '*', '[', ']', ':', '?']:
        title = title.replace(char, ' ')
    return (title.strip() or 'Лист')[:31]


def format_sheet(worksheet) -> None:
    worksheet.freeze_panes = 'A2'
    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, DEFAULT_COLUMN_WIDTH),
            MAX_COLUMN_WIDTH,
        )
