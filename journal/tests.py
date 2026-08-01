from __future__ import annotations

import logging
import json
from datetime import date
from threading import Barrier, Lock, Thread
from io import BytesIO, StringIO
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest import skipUnless
from unittest.mock import patch

from django import forms
from django.forms.models import inlineformset_factory
from django.apps import apps
from django.conf import settings
from django.contrib import admin as django_admin
from django.contrib.auth import get_user_model
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import Group
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.management import CommandError, call_command
from django.db import IntegrityError, close_old_connections, connection, transaction
from django.db.models import Count, Q
from django.http import HttpResponse
from django.test import Client, RequestFactory, SimpleTestCase, TestCase, TransactionTestCase, override_settings, tag
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from openpyxl import load_workbook

from journal.academic_year_context import (
    academic_year_ids_for_user,
    filter_temporary_credentials_for_year,
    get_admin_academic_year_context,
)
from journal.assignment_availability import (
    available_groups,
    available_students,
    available_subjects,
    available_teachers,
)
from journal.assessment_services import (
    assessment_items_for_teacher,
    assessment_sections_for_teacher,
    available_assessment_items_for_student,
    set_assessment_result,
)
from journal.birthday_notifications import birthday_notifications_for_user
from journal.services.excel_export import build_full_export_workbook
from journal.assignment_options import assignment_teacher_queryset
from journal.account_utils import (
    build_course_application_login,
    build_display_name_from_full_name,
    build_username_from_full_name,
    display_name_for_user,
    ensure_temporary_credential_for_user,
    generate_temporary_password,
    split_user_name,
)
from journal.admin import (
    AcademicYearHistoryInlineForm,
    AcademicYearAdmin,
    AssessmentGroupForSubjectAdminForm,
    AssessmentItemAdminForm,
    AssessmentItemInlineFormSet,
    AssessmentResultAdminForm,
    CourseRegistrationSettingsAdmin,
    FinalGradeRuleAdminForm,
    GradeAdmin,
    GradeAdminForm,
    GroupSubjectAdminForm,
    GroupSubjectForSubjectAdminForm,
    GroupSubjectForSubjectInline,
    JournalAdminDescriptionMixin,
    StudentAdmin,
    StudentAdminForm,
    StudentEnrollmentHistoryInline,
    StudentAssessmentGroupAdminForm,
    StudentAssessmentGroupInline,
    StudentInline,
    StudentSubjectAdminForm,
    StudyGroupAdmin,
    PasswordRecoveryContactAdmin,
    TemporaryCredentialAdmin,
    TeacherAdmin,
    TeacherEnrollmentHistoryInline,
    SubjectResultAdminForm,
    TeacherAdminForm,
)
from journal.middleware import ErrorLoggingMiddleware
from journal.error_logging import log_handled_error
from journal.user_error_messages import build_admin_form_user_message
from journal.forms import (
    CourseApplicationAdminForm,
    CourseApplicationPublicForm,
    CourseRegistrationSettingsForm,
    DetailedPasswordChangeForm,
    GradeCreateForm,
    SubjectResultForm,
    get_student_allowed_subjects,
    get_student_subject_teachers,
    get_teacher_groups,
    get_teacher_subjects,
)
from journal.grade_options import (
    get_grade_form_options,
    get_grade_groups,
    get_grade_students,
    get_grade_subjects,
    get_grade_teachers,
)
from journal.registration_utils import (
    latest_birth_date_for_age_in_year,
    minimum_birth_date_for_age,
    normalize_parent_contacts,
    reaches_age_in_calendar_year,
)
from journal.templatetags.admin_dashboard import journal_admin_dashboard
from journal.templatetags.journal_extras import short_person_name
from journal.views import (
    _build_journal_tables,
    _calculate_average,
    _is_duplicate_course_application_phone_error,
)
from journal.models import (
    AccountProfile,
    AcademicYear,
    AssessmentElement,
    AssessmentGroup,
    AssessmentItem,
    AssessmentResult,
    CourseApplication,
    CourseRegistrationSettings,
    ErrorLog,
    Grade,
    GroupSubject,
    FinalGradeRule,
    Instrument,
    OrchestraPart,
    PasswordRecoveryContact,
    Student,
    StudentAssessmentGroup,
    StudentEnrollment,
    StudentSubject,
    StudyGroup,
    Subject,
    SubjectResult,
    Teacher,
    TeacherEnrollment,
    TeacherSubject,
    TemporaryCredential,
    UserAcademicYearMembership,
)


User = get_user_model()


class JournalTestDataMixin:
    """Фабрики для тестов новой архитектуры журнала."""

    def create_academic_year(self, *, name='2025/2026', is_active=True):
        if not getattr(self, '_explicit_test_year_initialized', False):
            # Data migration 0009 creates a fallback year whose dates depend on
            # the day the test database is built.  Each test that constructs an
            # explicit chronology starts without that migration-only placeholder.
            AcademicYear.objects.all().delete()
            self._explicit_test_year_initialized = True
        start_year = int(name.split('/', 1)[0])
        return AcademicYear.objects.create(
            name=name,
            starts_on=date(start_year, 9, 1),
            ends_on=date(start_year + 1, 8, 31),
            is_active=is_active,
        )

    def create_group(self, *, name='Группа А', academic_year=None):
        academic_year = academic_year or self.create_academic_year()
        return StudyGroup.objects.create(name=name, academic_year=academic_year)

    def create_instrument(self, *, name='Баян'):
        return Instrument.objects.create(name=name)

    def create_subject(
        self,
        *,
        name='Сольфеджио',
        is_specialty=False,
        final_grade_type=None,
    ):
        return Subject.objects.create(
            name=name,
            is_specialty=is_specialty,
            final_grade_type=final_grade_type or Subject.FINAL_GRADE_TYPE_NUMERIC,
        )

    def create_teacher(
        self,
        *,
        full_name='Иванов Иван Иванович',
        username='teacher',
    ):
        user = User.objects.create_user(username=username, password='Pass12345!')
        return Teacher.objects.create(full_name=full_name, user=user)

    def create_student(
        self,
        *,
        full_name='Петров Пётр Петрович',
        group=None,
        instrument=None,
        username='student',
    ):
        group = group or self.create_group()
        instrument = instrument or self.create_instrument()
        user = User.objects.create_user(username=username, password='Pass12345!')
        return Student.objects.create(
            full_name=full_name,
            group=group,
            instrument=instrument,
            user=user,
        )

    def create_group_assignment(self, *, group=None, subject=None, teacher=None):
        group = group or self.create_group()
        subject = subject or self.create_subject()
        teacher = teacher or self.create_teacher()
        return GroupSubject.objects.create(
            group=group,
            subject=subject,
            teacher=teacher,
        )

    def create_individual_assignment(
        self,
        *,
        student=None,
        subject=None,
        teacher=None,
    ):
        student = student or self.create_student()
        subject = subject or self.create_subject(
            name='Специальность',
            is_specialty=True,
        )
        teacher = teacher or self.create_teacher(username='specialty_teacher')
        return StudentSubject.objects.create(
            student=student,
            subject=subject,
            teacher=teacher,
        )

    def create_base_journal(self):
        year = self.create_academic_year()
        group = self.create_group(academic_year=year)
        instrument = self.create_instrument(name='Баян')
        solfeggio = self.create_subject(name='Сольфеджио')
        literature = self.create_subject(name='Музыкальная литература')
        specialty = self.create_subject(name='Специальность', is_specialty=True)

        teacher = self.create_teacher(
            full_name='Иванов Иван Иванович',
            username='teacher_ivanov',
        )
        other_teacher = self.create_teacher(
            full_name='Петров Пётр Петрович',
            username='teacher_petrov',
        )
        student = self.create_student(
            full_name='Сидоров Семён Семёнович',
            group=group,
            instrument=instrument,
            username='student_sidorov',
        )

        GroupSubject.objects.create(
            group=group,
            subject=solfeggio,
            teacher=teacher,
        )
        GroupSubject.objects.create(
            group=group,
            subject=literature,
            teacher=other_teacher,
        )
        StudentSubject.objects.create(
            student=student,
            subject=specialty,
            teacher=other_teacher,
        )

        return {
            'year': year,
            'group': group,
            'instrument': instrument,
            'solfeggio': solfeggio,
            'literature': literature,
            'specialty': specialty,
            'teacher': teacher,
            'other_teacher': other_teacher,
            'student': student,
        }

    def application_payload(self, **overrides):
        if AcademicYear.get_active() is None:
            self.create_academic_year()
        instrument_name = overrides.pop('instrument', 'Баян I')
        custom_instrument = overrides.pop('custom_instrument', '')
        instrument_reference = overrides.pop('instrument_reference', None)
        if instrument_reference is None and not custom_instrument:
            instrument_reference, _created = Instrument.objects.get_or_create(
                name=instrument_name,
            )

        payload = {
            'last_name': 'Иванов',
            'first_name': 'Иван',
            'middle_name': 'Иванович',
            'gender': CourseApplication.GENDER_MALE,
            'birth_date': date(2000, 1, 1),
            'city_church': 'Тамбов',
            'instrument': instrument_name,
            'instrument_reference': instrument_reference,
            'custom_instrument': custom_instrument,
            'music_education': CourseApplication.MUSIC_EDUCATION_NONE,
            'student_phone': '+7 (999) 123-45-67',
            'parent_contacts': '',
            'comments': '',
        }
        payload.update(overrides)
        return payload

    def application_form_payload(self, **overrides):
        payload = self.application_payload(**overrides)
        birth_date = payload['birth_date']

        if hasattr(birth_date, 'isoformat'):
            payload['birth_date'] = birth_date.isoformat()
        reference = payload.get('instrument_reference')
        payload['instrument_reference'] = reference.pk if reference is not None else ''
        payload.pop('instrument', None)

        return payload


class AcademicStructureModelTests(JournalTestDataMixin, TestCase):
    def test_only_one_academic_year_can_be_active_after_save(self):
        first = self.create_academic_year(name='2025/2026', is_active=True)
        second = AcademicYear.objects.create(
            name='2026/2027',
            starts_on=date(2026, 9, 1),
            ends_on=date(2027, 8, 31),
            is_active=True,
        )

        first.refresh_from_db()
        second.refresh_from_db()

        self.assertFalse(first.is_active)
        self.assertTrue(second.is_active)

    def test_orchestra_parts_belong_to_one_instrument(self):
        domra = self.create_instrument(name='Домра')
        bayan = self.create_instrument(name='Баян')
        domra_part = OrchestraPart.objects.create(
            instrument=domra,
            name='Малая первая',
        )
        bayan_part = OrchestraPart.objects.create(
            instrument=bayan,
            name='Первый',
        )

        self.assertEqual(list(domra.orchestra_parts.all()), [domra_part])
        self.assertEqual(list(bayan.orchestra_parts.all()), [bayan_part])
        with self.assertRaises(ValidationError):
            OrchestraPart.objects.create(
                instrument=domra,
                name='Малая первая',
            )

    def test_student_rejects_part_from_another_instrument(self):
        year = self.create_academic_year()
        group = self.create_group(academic_year=year)
        domra = self.create_instrument(name='Домра')
        bayan = self.create_instrument(name='Баян')
        wrong_part = OrchestraPart.objects.create(
            instrument=bayan,
            name='Первый',
        )
        student = self.create_student(
            group=group,
            instrument=domra,
            username='orchestra_mismatch_student',
        )
        student.orchestra_part = wrong_part

        with self.assertRaises(ValidationError) as error:
            student.save()

        self.assertIn('orchestra_part', error.exception.message_dict)

    def test_custom_instrument_clears_orchestra_part(self):
        instrument = self.create_instrument(name='Домра')
        part = OrchestraPart.objects.create(
            instrument=instrument,
            name='Малая вторая',
        )
        student = self.create_student(
            instrument=instrument,
            username='custom_instrument_part_student',
        )
        student.orchestra_part = part
        student.instrument = None
        student.custom_instrument = 'Скрипка'

        student.save()
        student.refresh_from_db()

        self.assertIsNone(student.orchestra_part)

    def test_database_rejects_part_without_reference_instrument(self):
        instrument = self.create_instrument(name='Домра')
        part = OrchestraPart.objects.create(
            instrument=instrument,
            name='Альтовая первая',
        )
        student = self.create_student(
            instrument=instrument,
            username='orchestra_constraint_student',
        )

        with transaction.atomic():
            with self.assertRaises(IntegrityError):
                Student.objects.filter(pk=student.pk).update(
                    instrument=None,
                    custom_instrument='Собственный инструмент',
                    orchestra_part=part,
                )

    def test_newest_academic_year_becomes_active_even_when_created_inactive(self):
        first = self.create_academic_year(name='2025/2026', is_active=True)
        second = AcademicYear.objects.create(
            name='2026/2027',
            starts_on=date(2026, 9, 1),
            ends_on=date(2027, 8, 31),
            is_active=False,
        )

        first.refresh_from_db()
        second.refresh_from_db()

        self.assertFalse(first.is_active)
        self.assertTrue(second.is_active)

    def test_new_year_preserves_old_enrollment_and_grade_snapshots(self):
        data = self.create_base_journal()
        old_name = data['student'].full_name
        grade = Grade.objects.create(
            student=data['student'],
            subject=data['solfeggio'],
            teacher=data['teacher'],
            academic_year=data['year'],
            date=date(2025, 10, 10),
            value='5',
        )

        new_year = self.create_academic_year(name='2026/2027')
        data['student'].refresh_from_db()
        data['year'].refresh_from_db()
        grade.refresh_from_db()

        self.assertFalse(data['year'].is_active)
        self.assertTrue(new_year.is_active)
        self.assertIsNone(data['student'].group_id)
        self.assertFalse(data['student'].is_active)
        self.assertEqual(grade.enrollment.group_id, data['group'].pk)
        self.assertEqual(grade.student_name_snapshot, old_name)

        data['student'].full_name = 'Новое имя ученика'
        data['student'].save()
        old_enrollment = StudentEnrollment.objects.get(
            student=data['student'],
            academic_year=data['year'],
        )
        self.assertEqual(old_enrollment.full_name, old_name)

    def test_new_year_resets_group_for_every_student_from_previous_year(self):
        year = self.create_academic_year()
        first_group = self.create_group(name='Первая группа', academic_year=year)
        second_group = self.create_group(name='Вторая группа', academic_year=year)
        instrument = self.create_instrument()
        first = self.create_student(
            full_name='Первый Ученик',
            group=first_group,
            instrument=instrument,
            username='first_student',
        )
        second = self.create_student(
            full_name='Второй Ученик',
            group=second_group,
            instrument=instrument,
            username='second_student',
        )

        self.create_academic_year(name='2026/2027')
        first.refresh_from_db()
        second.refresh_from_db()

        self.assertIsNone(first.group_id)
        self.assertIsNone(second.group_id)
        self.assertFalse(first.is_active)
        self.assertFalse(second.is_active)

    def test_teacher_membership_is_scoped_to_academic_year(self):
        old_year = self.create_academic_year()
        teacher = self.create_teacher()

        self.assertTrue(
            TeacherEnrollment.objects.filter(
                teacher=teacher,
                academic_year=old_year,
                is_active=True,
            ).exists(),
        )

        new_year = self.create_academic_year(name='2026/2027')
        teacher.refresh_from_db()
        self.assertFalse(teacher.is_active)
        self.assertFalse(
            TeacherEnrollment.objects.filter(
                teacher=teacher,
                academic_year=new_year,
            ).exists(),
        )

        new_group = self.create_group(name='Новая группа', academic_year=new_year)
        subject = self.create_subject()
        GroupSubject.objects.create(group=new_group, subject=subject, teacher=teacher)

        teacher.refresh_from_db()
        self.assertTrue(teacher.is_active)
        self.assertTrue(
            TeacherEnrollment.objects.filter(
                teacher=teacher,
                academic_year=new_year,
                is_active=True,
            ).exists(),
        )

    def test_new_year_finalizes_current_names_before_archiving(self):
        data = self.create_base_journal()
        grade = Grade.objects.create(
            student=data['student'],
            subject=data['solfeggio'],
            teacher=data['teacher'],
            academic_year=data['year'],
            date=date(2025, 10, 10),
            value='5',
        )
        data['student'].full_name = 'Итоговое имя ученика'
        data['student'].save()
        data['solfeggio'].name = 'Итоговое название предмета'
        data['solfeggio'].save()
        data['teacher'].full_name = 'Итоговое имя преподавателя'
        data['teacher'].save()

        self.create_academic_year(name='2026/2027')
        grade.refresh_from_db()
        assignment = GroupSubject.objects.get(
            group=data['group'],
            subject=data['solfeggio'],
        )
        enrollment = StudentEnrollment.objects.get(
            student=data['student'],
            academic_year=data['year'],
        )

        self.assertEqual(enrollment.full_name, 'Итоговое имя ученика')
        self.assertEqual(assignment.subject_name_snapshot, 'Итоговое название предмета')
        self.assertEqual(assignment.teacher_name_snapshot, 'Итоговое имя преподавателя')
        self.assertEqual(grade.student_name_snapshot, 'Итоговое имя ученика')
        self.assertEqual(grade.subject_name_snapshot, 'Итоговое название предмета')
        self.assertEqual(grade.teacher_name_snapshot, 'Итоговое имя преподавателя')

    def test_reordering_active_year_finalizes_it_and_restores_latest_enrollment(self):
        data = self.create_base_journal()
        current_year = self.create_academic_year(name='2027/2028')
        current_group = self.create_group(name='Группа Б', academic_year=current_year)
        GroupSubject.objects.create(
            group=current_group,
            subject=data['solfeggio'],
            teacher=data['teacher'],
        )

        data['student'].group = current_group
        data['student'].full_name = 'Финальное имя второго года'
        data['student'].save()
        current_grade = Grade.objects.create(
            student=data['student'],
            subject=data['solfeggio'],
            teacher=data['teacher'],
            academic_year=current_year,
            date=date(2027, 10, 10),
            value='5',
        )

        current_year.name = '2024/2025'
        current_year.starts_on = date(2024, 9, 1)
        current_year.ends_on = date(2025, 8, 31)
        current_year.save()

        data['year'].refresh_from_db()
        current_year.refresh_from_db()
        data['student'].refresh_from_db()
        current_grade.refresh_from_db()
        current_enrollment = StudentEnrollment.objects.get(
            student=data['student'],
            academic_year=current_year,
        )

        self.assertTrue(data['year'].is_active)
        self.assertFalse(current_year.is_active)
        self.assertEqual(data['student'].group_id, data['group'].pk)
        self.assertEqual(current_enrollment.full_name, 'Финальное имя второго года')
        self.assertEqual(current_grade.student_name_snapshot, 'Финальное имя второго года')

    def test_deleting_empty_active_year_restores_previous_student_groups(self):
        data = self.create_base_journal()
        empty_year = self.create_academic_year(name='2026/2027')
        data['student'].refresh_from_db()
        self.assertIsNone(data['student'].group_id)

        empty_year.delete()

        data['year'].refresh_from_db()
        data['student'].refresh_from_db()
        self.assertTrue(data['year'].is_active)
        self.assertEqual(data['student'].group_id, data['group'].pk)
        self.assertTrue(data['student'].is_active)

    def test_archived_grade_cannot_be_changed_but_can_be_deleted(self):
        data = self.create_base_journal()
        grade = Grade.objects.create(
            student=data['student'],
            subject=data['solfeggio'],
            teacher=data['teacher'],
            academic_year=data['year'],
            date=date(2025, 10, 10),
            value='5',
        )
        self.create_academic_year(name='2026/2027')

        grade.value = '4'
        with self.assertRaisesMessage(ValidationError, 'Архивный учебный год'):
            grade.save()
        grade.delete()

        self.assertFalse(Grade.objects.filter(pk=grade.pk).exists())

    def test_academic_year_periods_cannot_overlap(self):
        self.create_academic_year()

        with self.assertRaisesMessage(ValidationError, 'пересекается'):
            AcademicYear.objects.create(
                name='Пересекающийся',
                starts_on=date(2026, 8, 1),
                ends_on=date(2027, 7, 31),
            )

    def test_cannot_save_group_in_archived_academic_year(self):
        group = self.create_group()
        AcademicYear.objects.create(
            name='2026/2027',
            starts_on=date(2026, 9, 1),
            ends_on=date(2027, 8, 31),
        )

        group.name = 'Переименованная группа'

        with self.assertRaisesMessage(ValidationError, 'Архивный учебный год'):
            group.save()

    def test_grade_date_must_be_inside_selected_academic_year(self):
        data = self.create_base_journal()

        with self.assertRaisesMessage(ValidationError, 'Дата оценки должна попадать в период'):
            Grade.objects.create(
                student=data['student'],
                subject=data['solfeggio'],
                teacher=data['teacher'],
                academic_year=data['year'],
                date=date(2024, 10, 1),
                value='5',
            )

    def test_grade_cannot_be_created_in_archived_academic_year(self):
        data = self.create_base_journal()
        AcademicYear.objects.create(
            name='2026/2027',
            starts_on=date(2026, 9, 1),
            ends_on=date(2027, 8, 31),
        )

        with self.assertRaisesMessage(ValidationError, 'Архивный учебный год'):
            Grade.objects.create(
                student=data['student'],
                subject=data['solfeggio'],
                teacher=data['teacher'],
                academic_year=data['year'],
                date=date(2025, 10, 1),
                value='5',
            )

    def test_subject_final_grade_type_cannot_break_existing_results(self):
        data = self.create_base_journal()
        pass_fail_subject = self.create_subject(
            name='Зачетный предмет',
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )
        GroupSubject.objects.create(
            group=data['group'],
            subject=pass_fail_subject,
            teacher=data['teacher'],
        )
        SubjectResult.objects.create(
            student=data['student'],
            subject=pass_fail_subject,
            academic_year=data['year'],
            final_grade='Зачет',
        )

        pass_fail_subject.final_grade_type = Subject.FINAL_GRADE_TYPE_NUMERIC

        with self.assertRaisesMessage(ValidationError, 'Нельзя изменить тип итоговой оценки'):
            pass_fail_subject.save()

    def test_group_subject_links_group_subject_and_teacher(self):
        data = self.create_base_journal()

        assignment = GroupSubject.objects.get(
            group=data['group'],
            subject=data['solfeggio'],
        )

        self.assertEqual(assignment.teacher, data['teacher'])
        self.assertIn(data['solfeggio'], data['group'].subjects.all())
        self.assertEqual(data['teacher'].group_subjects.count(), 1)

    def test_student_group_is_optional_and_cleared_when_group_is_deleted(self):
        group = self.create_group()
        student = self.create_student(group=group)

        group.delete()
        student.refresh_from_db()

        self.assertIsNone(student.group)

    def test_deleting_active_enrollment_clears_current_student_group(self):
        group = self.create_group()
        student = self.create_student(group=group)
        enrollment = student.enrollment_for_year(group.academic_year)

        enrollment.delete()
        student.refresh_from_db()

        self.assertIsNone(student.group)
        self.assertFalse(
            StudentEnrollment.objects.filter(
                student=student,
                academic_year=group.academic_year,
            ).exists(),
        )

    def test_only_academic_year_can_be_deleted(self):
        academic_year = self.create_academic_year()

        academic_year.delete()

        self.assertFalse(AcademicYear.objects.filter(pk=academic_year.pk).exists())

    def test_archived_academic_year_can_be_deleted(self):
        archived_year = self.create_academic_year()
        self.create_academic_year(name='2026/2027')

        archived_year.delete()

        self.assertFalse(AcademicYear.objects.filter(pk=archived_year.pk).exists())

    def test_student_without_group_keeps_individual_subjects_display(self):
        data = self.create_base_journal()
        data['student'].group = None
        data['student'].save(update_fields=['group'])

        self.assertIn('Специальность', data['student'].subjects_display)

    def test_group_subject_rejects_specialty_subject(self):
        group = self.create_group()
        specialty = self.create_subject(name='Специальность', is_specialty=True)
        teacher = self.create_teacher()

        with self.assertRaises(ValidationError):
            GroupSubject.objects.create(
                group=group,
                subject=specialty,
                teacher=teacher,
            )

    def test_subject_specialty_flag_is_labeled_as_individual_subject(self):
        self.assertEqual(
            Subject._meta.get_field('is_specialty').verbose_name,
            'Индивидуальный предмет',
        )

    def test_student_subject_accepts_specialty_subject(self):
        data = self.create_base_journal()
        student = data['student']

        self.assertEqual(student.specialty_subject, data['specialty'])
        self.assertEqual(student.specialty_teacher, data['other_teacher'])
        self.assertIn('Специальность', student.subjects_display)

    def test_student_subject_rejects_group_subject(self):
        data = self.create_base_journal()

        with self.assertRaises(ValidationError):
            StudentSubject.objects.create(
                student=data['student'],
                subject=data['solfeggio'],
                teacher=data['teacher'],
            )

    def test_subject_cannot_be_switched_to_individual_with_group_assignments(self):
        data = self.create_base_journal()
        subject = data['solfeggio']
        subject.is_specialty = True

        with self.assertRaises(ValidationError):
            subject.save()

    def test_subject_cannot_be_switched_to_group_with_individual_assignments(self):
        data = self.create_base_journal()
        subject = data['specialty']
        subject.is_specialty = False

        with self.assertRaises(ValidationError):
            subject.save()

    def test_student_can_have_multiple_individual_subjects(self):
        data = self.create_base_journal()
        another_subject = self.create_subject(
            name='Индивидуальная импровизация',
            is_specialty=True,
        )

        assignment = StudentSubject.objects.create(
            student=data['student'],
            subject=another_subject,
            teacher=data['teacher'],
            is_active=True,
        )

        self.assertEqual(assignment.subject, another_subject)
        self.assertEqual(
            StudentSubject.objects.filter(student=data['student'], is_active=True).count(),
            2,
        )

    def test_teacher_subject_stores_qualification_not_assignment(self):
        subject = self.create_subject()
        teacher = self.create_teacher()

        TeacherSubject.objects.create(teacher=teacher, subject=subject)

        self.assertIn(subject, teacher.qualified_subjects.all())
        self.assertEqual(teacher.group_subjects.count(), 0)

    def test_group_subject_teacher_change_syncs_teacher_subjects_and_grades(self):
        data = self.create_base_journal()
        assignment = GroupSubject.objects.get(
            group=data['group'],
            subject=data['solfeggio'],
        )
        grade = Grade.objects.create(
            student=data['student'],
            subject=data['solfeggio'],
            teacher=data['teacher'],
            date=date(2025, 10, 8),
            value='5',
        )

        assignment.teacher = data['other_teacher']
        assignment.save()

        grade.refresh_from_db()
        self.assertEqual(grade.teacher, data['other_teacher'])
        self.assertTrue(
            TeacherSubject.objects.filter(
                teacher=data['other_teacher'],
                subject=data['solfeggio'],
            ).exists(),
        )
        self.assertFalse(
            TeacherSubject.objects.filter(
                teacher=data['teacher'],
                subject=data['solfeggio'],
            ).exists(),
        )

    def test_individual_subject_teacher_change_syncs_teacher_subjects_and_grades(self):
        data = self.create_base_journal()
        assignment = StudentSubject.objects.get(
            student=data['student'],
            subject=data['specialty'],
        )
        grade = Grade.objects.create(
            student=data['student'],
            subject=data['specialty'],
            teacher=data['other_teacher'],
            date=date(2025, 10, 9),
            value='4',
        )

        assignment.teacher = data['teacher']
        assignment.save()

        grade.refresh_from_db()
        self.assertEqual(grade.teacher, data['teacher'])
        self.assertTrue(
            TeacherSubject.objects.filter(
                teacher=data['teacher'],
                subject=data['specialty'],
            ).exists(),
        )
        self.assertFalse(
            TeacherSubject.objects.filter(
                teacher=data['other_teacher'],
                subject=data['specialty'],
            ).exists(),
        )


class GradeModelTests(JournalTestDataMixin, TestCase):
    def test_group_subject_teacher_can_create_grade(self):
        data = self.create_base_journal()

        grade = Grade.objects.create(
            student=data['student'],
            subject=data['solfeggio'],
            teacher=data['teacher'],
            date=date(2025, 10, 1),
            value='5',
        )

        self.assertEqual(grade.academic_year, data['year'])
        self.assertTrue(grade.is_group_subject)
        self.assertFalse(grade.is_individual_subject)

    def test_individual_subject_teacher_can_create_grade(self):
        data = self.create_base_journal()

        grade = Grade.objects.create(
            student=data['student'],
            subject=data['specialty'],
            teacher=data['other_teacher'],
            date=date(2025, 10, 2),
            value='4',
        )

        self.assertTrue(grade.is_individual_subject)
        self.assertFalse(grade.is_group_subject)

    def test_individual_subject_grade_does_not_require_student_group(self):
        data = self.create_base_journal()
        data['student'].group = None
        data['student'].save(update_fields=['group'])

        grade = Grade.objects.create(
            student=data['student'],
            subject=data['specialty'],
            teacher=data['other_teacher'],
            academic_year=data['year'],
            date=date(2025, 10, 3),
            value='5',
        )

        self.assertIsNone(grade.enrollment.group)
        self.assertTrue(grade.is_individual_subject)
        self.assertEqual(
            list(get_grade_subjects(student=data['student'], academic_year=data['year'])),
            [data['specialty']],
        )

    def test_unassigned_teacher_cannot_create_grade(self):
        data = self.create_base_journal()

        with self.assertRaises(ValidationError):
            Grade.objects.create(
                student=data['student'],
                subject=data['solfeggio'],
                teacher=data['other_teacher'],
                date=date(2025, 10, 3),
                value='5',
            )

    def test_student_cannot_receive_grade_for_unassigned_subject(self):
        data = self.create_base_journal()
        unassigned_subject = self.create_subject(name='Хор')

        with self.assertRaises(ValidationError):
            Grade.objects.create(
                student=data['student'],
                subject=unassigned_subject,
                teacher=data['teacher'],
                date=date(2025, 10, 4),
                value='5',
            )

    def test_duplicate_grade_for_same_student_subject_and_date_is_rejected(self):
        data = self.create_base_journal()

        Grade.objects.create(
            student=data['student'],
            subject=data['solfeggio'],
            teacher=data['teacher'],
            date=date(2025, 10, 5),
            value='4',
        )

        with self.assertRaises(ValidationError):
            Grade.objects.create(
                student=data['student'],
                subject=data['solfeggio'],
                teacher=data['teacher'],
                date=date(2025, 10, 5),
                value='5',
            )

    def test_grade_value_is_normalized_and_limited(self):
        data = self.create_base_journal()

        grade = Grade.objects.create(
            student=data['student'],
            subject=data['solfeggio'],
            teacher=data['teacher'],
            date=date(2025, 10, 6),
            value='н',
        )

        self.assertEqual(grade.value, 'Н')

        with self.assertRaises(ValidationError):
            Grade.objects.create(
                student=data['student'],
                subject=data['solfeggio'],
                teacher=data['teacher'],
                date=date(2025, 10, 7),
                value='6',
            )

    def test_grade_accepts_plus_and_minus_for_every_numeric_mark(self):
        self.assertEqual(
            Grade.ALLOWED_VALUES,
            {
                '1', '1+', '1-',
                '2', '2+', '2-',
                '3', '3+', '3-',
                '4', '4+', '4-',
                '5', '5+', '5-',
                'Н',
            },
        )

    def test_average_applies_modifiers_and_ignores_absence(self):
        self.assertEqual(
            _calculate_average(['4+', '5-', 'Н', 'n']),
            '4.50',
        )
        self.assertEqual(_calculate_average(['Н']), '')


class SubjectResultModelTests(JournalTestDataMixin, TestCase):
    def test_subject_result_is_unique_for_student_subject_and_year(self):
        data = self.create_base_journal()

        SubjectResult.objects.create(
            student=data['student'],
            subject=data['solfeggio'],
            academic_year=data['year'],
            exam_grade='5',
            final_grade='5',
        )

        with self.assertRaises(ValidationError):
            SubjectResult.objects.create(
                student=data['student'],
                subject=data['solfeggio'],
                academic_year=data['year'],
                exam_grade='4',
                final_grade='4',
            )

    def test_subject_result_rejects_unassigned_subject(self):
        data = self.create_base_journal()
        unassigned_subject = self.create_subject(name='Хор')

        with self.assertRaises(ValidationError):
            SubjectResult.objects.create(
                student=data['student'],
                subject=unassigned_subject,
                academic_year=data['year'],
                final_grade='5',
            )

    def test_pass_fail_result_accepts_only_pass_fail_values(self):
        data = self.create_base_journal()
        pass_fail_subject = self.create_subject(
            name='Зачетный предмет',
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )

        GroupSubject.objects.create(
            group=data['group'],
            subject=pass_fail_subject,
            teacher=data['teacher'],
        )

        result = SubjectResult.objects.create(
            student=data['student'],
            subject=pass_fail_subject,
            academic_year=data['year'],
            exam_grade='зачет',
            final_grade='незачет',
        )

        self.assertEqual(result.exam_grade, 'Зачет')
        self.assertEqual(result.final_grade, 'Незачет')

        with self.assertRaises(ValidationError):
            SubjectResult.objects.create(
                student=data['student'],
                subject=pass_fail_subject,
                academic_year=data['year'],
                exam_grade='5',
                final_grade='5',
            )

    def test_numeric_exam_and_final_accept_modifiers_and_absence(self):
        data = self.create_base_journal()

        result = SubjectResult.objects.create(
            student=data['student'],
            subject=data['solfeggio'],
            academic_year=data['year'],
            exam_grade='4+',
            final_grade='n',
        )

        self.assertEqual(result.exam_grade, '4+')
        self.assertEqual(result.final_grade, 'Н')

        result.exam_grade = '5-'
        result.final_grade = '3+'
        result.save()
        result.refresh_from_db()

        self.assertEqual(result.exam_grade, '5-')
        self.assertEqual(result.final_grade, '3+')


class CourseApplicationLifecycleTests(JournalTestDataMixin, TestCase):
    def test_default_status_is_confirmed_and_creates_journal_records(self):
        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='Temp12345!',
        ):
            application = CourseApplication.objects.create(
                **self.application_payload(),
            )

        application.refresh_from_db()
        credential = TemporaryCredential.objects.get(
            course_application=application,
        )
        student = Student.objects.get(pk=application.student_id)
        user = User.objects.get(pk=application.user_id)

        self.assertEqual(application.status, CourseApplication.STATUS_CONFIRMED)
        self.assertEqual(application.generated_login, 'Иванов Иван')
        self.assertEqual(credential.login, 'Иванов Иван')
        self.assertEqual(credential.temporary_password, 'Temp12345!')
        self.assertEqual(credential.student_phone, '+7 (999) 123-45-67')
        self.assertEqual(credential.user, user)
        self.assertTrue(user.check_password('Temp12345!'))
        self.assertEqual(student.full_name, 'Иванов Иван Иванович')
        self.assertEqual(student.gender, application.gender)
        self.assertEqual(student.birth_date, application.birth_date)
        self.assertEqual(student.city_church, application.city_church)
        self.assertEqual(student.music_education, application.music_education)
        self.assertEqual(student.student_phone, application.student_phone)
        self.assertEqual(student.parent_contacts, application.parent_contacts)
        self.assertEqual(student.comments, application.comments)
        self.assertEqual(student.user, user)
        self.assertEqual(student.group.name, CourseApplication.STUDENT_COURSE_GROUP_NAME)
        self.assertEqual(student.instrument.name, 'Баян I')
        self.assertIsNotNone(application.journal_created_at)
        self.assertIsNone(application.journal_removed_at)

    def test_confirmed_application_updates_existing_student_profile_details(self):
        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='Temp12345!',
        ):
            application = CourseApplication.objects.create(
                **self.application_payload(),
            )

        application.birth_date = date(1999, 5, 4)
        application.city_church = 'Воронеж / Отрожка'
        application.instrument_reference = self.create_instrument(name='Фортепиано')
        application.custom_instrument = ''
        application.music_education = CourseApplication.MUSIC_EDUCATION_HIGHER
        application.student_phone = '+7 (999) 123-45-69'
        application.parent_contacts = 'Отец - +7 (999) 111-22-33'
        application.comments = 'Нужен вечерний поток'
        application.save()

        student = application.student
        student.refresh_from_db()
        credential = TemporaryCredential.objects.get(course_application=application)

        self.assertEqual(student.birth_date, date(1999, 5, 4))
        self.assertEqual(student.city_church, 'Воронеж / Отрожка')
        self.assertEqual(student.instrument.name, 'Фортепиано')
        self.assertEqual(student.music_education, CourseApplication.MUSIC_EDUCATION_HIGHER)
        self.assertEqual(student.student_phone, '+7 (999) 123-45-69')
        self.assertEqual(student.parent_contacts, 'Отец - +7 (999) 111-22-33')
        self.assertEqual(student.comments, 'Нужен вечерний поток')
        self.assertEqual(credential.student_phone, '+7 (999) 123-45-69')

    def test_confirmed_application_update_preserves_password_and_temporary_password(self):
        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='Temp12345!',
        ):
            application = CourseApplication.objects.create(
                **self.application_payload(),
            )

        user = application.user
        credential = TemporaryCredential.objects.get(course_application=application)
        original_password_hash = user.password
        original_temporary_password = credential.temporary_password

        application.first_name = 'Пётр'
        application.comments = 'Данные заявки изменены'
        application.save()

        user.refresh_from_db()
        credential.refresh_from_db()
        self.assertEqual(user.password, original_password_hash)
        self.assertEqual(credential.temporary_password, original_temporary_password)
        self.assertTrue(user.check_password(original_temporary_password))

    def test_same_person_in_later_year_reuses_student_and_user(self):
        first_year = self.create_academic_year(name='2025/2026')
        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='Temp12345!',
        ):
            first_application = CourseApplication.objects.create(
                **self.application_payload(),
            )

        original_student_id = first_application.student_id
        original_user_id = first_application.user_id
        original_password_hash = first_application.user.password

        second_year = self.create_academic_year(name='2026/2027')
        second_application = CourseApplication.objects.create(
            **self.application_payload(
                student_phone='+7 (999) 765-43-21',
                city_church='Новый город / Новая церковь',
            ),
        )

        second_application.refresh_from_db()
        second_application.user.refresh_from_db()
        self.assertEqual(second_application.student_id, original_student_id)
        self.assertEqual(second_application.user_id, original_user_id)
        self.assertEqual(second_application.user.password, original_password_hash)
        self.assertEqual(Student.objects.count(), 1)
        self.assertEqual(User.objects.filter(pk=original_user_id).count(), 1)
        self.assertEqual(TemporaryCredential.objects.count(), 1)
        self.assertTrue(
            StudentEnrollment.objects.filter(
                student_id=original_student_id,
                academic_year=first_year,
            ).exists(),
        )
        self.assertTrue(
            StudentEnrollment.objects.filter(
                student_id=original_student_id,
                academic_year=second_year,
            ).exists(),
        )

    def test_different_birth_date_creates_different_student_even_with_same_name(self):
        self.create_academic_year(name='2025/2026')
        first_application = CourseApplication.objects.create(**self.application_payload())
        self.create_academic_year(name='2026/2027')
        second_application = CourseApplication.objects.create(
            **self.application_payload(
                birth_date=date(2001, 1, 1),
                student_phone='+7 (999) 765-43-21',
            ),
        )

        self.assertNotEqual(first_application.student_id, second_application.student_id)
        self.assertEqual(Student.objects.count(), 2)

    def test_confirmed_applications_add_suffix_for_duplicate_login(self):
        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='Temp12345!',
        ):
            first = CourseApplication.objects.create(
                **self.application_payload(),
            )
            second = CourseApplication.objects.create(
                **self.application_payload(
                    birth_date=date(2001, 1, 1),
                    student_phone='+7 (999) 123-45-68',
                ),
            )

        self.assertEqual(first.generated_login, 'Иванов Иван')
        self.assertEqual(second.generated_login, 'Иванов Иван 2')
        self.assertEqual(
            list(
                TemporaryCredential.objects.order_by('id').values_list(
                    'login',
                    flat=True,
                ),
            ),
            ['Иванов Иван', 'Иванов Иван 2'],
        )

    def test_rejecting_one_of_two_confirmed_applications_preserves_shared_account(self):
        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='Temp12345!',
        ):
            first = CourseApplication.objects.create(**self.application_payload())
        second = CourseApplication.objects.create(
            **self.application_payload(student_phone='+7 (999) 765-43-21'),
        )
        student_id = first.student_id
        user_id = first.user_id
        enrollment_id = StudentEnrollment.objects.get(
            student_id=student_id,
            academic_year=first.academic_year,
        ).pk

        first.status = CourseApplication.STATUS_REJECTED
        first.save()
        first.refresh_from_db()
        second.refresh_from_db()
        credential = TemporaryCredential.objects.get(user_id=user_id)

        self.assertIsNone(first.student_id)
        self.assertIsNone(first.user_id)
        self.assertEqual(second.student_id, student_id)
        self.assertEqual(second.user_id, user_id)
        self.assertTrue(Student.objects.filter(pk=student_id).exists())
        self.assertTrue(User.objects.filter(pk=user_id).exists())
        self.assertTrue(StudentEnrollment.objects.filter(pk=enrollment_id).exists())
        self.assertEqual(credential.course_application, second)
        self.assertEqual(credential.temporary_password, 'Temp12345!')

    def test_deleting_one_of_two_confirmed_applications_preserves_shared_account(self):
        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='Temp12345!',
        ):
            first = CourseApplication.objects.create(**self.application_payload())
        second = CourseApplication.objects.create(
            **self.application_payload(student_phone='+7 (999) 765-43-21'),
        )
        student_id = first.student_id
        user_id = first.user_id

        first.delete()
        second.refresh_from_db()
        credential = TemporaryCredential.objects.get(user_id=user_id)

        self.assertEqual(second.student_id, student_id)
        self.assertEqual(second.user_id, user_id)
        self.assertTrue(Student.objects.filter(pk=student_id).exists())
        self.assertTrue(User.objects.filter(pk=user_id).exists())
        self.assertTrue(
            StudentEnrollment.objects.filter(
                student_id=student_id,
                academic_year=second.academic_year,
            ).exists(),
        )
        self.assertEqual(credential.course_application, second)

    def test_rejected_application_does_not_create_journal_records(self):
        application = CourseApplication.objects.create(
            **self.application_payload(status=CourseApplication.STATUS_REJECTED),
        )

        application.refresh_from_db()

        self.assertEqual(application.status, CourseApplication.STATUS_REJECTED)
        self.assertIsNone(application.student)
        self.assertIsNone(application.user)
        self.assertEqual(Student.objects.count(), 0)
        self.assertEqual(TemporaryCredential.objects.count(), 0)

    def test_changing_status_to_rejected_removes_student_user_and_temporary_credentials(
        self,
    ):
        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='Temp12345!',
        ):
            application = CourseApplication.objects.create(
                **self.application_payload(),
            )

        login = application.generated_login
        student_id = application.student_id
        user_id = application.user_id

        self.assertTrue(Student.objects.filter(pk=student_id).exists())
        self.assertTrue(User.objects.filter(pk=user_id).exists())
        self.assertTrue(TemporaryCredential.objects.filter(login=login).exists())

        application.status = CourseApplication.STATUS_REJECTED
        application.save()
        application.refresh_from_db()

        self.assertEqual(CourseApplication.objects.count(), 1)
        self.assertEqual(application.status, CourseApplication.STATUS_REJECTED)
        self.assertEqual(application.generated_login, login)
        self.assertIsNone(application.student)
        self.assertIsNone(application.user)
        self.assertIsNotNone(application.journal_removed_at)
        self.assertFalse(Student.objects.filter(pk=student_id).exists())
        self.assertFalse(User.objects.filter(pk=user_id).exists())
        self.assertFalse(TemporaryCredential.objects.filter(login=login).exists())

    def test_changing_rejected_application_back_to_confirmed_recreates_records(self):
        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='Temp12345!',
        ):
            application = CourseApplication.objects.create(
                **self.application_payload(),
            )

        application.status = CourseApplication.STATUS_REJECTED
        application.save()
        application.refresh_from_db()

        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='NewTemp12345!',
        ):
            application.status = CourseApplication.STATUS_CONFIRMED
            application.save()

        application.refresh_from_db()
        credential = TemporaryCredential.objects.get(
            course_application=application,
        )

        self.assertEqual(application.status, CourseApplication.STATUS_CONFIRMED)
        self.assertIsNotNone(application.student)
        self.assertIsNotNone(application.user)
        self.assertEqual(application.generated_login, 'Иванов Иван')
        self.assertEqual(credential.login, 'Иванов Иван')
        self.assertEqual(credential.temporary_password, 'NewTemp12345!')
        self.assertTrue(application.user.check_password('NewTemp12345!'))
        self.assertIsNone(application.journal_removed_at)

    def test_deleting_application_removes_created_journal_records(self):
        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='Temp12345!',
        ):
            application = CourseApplication.objects.create(
                **self.application_payload(),
            )

        user_id = application.user_id
        student_id = application.student_id

        application.delete()

        self.assertEqual(CourseApplication.objects.count(), 0)
        self.assertFalse(Student.objects.filter(pk=student_id).exists())
        self.assertFalse(User.objects.filter(pk=user_id).exists())
        self.assertEqual(TemporaryCredential.objects.count(), 0)

    def test_duplicate_student_phone_is_rejected(self):
        CourseApplication.objects.create(**self.application_payload())

        with self.assertRaises(ValidationError):
            CourseApplication.objects.create(
                **self.application_payload(
                    last_name='Петров',
                    first_name='Пётр',
                    middle_name='Петрович',
                    student_phone='8 999 123 45 67',
                ),
            )

    def test_editing_confirmed_application_does_not_reset_existing_user_password(self):
        application = CourseApplication.objects.create(**self.application_payload())
        user = application.user
        original_password_hash = user.password
        TemporaryCredential.objects.filter(course_application=application).delete()

        application.comments = 'Обновленный комментарий'
        application.save()
        user.refresh_from_db()

        self.assertEqual(user.password, original_password_hash)
        self.assertFalse(
            TemporaryCredential.objects.filter(course_application=application).exists(),
        )


class FormTests(JournalTestDataMixin, TestCase):
    def test_public_course_application_form_hides_status_and_normalizes_phone(self):
        form = CourseApplicationPublicForm(
            data=self.application_form_payload(
                student_phone='8 999 123 45 67',
            ),
        )

        self.assertNotIn('status', form.fields)
        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['student_phone'], '+7 (999) 123-45-67')

    def test_admin_course_application_form_includes_status(self):
        form = CourseApplicationAdminForm(
            data=self.application_form_payload(
                status=CourseApplication.STATUS_REJECTED,
            ),
        )

        self.assertIn('status', form.fields)
        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['status'], CourseApplication.STATUS_REJECTED)

    def test_course_application_edit_form_keeps_existing_field_values(self):
        application = CourseApplication(
            **self.application_payload(
                birth_date=date(2000, 1, 2),
                city_church='Воронеж, Отрожка',
                instrument='Фортепиано',
                music_education=CourseApplication.MUSIC_EDUCATION_BASIC,
                comments='Нужен вечерний поток',
            ),
        )

        form = CourseApplicationAdminForm(instance=application)

        self.assertIn('value="Иванов"', str(form['last_name']))
        self.assertIn('value="2000-01-02"', str(form['birth_date']))
        self.assertIn('Воронеж, Отрожка', str(form['city_church']))
        self.assertIn('Фортепиано', str(form['instrument_reference']))
        self.assertIn('value="basic" selected', str(form['music_education']))
        self.assertIn('Нужен вечерний поток', str(form['comments']))

    def test_course_application_form_supports_directory_and_custom_instruments(self):
        bayan = self.create_instrument(name='Баян')

        directory_form = CourseApplicationPublicForm(
            data=self.application_form_payload(
                instrument='Баян',
                instrument_reference=bayan,
            ),
        )
        custom_form = CourseApplicationPublicForm(
            data=self.application_form_payload(
                instrument='Домра малая II',
                instrument_reference=None,
                custom_instrument='Домра малая II',
            ),
        )

        self.assertIn('<select', str(directory_form['instrument_reference']))
        self.assertTrue(directory_form.is_valid(), directory_form.errors)
        self.assertTrue(custom_form.is_valid(), custom_form.errors)
        application = custom_form.save()
        self.assertEqual(application.custom_instrument, 'Домра малая II')
        self.assertIsNone(application.instrument_reference)
        self.assertFalse(Instrument.objects.filter(name='Домра малая II').exists())

    def test_course_application_form_stores_optional_orchestra_part(self):
        instrument = self.create_instrument(name='Альт')
        orchestra_part = OrchestraPart.objects.create(
            instrument=instrument,
            name='Партия второго альта',
        )
        form = CourseApplicationPublicForm(
            data=self.application_form_payload(
                instrument_reference=instrument,
                orchestra_part=orchestra_part.pk,
            ),
        )

        self.assertTrue(form.is_valid(), form.errors)
        application = form.save()
        self.assertEqual(application.orchestra_part, orchestra_part)
        self.assertEqual(application.student.orchestra_part, orchestra_part)
        self.assertEqual(
            application.student.enrollment_for_year(application.academic_year).orchestra_part,
            'Партия второго альта',
        )
        self.assertFalse(form.fields['orchestra_part'].required)
        self.assertIn('едет на курсы впервые', form.fields['orchestra_part'].help_text)

    def test_course_application_form_filters_parts_by_instrument(self):
        domra = self.create_instrument(name='Домра')
        bayan = self.create_instrument(name='Баян')
        domra_part = OrchestraPart.objects.create(
            instrument=domra,
            name='Малая первая',
        )
        bayan_part = OrchestraPart.objects.create(
            instrument=bayan,
            name='Первый',
        )
        inactive_part = OrchestraPart.objects.create(
            instrument=domra,
            name='Старая партия',
            is_active=False,
        )

        form = CourseApplicationPublicForm(
            data=self.application_form_payload(
                instrument_reference=domra,
                orchestra_part=domra_part.pk,
            ),
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(
            list(form.fields['orchestra_part'].queryset),
            [domra_part],
        )
        self.assertNotIn(bayan_part, form.fields['orchestra_part'].queryset)
        self.assertNotIn(inactive_part, form.fields['orchestra_part'].queryset)

    def test_course_application_form_rejects_part_from_another_instrument(self):
        domra = self.create_instrument(name='Домра')
        bayan = self.create_instrument(name='Баян')
        bayan_part = OrchestraPart.objects.create(
            instrument=bayan,
            name='Первый',
        )

        form = CourseApplicationPublicForm(
            data=self.application_form_payload(
                instrument_reference=domra,
                orchestra_part=bayan_part.pk,
            ),
        )

        self.assertFalse(form.is_valid())
        self.assertIn('orchestra_part', form.errors)

    def test_custom_instrument_disables_and_clears_orchestra_part(self):
        instrument = self.create_instrument(name='Домра')
        part = OrchestraPart.objects.create(
            instrument=instrument,
            name='Малая первая',
        )

        form = CourseApplicationPublicForm(
            data=self.application_form_payload(
                instrument_reference=None,
                custom_instrument='Скрипка',
                orchestra_part=part.pk,
            ),
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertTrue(form.fields['orchestra_part'].disabled)
        self.assertIsNone(form.cleaned_data['orchestra_part'])

    def test_parent_contacts_accepts_dash_from_form_placeholder(self):
        normalized_contacts = normalize_parent_contacts(
            'Иванов Иван Иванович — +7 (999) 123-45-67',
        )

        self.assertEqual(
            normalized_contacts,
            'Иванов Иван Иванович - +7 (999) 123-45-67',
        )

    def test_parent_contacts_placeholder_contains_only_examples(self):
        self.create_academic_year()
        placeholder = CourseApplicationPublicForm().fields['parent_contacts'].widget.attrs[
            'placeholder'
        ]

        self.assertEqual(
            placeholder,
            (
                'Иванов Иван Иванович — +7 (999) 123-45-67\n'
                'Петрова Анна Сергеевна — +7 (999) 987-65-43'
            ),
        )
        self.assertNotIn('Родитель', placeholder)

    def test_minimum_birth_date_for_age_handles_leap_course_start_date(self):
        self.assertEqual(
            minimum_birth_date_for_age(14, today=date(2024, 2, 29)),
            date(2010, 2, 28),
        )

    def test_public_course_application_form_enforces_age_limit(self):
        too_young_birth_date = date.today().replace(
            year=date.today().year - 10,
        ).isoformat()

        form = CourseApplicationPublicForm(
            data=self.application_form_payload(
                birth_date=too_young_birth_date,
            ),
        )

        self.assertFalse(form.is_valid())
        self.assertIn('birth_date', form.errors)

    def test_public_course_application_form_uses_age_reached_in_course_start_year(self):
        academic_year = self.create_academic_year(name='2025/2026')
        registration_settings = CourseRegistrationSettings.load(academic_year)
        registration_settings.telegram_group_url = 'https://t.me/test_group'
        registration_settings.minimum_registration_age = 15
        registration_settings.save()

        allowed_form = CourseApplicationPublicForm(
            data=self.application_form_payload(
                birth_date=date(2010, 12, 31),
            ),
            registration_settings=registration_settings,
        )
        too_young_form = CourseApplicationPublicForm(
            data=self.application_form_payload(
                birth_date=date(2011, 1, 1),
            ),
            registration_settings=registration_settings,
        )

        self.assertTrue(allowed_form.is_valid(), allowed_form.errors)
        self.assertFalse(too_young_form.is_valid())
        self.assertIn('birth_date', too_young_form.errors)
        self.assertEqual(
            allowed_form.fields['birth_date'].widget.attrs['max'],
            '2010-12-31',
        )
        self.assertEqual(
            allowed_form.fields['birth_date'].widget.attrs['data-age-limit'],
            '15',
        )
        self.assertEqual(
            allowed_form.fields['birth_date'].widget.attrs['data-age-reference-year'],
            '2025',
        )

    def test_calendar_year_age_helpers_accept_any_birthday_in_qualifying_year(self):
        self.assertEqual(
            latest_birth_date_for_age_in_year(14, year=2026),
            date(2012, 12, 31),
        )
        self.assertTrue(
            reaches_age_in_calendar_year(date(2012, 12, 31), 14, year=2026),
        )
        self.assertFalse(
            reaches_age_in_calendar_year(date(2013, 1, 1), 14, year=2026),
        )

    def test_music_education_choices_use_clear_education_levels(self):
        self.assertEqual(
            list(CourseApplication.MUSIC_EDUCATION_CHOICES),
            [
                ('self_taught', 'Самоучка'),
                ('basic', 'Музыкальная школа'),
                ('secondary', 'Колледж'),
                ('higher', 'Институт'),
            ],
        )

    def test_course_registration_settings_form_stores_age_and_uses_active_year_dates(self):
        academic_year = self.create_academic_year(name='2026/2027')
        form = CourseRegistrationSettingsForm(
            instance=CourseRegistrationSettings.load(),
            data={
                'telegram_group_url': ' https://t.me/test_group ',
                'minimum_registration_age': 16,
                'registration_mode': CourseRegistrationSettings.REGISTRATION_MODE_OPEN,
                'application_limit': '',
            },
        )

        self.assertTrue(form.is_valid(), form.errors)
        settings_obj = form.save()
        self.assertEqual(settings_obj.telegram_group_url, 'https://t.me/test_group')
        self.assertEqual(settings_obj.minimum_registration_age, 16)
        self.assertFalse(hasattr(settings_obj, 'course_starts_on'))
        self.assertFalse(hasattr(settings_obj, 'course_ends_on'))
        self.assertEqual(AcademicYear.get_active(), academic_year)

    def test_course_registration_settings_form_does_not_accept_course_dates(self):
        form = CourseRegistrationSettingsForm(
            data={
                'telegram_group_url': 'https://t.me/test_group',
                'minimum_registration_age': 14,
                'registration_mode': CourseRegistrationSettings.REGISTRATION_MODE_OPEN,
                'application_limit': '',
                'course_starts_on': '2026-08-31',
                'course_ends_on': '2025-09-01',
            },
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertNotIn('course_starts_on', form.fields)
        self.assertNotIn('course_ends_on', form.fields)

    def test_automatic_registration_mode_requires_positive_limit(self):
        form = CourseRegistrationSettingsForm(
            data={
                'telegram_group_url': 'https://t.me/test_group',
                'minimum_registration_age': 14,
                'registration_mode': CourseRegistrationSettings.REGISTRATION_MODE_AUTOMATIC,
                'application_limit': '',
            },
        )

        self.assertFalse(form.is_valid())
        self.assertIn('application_limit', form.errors)

    def test_grade_form_accepts_only_assigned_teacher_for_student_subject(self):
        data = self.create_base_journal()

        form = GradeCreateForm(
            data={
                'group': data['group'].pk,
                'student': data['student'].pk,
                'subject': data['solfeggio'].pk,
                'teacher': data['teacher'].pk,
                'academic_year': data['year'].pk,
                'date': '2025-10-10',
                'value': '5',
                'comment': '',
            },
        )

        self.assertTrue(form.is_valid(), form.errors)

        invalid_form = GradeCreateForm(
            data={
                'group': data['group'].pk,
                'student': data['student'].pk,
                'subject': data['solfeggio'].pk,
                'teacher': data['other_teacher'].pk,
                'academic_year': data['year'].pk,
                'date': '2025-10-10',
                'value': '5',
                'comment': '',
            },
        )

        self.assertFalse(invalid_form.is_valid())
        self.assertIn(
            'Этот преподаватель не назначен выбранному ученику',
            str(invalid_form.errors),
        )

    def test_grade_form_excludes_another_students_individual_subject(self):
        data = self.create_base_journal()
        unassigned_student = self.create_student(
            full_name='Ученик без индивидуального предмета',
            group=data['group'],
            instrument=data['student'].instrument,
            username='form_student_without_individual_subject',
        )

        form = GradeCreateForm(
            data={
                'group': data['group'].pk,
                'student': unassigned_student.pk,
                'subject': data['specialty'].pk,
                'teacher': data['other_teacher'].pk,
                'academic_year': data['year'].pk,
                'date': '2025-10-10',
                'value': '5',
                'comment': '',
            },
        )

        self.assertNotIn(data['specialty'], form.fields['subject'].queryset)
        self.assertIn(data['solfeggio'], form.fields['subject'].queryset)
        self.assertFalse(form.is_valid())
        self.assertIn('subject', form.errors)

    def test_grade_form_accepts_individual_subject_without_group(self):
        data = self.create_base_journal()
        data['student'].group = None
        data['student'].save(update_fields=['group'])

        form = GradeCreateForm(
            data={
                'group': '',
                'student': data['student'].pk,
                'subject': data['specialty'].pk,
                'teacher': data['other_teacher'].pk,
                'academic_year': data['year'].pk,
                'date': '2025-10-10',
                'value': '5',
                'comment': '',
            },
        )

        self.assertTrue(form.is_valid(), form.errors)
        grade = form.save()
        self.assertIsNone(grade.enrollment.group)

    def test_grade_form_with_fixed_teacher_removes_teacher_field(self):
        data = self.create_base_journal()

        form = GradeCreateForm(
            teacher=data['teacher'],
            group=data['group'],
            subject=data['solfeggio'],
            academic_year=data['year'],
        )

        self.assertNotIn('teacher', form.fields)
        self.assertNotIn('subject', form.fields)
        self.assertEqual(list(form.fields['student'].queryset), [data['student']])
        self.assertNotIn('data-searchable-select', form.fields['student'].widget.attrs)

    def test_grade_form_with_fixed_teacher_reports_invalid_subject_without_crash(self):
        data = self.create_base_journal()

        form = GradeCreateForm(
            data={
                'student': data['student'].pk,
                'subject': data['literature'].pk,
                'academic_year': data['year'].pk,
                'date': '2025-10-10',
                'value': '5',
                'comment': '',
            },
            teacher=data['teacher'],
            group=data['group'],
            academic_year=data['year'],
        )

        self.assertFalse(form.is_valid())
        self.assertIn(
            'Этот преподаватель не назначен выбранному ученику',
            str(form.errors),
        )

    def test_grade_form_rejects_student_from_another_group(self):
        data = self.create_base_journal()
        another_group = self.create_group(
            name='Другая группа',
            academic_year=data['year'],
        )

        form = GradeCreateForm(
            data={
                'group': another_group.pk,
                'student': data['student'].pk,
                'subject': data['solfeggio'].pk,
                'teacher': data['teacher'].pk,
                'academic_year': data['year'].pk,
                'date': '2025-10-10',
                'value': '5',
                'comment': '',
            },
        )

        self.assertFalse(form.is_valid())
        self.assertIn('Выбранный ученик недоступен', str(form.errors))

    def test_grade_form_rejects_inactive_student_from_forged_post(self):
        data = self.create_base_journal()
        data['student'].is_active = False
        data['student'].save()

        form = GradeCreateForm(
            data={
                'group': data['group'].pk,
                'student': data['student'].pk,
                'subject': data['solfeggio'].pk,
                'teacher': data['teacher'].pk,
                'academic_year': data['year'].pk,
                'date': '2025-10-10',
                'value': '5',
                'comment': '',
            },
        )

        self.assertFalse(form.is_valid())
        self.assertIn('Выбранный ученик недоступен', str(form.errors))

    def test_grade_form_rejects_group_from_another_academic_year(self):
        data = self.create_base_journal()
        another_year = AcademicYear.objects.create(
            name='2026/2027',
            starts_on=date(2026, 9, 1),
            ends_on=date(2027, 8, 31),
            is_active=False,
        )

        form = GradeCreateForm(
            data={
                'group': data['group'].pk,
                'student': data['student'].pk,
                'subject': data['solfeggio'].pk,
                'teacher': data['teacher'].pk,
                'academic_year': another_year.pk,
                'date': '2026-10-10',
                'value': '5',
                'comment': '',
            },
        )

        self.assertFalse(form.is_valid())
        self.assertIn('Группа относится к другому учебному году.', str(form.errors))

    def test_grade_admin_form_limits_related_fields_and_loads_dependency_script(self):
        data = self.create_base_journal()
        form = GradeAdminForm(
            data={
                'group': data['group'].pk,
                'student': data['student'].pk,
                'subject': data['solfeggio'].pk,
                'academic_year': data['year'].pk,
            },
        )

        self.assertEqual(
            set(form.fields['teacher'].queryset),
            {data['teacher'], data['other_teacher']},
        )
        self.assertEqual(list(form.fields['student'].queryset), [data['student']])
        self.assertFalse(form.fields['group'].required)
        self.assertIn('journal/grade_dependencies.js', GradeAdmin.Media.js)
        self.assertNotIn('journal/select_search.js', GradeAdmin.Media.js)
        dependency_script = Path(
            'journal/static/journal/grade_dependencies.js'
        ).read_text(encoding='utf-8')
        self.assertNotIn('clearDescendants', dependency_script)
        self.assertNotIn('individualMode', dependency_script)

    def test_blank_group_forms_render_all_year_assignments(self):
        data = self.create_base_journal()

        journal_form = GradeCreateForm(
            teacher=data['other_teacher'],
            academic_year=data['year'],
        )
        admin_form = GradeAdminForm(
            fixed_academic_year=data['year'],
        )

        self.assertEqual(
            set(journal_form.fields['subject'].queryset),
            {data['literature'], data['specialty']},
        )
        self.assertEqual(
            list(journal_form.fields['student'].queryset),
            [data['student']],
        )
        self.assertIn(data['specialty'], admin_form.fields['subject'].queryset)
        self.assertIn(data['solfeggio'], admin_form.fields['subject'].queryset)
        self.assertIn(data['literature'], admin_form.fields['subject'].queryset)

    def test_grade_dependency_options_show_year_values_before_group_selection(self):
        data = self.create_base_journal()
        other_year = self.create_academic_year(name='2026/2027')
        other_group = self.create_group(name='Группа другого года', academic_year=other_year)
        self.create_group_assignment(
            group=other_group,
            subject=data['solfeggio'],
            teacher=data['teacher'],
        )

        without_group = get_grade_form_options(academic_year=data['year'])
        with_group = get_grade_form_options(
            academic_year=data['year'],
            group=data['group'],
            student=data['student'],
            subject=data['literature'],
        )

        self.assertEqual(list(without_group['groups']), [data['group']])
        self.assertEqual(list(without_group['students']), [data['student']])
        self.assertEqual(
            set(without_group['subjects']),
            {data['solfeggio'], data['literature'], data['specialty']},
        )
        self.assertEqual(list(with_group['students']), [data['student']])
        self.assertEqual(
            set(with_group['subjects']),
            {data['solfeggio'], data['literature'], data['specialty']},
        )

    def test_grade_dependency_options_ignore_sibling_filters_without_group(self):
        data = self.create_base_journal()

        teacher_options = get_grade_form_options(
            academic_year=data['year'],
            teacher=data['other_teacher'],
        )
        student_options = get_grade_form_options(
            academic_year=data['year'],
            subject=data['specialty'],
            teacher=data['other_teacher'],
        )
        subject_options = get_grade_form_options(
            academic_year=data['year'],
            student=data['student'],
            teacher=data['teacher'],
        )

        self.assertEqual(list(teacher_options['students']), [data['student']])
        self.assertEqual(
            set(teacher_options['subjects']),
            {data['solfeggio'], data['literature'], data['specialty']},
        )
        self.assertEqual(list(student_options['students']), [data['student']])
        self.assertEqual(
            set(subject_options['subjects']),
            {data['solfeggio'], data['literature'], data['specialty']},
        )
        self.assertIn(data['specialty'], subject_options['subjects'])

    def test_grade_admin_uses_and_locks_year_selected_in_page_filter(self):
        data = self.create_base_journal()
        request = RequestFactory().get('/admin/journal/grade/add/', {'academic_year': data['year'].pk})
        request.user = User.objects.create_superuser(
            username='grade_form_admin',
            password='Pass12345!',
            email='admin@example.com',
        )
        request.session = {}
        model_admin = GradeAdmin(Grade, django_admin.site)

        form = model_admin.get_form(request)()

        self.assertTrue(form.fields['academic_year'].disabled)
        self.assertEqual(form.fields['academic_year'].initial, data['year'])
        for field_name in ('group', 'student', 'subject', 'teacher'):
            self.assertNotIn('data-searchable-select', form.fields[field_name].widget.attrs)

    def test_grade_edit_form_keeps_existing_date_value(self):
        form = GradeAdminForm(instance=Grade(date=date(2025, 10, 10), value='5'))

        self.assertIn('value="2025-10-10"', str(form['date']))

    def test_student_and_teacher_edit_forms_keep_existing_birth_dates(self):
        data = self.create_base_journal()
        data['student'].birth_date = date(2010, 3, 2)
        data['teacher'].birth_date = date(1980, 4, 3)

        student_form = StudentAdminForm(instance=data['student'])
        teacher_form = TeacherAdminForm(instance=data['teacher'])

        self.assertIn('value="2010-03-02"', str(student_form['birth_date']))
        self.assertIn('value="1980-04-03"', str(teacher_form['birth_date']))

    def test_subject_result_form_validates_allowed_subject_and_grade_type(self):
        data = self.create_base_journal()

        form = SubjectResultForm(
            data={
                'student': data['student'].pk,
                'subject': data['solfeggio'].pk,
                'academic_year': data['year'].pk,
                'exam_grade': '5',
                'final_grade': '4',
            },
        )

        self.assertTrue(form.is_valid(), form.errors)

        pass_fail_subject = self.create_subject(
            name='Зачетный предмет',
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )

        GroupSubject.objects.create(
            group=data['group'],
            subject=pass_fail_subject,
            teacher=data['teacher'],
        )

        invalid_form = SubjectResultForm(
            data={
                'student': data['student'].pk,
                'subject': pass_fail_subject.pk,
                'academic_year': data['year'].pk,
                'exam_grade': '5',
                'final_grade': '5',
            },
        )

        self.assertFalse(invalid_form.is_valid())
        self.assertIn('Допустимы значения', str(invalid_form.errors))

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_detailed_password_change_form_has_no_old_password_field_and_saves_new_password(
        self,
    ):
        user = User.objects.create_user(
            username='password_user',
            password='OldPass12345!',
        )

        form = DetailedPasswordChangeForm(
            user,
            data={
                'new_password1': 'NewPass12345!',
                'new_password2': 'NewPass12345!',
            },
        )

        self.assertNotIn('old_password', form.fields)
        self.assertEqual(
            form.fields['new_password1'].widget.attrs.get('autocomplete'),
            'new-password',
        )
        self.assertEqual(
            form.fields['new_password2'].widget.attrs.get('autocomplete'),
            'new-password',
        )
        self.assertTrue(form.is_valid(), form.errors)

        form.save()
        user.refresh_from_db()

        self.assertTrue(user.check_password('NewPass12345!'))

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_detailed_password_change_form_rejects_unchanged_password(self):
        user = User.objects.create_user(
            username='password_user',
            password='SamePass12345!',
        )

        form = DetailedPasswordChangeForm(
            user,
            data={
                'new_password1': 'SamePass12345!',
                'new_password2': 'SamePass12345!',
            },
        )

        self.assertFalse(form.is_valid())
        self.assertIn(
            'Новый пароль не должен совпадать со старым.',
            str(form.errors),
        )


class SelectorHelperTests(JournalTestDataMixin, TestCase):
    def test_helper_functions_return_only_real_assignments(self):
        data = self.create_base_journal()

        allowed_subjects = get_student_allowed_subjects(data['student'])

        self.assertIn(data['solfeggio'], allowed_subjects)
        self.assertIn(data['literature'], allowed_subjects)
        self.assertIn(data['specialty'], allowed_subjects)

        solfeggio_teachers = get_student_subject_teachers(
            data['student'],
            data['solfeggio'],
        )
        specialty_teachers = get_student_subject_teachers(
            data['student'],
            data['specialty'],
        )

        self.assertEqual(list(solfeggio_teachers), [data['teacher']])
        self.assertEqual(list(specialty_teachers), [data['other_teacher']])

        self.assertIn(data['group'], get_teacher_groups(data['teacher']))
        self.assertIn(data['solfeggio'], get_teacher_subjects(data['teacher']))
        self.assertNotIn(data['specialty'], get_teacher_subjects(data['teacher']))

    def test_grade_option_helpers_keep_only_complete_active_assignments(self):
        data = self.create_base_journal()

        self.assertEqual(
            list(get_grade_groups(teacher=data['teacher'])),
            [data['group']],
        )
        self.assertEqual(
            list(get_grade_students(
                group=data['group'],
                subject=data['solfeggio'],
                teacher=data['teacher'],
            )),
            [data['student']],
        )
        self.assertFalse(
            get_grade_students(
                group=data['group'],
                subject=data['solfeggio'],
                teacher=data['other_teacher'],
            ).exists(),
        )
        self.assertEqual(
            list(get_grade_subjects(
                group=data['group'],
                student=data['student'],
                teacher=data['teacher'],
            )),
            [data['solfeggio']],
        )
        self.assertEqual(
            list(get_grade_teachers(
                group=data['group'],
                student=data['student'],
                subject=data['specialty'],
            )),
            [data['other_teacher']],
        )

    def test_primary_group_options_ignore_legacy_individual_flag(self):
        data = self.create_base_journal()

        options = get_grade_form_options(
            academic_year=data['year'],
            fixed_teacher=data['other_teacher'],
            individual_only=True,
        )

        self.assertEqual(list(options['students']), [data['student']])
        self.assertEqual(
            set(options['subjects']),
            {data['literature'], data['specialty']},
        )
        self.assertEqual(list(options['teachers']), [data['other_teacher']])
        self.assertIn(data['literature'], options['subjects'])

    def test_grade_option_helpers_hide_archived_year_by_default_but_allow_explicit_view(self):
        data = self.create_base_journal()
        AcademicYear.objects.create(
            name='2026/2027',
            starts_on=date(2026, 9, 1),
            ends_on=date(2027, 8, 31),
        )

        self.assertNotIn(data['group'], get_grade_groups(teacher=data['teacher']))
        self.assertIn(
            data['group'],
            get_grade_groups(teacher=data['teacher'], academic_year=data['year']),
        )

    def test_assignment_availability_supports_standard_and_element_modes(self):
        data = self.create_base_journal()
        element_subject = Subject.objects.create(
            name='Оркестровые партии',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
        )
        GroupSubject.objects.create(
            group=data['group'],
            subject=element_subject,
            teacher=data['teacher'],
        )

        self.assertEqual(
            list(available_groups(
                subject=element_subject,
                teacher=data['teacher'],
                academic_year=data['year'],
            )),
            [data['group']],
        )
        self.assertEqual(
            list(available_students(
                group=data['group'],
                subject=element_subject,
                teacher=data['teacher'],
                academic_year=data['year'],
            )),
            [data['student']],
        )
        self.assertIn(
            element_subject,
            available_subjects(teacher=data['teacher'], academic_year=data['year']),
        )
        self.assertEqual(
            list(available_teachers(
                group=data['group'],
                student=data['student'],
                subject=element_subject,
                academic_year=data['year'],
            )),
            [data['teacher']],
        )
        self.assertNotIn(
            element_subject,
            get_grade_subjects(teacher=data['teacher'], academic_year=data['year']),
        )


class GradeOptionsApiTests(JournalTestDataMixin, TestCase):
    def setUp(self):
        self.data = self.create_base_journal()
        self.admin_user = User.objects.create_superuser(
            username='grade_options_admin',
            password='Pass12345!',
        )

    def create_alternative_group_assignment(self):
        group = self.create_group(name='Другая группа', academic_year=self.data['year'])
        student = self.create_student(
            full_name='Другой Ученик',
            group=group,
            instrument=self.data['student'].instrument,
            username='other_group_student',
        )
        subject = self.create_subject(name='Другой предмет')
        self.create_group_assignment(
            group=group,
            subject=subject,
            teacher=self.data['other_teacher'],
        )
        return group, student, subject

    def test_admin_options_narrow_teachers_for_selected_assignment(self):
        self.client.login(username='grade_options_admin', password='Pass12345!')

        response = self.client.get(
            reverse('grade_options_api'),
            {
                'group': self.data['group'].pk,
                'student': self.data['student'].pk,
                'subject': self.data['solfeggio'].pk,
                'academic_year': self.data['year'].pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(
            payload['teachers'],
            [{'id': self.data['teacher'].pk, 'label': self.data['teacher'].full_name}],
        )
        self.assertEqual(
            [item['id'] for item in payload['students']],
            [self.data['student'].pk],
        )
        self.assertEqual(payload['defaults']['group_id'], self.data['group'].pk)
        self.assertEqual(payload['defaults']['academic_year_id'], self.data['year'].pk)

    def test_grade_mode_returns_all_values_without_group(self):
        other_group, other_student, other_subject = self.create_alternative_group_assignment()
        self.client.login(username='grade_options_admin', password='Pass12345!')

        response = self.client.get(
            reverse('grade_options_api'),
            {
                'mode': 'grade',
                'teacher': self.data['other_teacher'].pk,
                'academic_year': self.data['year'].pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(
            {item['id'] for item in payload['students']},
            {self.data['student'].pk, other_student.pk},
        )
        self.assertEqual(
            {item['id'] for item in payload['subjects']},
            {
                self.data['solfeggio'].pk,
                self.data['literature'].pk,
                self.data['specialty'].pk,
                other_subject.pk,
            },
        )
        self.assertEqual(
            {item['id'] for item in payload['groups']},
            {self.data['group'].pk, other_group.pk},
        )

    def test_selecting_group_first_limits_students_and_subjects(self):
        _, other_student, other_subject = self.create_alternative_group_assignment()
        self.client.login(username='grade_options_admin', password='Pass12345!')

        response = self.client.get(
            reverse('grade_options_api'),
            {
                'mode': 'grade',
                'group': self.data['group'].pk,
                'student': other_student.pk,
                'subject': other_subject.pk,
                'academic_year': self.data['year'].pk,
                'changed': 'group',
                'strict': '1',
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(
            [item['id'] for item in payload['students']],
            [self.data['student'].pk],
        )
        self.assertNotIn(other_student.pk, [item['id'] for item in payload['students']])
        self.assertIn(
            self.data['solfeggio'].pk,
            [item['id'] for item in payload['subjects']],
        )
        self.assertNotIn(other_subject.pk, [item['id'] for item in payload['subjects']])

    def test_selecting_subject_first_limits_groups_and_students(self):
        other_group, other_student, other_subject = self.create_alternative_group_assignment()
        self.client.login(username='grade_options_admin', password='Pass12345!')

        response = self.client.get(
            reverse('grade_options_api'),
            {'subject': other_subject.pk, 'academic_year': self.data['year'].pk},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual([item['id'] for item in payload['groups']], [other_group.pk])
        self.assertEqual([item['id'] for item in payload['students']], [other_student.pk])

    def test_selecting_student_first_limits_group_and_subjects(self):
        other_subject = self.create_subject(
            name='Индивидуальный предмет',
            is_specialty=True,
        )
        self.create_individual_assignment(
            student=self.data['student'],
            subject=other_subject,
            teacher=self.data['other_teacher'],
        )
        unrelated_group = self.create_group(
            name='Другая группа',
            academic_year=self.data['year'],
        )
        self.client.login(username='grade_options_admin', password='Pass12345!')

        response = self.client.get(
            reverse('grade_options_api'),
            {'student': self.data['student'].pk, 'academic_year': self.data['year'].pk},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(
            [item['id'] for item in payload['groups']],
            [self.data['group'].pk],
        )
        self.assertNotIn(unrelated_group.pk, [item['id'] for item in payload['groups']])
        subject_ids = [item['id'] for item in payload['subjects']]
        self.assertIn(self.data['solfeggio'].pk, subject_ids)
        self.assertIn(other_subject.pk, subject_ids)

    def test_grade_mode_sibling_fields_do_not_hide_year_options(self):
        self.client.login(username='grade_options_admin', password='Pass12345!')

        response = self.client.get(
            reverse('grade_options_api'),
            {
                'mode': 'grade',
                'individual': '1',
                'student': self.data['student'].pk,
                'subject': self.data['specialty'].pk,
                'teacher': self.data['other_teacher'].pk,
                'academic_year': self.data['year'].pk,
                'changed': 'student',
                'strict': '1',
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn(
            self.data['specialty'].pk,
            [item['id'] for item in payload['subjects']],
        )
        self.assertIn(
            self.data['literature'].pk,
            [item['id'] for item in payload['subjects']],
        )
        self.assertEqual(
            [item['id'] for item in payload['students']],
            [self.data['student'].pk],
        )

    def test_grade_mode_hides_other_students_individual_subjects(self):
        unassigned_student = self.create_student(
            full_name='Ученик без индивидуального предмета',
            group=self.data['group'],
            instrument=self.data['student'].instrument,
            username='student_without_individual_subject',
        )
        self.client.login(username='grade_options_admin', password='Pass12345!')

        response = self.client.get(
            reverse('grade_options_api'),
            {
                'mode': 'grade',
                'group': self.data['group'].pk,
                'student': unassigned_student.pk,
                'academic_year': self.data['year'].pk,
                'changed': 'student',
                'strict': '1',
            },
        )

        self.assertEqual(response.status_code, 200)
        subject_ids = {item['id'] for item in response.json()['subjects']}
        self.assertIn(self.data['solfeggio'].pk, subject_ids)
        self.assertIn(self.data['literature'].pk, subject_ids)
        self.assertNotIn(self.data['specialty'].pk, subject_ids)

    def test_blank_group_does_not_enable_individual_only_mode(self):
        self.client.login(username='grade_options_admin', password='Pass12345!')

        response = self.client.get(
            reverse('grade_options_api'),
            {
                'mode': 'grade',
                'group': '',
                'teacher': self.data['other_teacher'].pk,
                'academic_year': self.data['year'].pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(
            {item['id'] for item in payload['subjects']},
            {
                self.data['solfeggio'].pk,
                self.data['literature'].pk,
                self.data['specialty'].pk,
            },
        )
        self.assertIn(
            self.data['literature'].pk,
            [item['id'] for item in payload['subjects']],
        )

    def test_changing_subject_drops_incompatible_group_and_student(self):
        other_group, other_student, other_subject = self.create_alternative_group_assignment()
        self.client.login(username='grade_options_admin', password='Pass12345!')

        response = self.client.get(
            reverse('grade_options_api'),
            {
                'group': self.data['group'].pk,
                'student': self.data['student'].pk,
                'subject': other_subject.pk,
                'academic_year': self.data['year'].pk,
                'changed': 'subject',
                'strict': '1',
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn(other_subject.pk, [item['id'] for item in payload['subjects']])
        self.assertNotIn(self.data['group'].pk, [item['id'] for item in payload['groups']])
        self.assertNotIn(self.data['student'].pk, [item['id'] for item in payload['students']])
        self.assertNotIn(other_group.pk, [item['id'] for item in payload['groups']])
        self.assertNotIn(other_student.pk, [item['id'] for item in payload['students']])

    def test_options_keep_currently_selected_values_when_other_field_changes(self):
        self.client.login(username='grade_options_admin', password='Pass12345!')

        response = self.client.get(
            reverse('grade_options_api'),
            {
                'group': self.data['group'].pk,
                'student': self.data['student'].pk,
                'subject': self.data['solfeggio'].pk,
                'teacher': self.data['other_teacher'].pk,
                'academic_year': self.data['year'].pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn(self.data['group'].pk, [item['id'] for item in payload['groups']])
        self.assertIn(self.data['student'].pk, [item['id'] for item in payload['students']])
        self.assertIn(self.data['solfeggio'].pk, [item['id'] for item in payload['subjects']])
        self.assertIn(self.data['other_teacher'].pk, [item['id'] for item in payload['teachers']])

    def test_strict_options_drop_incompatible_dependent_values(self):
        other_year = self.create_academic_year(name='2026/2027')
        other_group = self.create_group(name='Другая группа', academic_year=other_year)
        self.client.login(username='grade_options_admin', password='Pass12345!')

        response = self.client.get(
            reverse('grade_options_api'),
            {
                'group': other_group.pk,
                'student': self.data['student'].pk,
                'subject': self.data['solfeggio'].pk,
                'teacher': self.data['teacher'].pk,
                'changed': 'group',
                'strict': '1',
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn(other_group.pk, [item['id'] for item in payload['groups']])
        self.assertNotIn(self.data['student'].pk, [item['id'] for item in payload['students']])
        self.assertNotIn(self.data['solfeggio'].pk, [item['id'] for item in payload['subjects']])
        self.assertNotIn(self.data['teacher'].pk, [item['id'] for item in payload['teachers']])

    def test_teacher_options_are_always_limited_to_own_assignments(self):
        self.client.login(username='teacher_ivanov', password='Pass12345!')

        response = self.client.get(reverse('grade_options_api'))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(
            [item['id'] for item in payload['teachers']],
            [self.data['teacher'].pk],
        )
        self.assertIn(
            self.data['solfeggio'].pk,
            [item['id'] for item in payload['subjects']],
        )
        self.assertNotIn(
            self.data['literature'].pk,
            [item['id'] for item in payload['subjects']],
        )
        self.assertNotIn(
            self.data['specialty'].pk,
            [item['id'] for item in payload['subjects']],
        )

    def test_student_cannot_request_grade_entry_options(self):
        self.client.login(username='student_sidorov', password='Pass12345!')

        response = self.client.get(reverse('grade_options_api'))

        self.assertEqual(response.status_code, 403)


class AssignmentOptionsApiTests(JournalTestDataMixin, TestCase):
    def setUp(self):
        self.data = self.create_base_journal()
        self.admin_user = User.objects.create_superuser(
            username='assignment_options_admin',
            password='Pass12345!',
        )

    def test_student_subject_options_return_defaults_for_selected_subject(self):
        extra_subject = self.create_subject(
            name='Индивидуальная импровизация',
            is_specialty=True,
        )
        self.client.login(username='assignment_options_admin', password='Pass12345!')

        response = self.client.get(
            reverse('assignment_options_api'),
            {
                'type': 'student_subject',
                'student': self.data['student'].pk,
                'subject': extra_subject.pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['defaults']['subject_is_individual'])
        self.assertEqual(payload['defaults']['group_id'], self.data['group'].pk)
        self.assertEqual(payload['defaults']['academic_year_id'], self.data['year'].pk)
        self.assertIn(extra_subject.pk, [item['id'] for item in payload['subjects']])

    def test_group_subject_options_return_next_sort_order_and_group_year(self):
        self.client.login(username='assignment_options_admin', password='Pass12345!')

        response = self.client.get(
            reverse('assignment_options_api'),
            {
                'type': 'group_subject',
                'group': self.data['group'].pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['defaults']['academic_year_id'], self.data['year'].pk)
        self.assertEqual(payload['defaults']['sort_order'], 110)
        self.assertIn(self.data['group'].pk, [item['id'] for item in payload['groups']])


class RegistrationSettingsByAcademicYearTests(JournalTestDataMixin, TestCase):
    def test_each_academic_year_has_independent_registration_settings(self):
        old_year = self.create_academic_year(name='2025/2026')
        old_settings = CourseRegistrationSettings.load(old_year)
        old_settings.telegram_group_url = 'https://t.me/old_year'
        old_settings.minimum_registration_age = 16
        old_settings.registration_mode = CourseRegistrationSettings.REGISTRATION_MODE_CLOSED
        old_settings.save()

        new_year = self.create_academic_year(name='2026/2027')
        new_settings = CourseRegistrationSettings.load(new_year)

        self.assertNotEqual(old_settings.pk, new_settings.pk)
        self.assertEqual(new_settings.minimum_registration_age, 14)
        self.assertEqual(
            new_settings.registration_mode,
            CourseRegistrationSettings.REGISTRATION_MODE_OPEN,
        )
        old_settings.refresh_from_db()
        self.assertEqual(old_settings.telegram_group_url, 'https://t.me/old_year')
        self.assertEqual(old_settings.minimum_registration_age, 16)
        self.assertEqual(CourseRegistrationSettings.objects.count(), 2)


class BirthdayNotificationTests(JournalTestDataMixin, TestCase):
    def setUp(self):
        self.data = self.create_base_journal()
        self.admin_user = User.objects.create_superuser(
            username='birthday_admin',
            password='Pass12345!',
            first_name='Анна',
            last_name='Администраторова',
        )
        self.data['student'].birth_date = date(2008, 7, 30)
        self.data['student'].save()
        self.data['teacher'].birth_date = date(1980, 7, 31)
        self.data['teacher'].save()
        AccountProfile.objects.create(
            user=self.admin_user,
            birth_date=date(1990, 7, 30),
        )

    def test_admin_and_teacher_receive_today_and_tomorrow_birthdays_with_age(self):
        today = date(2026, 7, 30)

        admin_messages = [
            item['message']
            for item in birthday_notifications_for_user(self.admin_user, today=today)
        ]
        teacher_messages = [
            item['message']
            for item in birthday_notifications_for_user(self.data['teacher'].user, today=today)
        ]

        expected_fragments = (
            'Сегодня день рождения: Сидоров Семён Семёнович (ученик) — исполнилось 18 лет.',
            'Сегодня день рождения: Администраторова Анна (администратор) — исполнилось 36 лет.',
            'Завтра день рождения: Иванов Иван Иванович (преподаватель) — исполнится 46 лет.',
        )
        for expected in expected_fragments:
            self.assertIn(expected, admin_messages)
            self.assertIn(expected, teacher_messages)

    def test_student_receives_all_birthday_notifications(self):
        messages = [
            item['message']
            for item in birthday_notifications_for_user(
                self.data['student'].user,
                today=date(2026, 7, 30),
            )
        ]

        self.assertIn(
            'Сегодня день рождения: Сидоров Семён Семёнович (ученик) — исполнилось 18 лет.',
            messages,
        )
        self.assertIn(
            'Завтра день рождения: Иванов Иван Иванович (преподаватель) — исполнится 46 лет.',
            messages,
        )

    def test_birthday_notifications_render_for_admin_teacher_and_student(self):
        with patch('journal.birthday_notifications.timezone.localdate', return_value=date(2026, 7, 30)):
            self.client.force_login(self.admin_user)
            admin_response = self.client.get(reverse('admin:index'))
            self.client.force_login(self.data['teacher'].user)
            teacher_response = self.client.get(reverse('journal'))
            self.client.force_login(self.data['student'].user)
            student_response = self.client.get(reverse('journal'))

        self.assertContains(admin_response, 'Сидоров Семён Семёнович')
        self.assertContains(admin_response, 'исполнилось 18 лет')
        self.assertContains(teacher_response, 'Иванов Иван Иванович')
        self.assertContains(teacher_response, 'исполнится 46 лет')
        self.assertContains(student_response, 'Администраторова Анна')
        self.assertContains(student_response, 'исполнилось 36 лет')


class ViewTests(JournalTestDataMixin, TestCase):
    def setUp(self):
        self.data = self.create_base_journal()
        self.admin_user = User.objects.create_superuser(
            username='admin_test',
            password='Pass12345!',
            email='admin@example.com',
        )

    def test_teacher_can_open_journal_only_with_assigned_data(self):
        self.client.login(username='teacher_ivanov', password='Pass12345!')

        response = self.client.get(
            reverse('journal'),
            {'group': self.data['group'].pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Сольфеджио')
        self.assertNotContains(response, 'Регистрация на курсы')

    def test_student_table_row_shows_two_name_words_without_icon(self):
        self.client.force_login(self.admin_user)

        response = self.client.get(
            reverse('journal'),
            {
                'group': self.data['group'].pk,
                'subject': self.data['solfeggio'].pk,
                'academic_year': self.data['year'].pk,
            },
        )
        html = response.content.decode()
        row_start = html.index('<span class="student-name">')
        row_end = html.index('</span>', html.index('<span>', row_start)) + len('</span>')
        student_cell = html[row_start:row_end]

        self.assertEqual(short_person_name(self.data['student'].full_name), 'Сидоров Семён')
        self.assertIn('Сидоров Семён', student_cell)
        self.assertNotIn('Семёнович', student_cell)
        self.assertNotIn('<svg', student_cell)

    def test_orchestra_part_options_api_returns_only_active_instrument_parts(self):
        domra = self.create_instrument(name='Домра')
        bayan = self.data['instrument']
        domra_part = OrchestraPart.objects.create(
            instrument=domra,
            name='Малая первая',
        )
        OrchestraPart.objects.create(
            instrument=domra,
            name='Старая партия',
            is_active=False,
        )
        OrchestraPart.objects.create(
            instrument=bayan,
            name='Первый',
        )

        response = self.client.get(
            reverse('orchestra_part_options_api'),
            {'instrument': domra.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.json(),
            {'parts': [{'id': domra_part.pk, 'name': 'Малая первая'}]},
        )

    def test_registration_page_loads_orchestra_part_dependency_script(self):
        response = self.client.get(reverse('course_registration'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'journal/orchestra_part_dependencies_v5.js')
        self.assertContains(response, 'data-orchestra-part="1"')

    def test_grade_options_api_keeps_upstream_groups_independent_of_children(self):
        second_group = self.create_group(name='Вторая группа', academic_year=self.data['year'])
        self.create_group_assignment(
            group=second_group,
            subject=self.data['solfeggio'],
            teacher=self.data['teacher'],
        )
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse('grade_options_api'), {
            'mode': 'grade',
            'academic_year': self.data['year'].pk,
            'group': self.data['group'].pk,
            'student': self.data['student'].pk,
            'subject': self.data['literature'].pk,
        })

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(
            {item['id'] for item in payload['groups']},
            {self.data['group'].pk, second_group.pk},
        )
        self.assertEqual(
            {item['id'] for item in payload['students']},
            {self.data['student'].pk},
        )
        self.assertEqual(
            {item['id'] for item in payload['subjects']},
            {
                self.data['solfeggio'].pk,
                self.data['literature'].pk,
                self.data['specialty'].pk,
            },
        )

    def test_teacher_sees_all_assigned_tables_without_selecting_group(self):
        self.client.login(username='teacher_ivanov', password='Pass12345!')

        response = self.client.get(
            reverse('journal'),
            {'academic_year': self.data['year'].pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['journal_tables'])
        self.assertIsNotNone(response.context['grade_form'])
        self.assertContains(response, 'id="grade-create-form"')
        self.assertContains(response, 'data-save-context="grade-create"')
        self.assertContains(response, 'name="academic_year"')
        self.assertContains(response, 'data-workspace-search')
        self.assertContains(response, 'data-grade-dependency-mode="journal_filter"')
        self.assertContains(response, 'id="journal-subject-blocks"')
        self.assertContains(response, 'Свернуть все таблицы оценок по предметам')
        self.assertContains(response, 'Развернуть все таблицы оценок по предметам')
        self.assertContains(response, 'data-collapse-target="journal-subject-blocks"', count=1)
        self.assertContains(response, 'data-save-context="journal-table-')
        self.assertContains(
            response,
            (
                'data-collapse-key="journal-subject-'
                f'{self.data["teacher"].user_id}-{self.data["year"].pk}-'
                f'{self.data["group"].pk}-{self.data["solfeggio"].pk}"'
            ),
        )
        self.assertContains(
            response,
            (
                f'{self.data["group"].name} — '
                f'{self.data["solfeggio"].name} — '
                f'{self.data["year"].name}'
            ),
        )

    def test_admin_journal_has_admin_link_and_complete_table_title(self):
        self.client.login(username='admin_test', password='Pass12345!')

        response = self.client.get(
            reverse('journal'),
            {'academic_year': self.data['year'].pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse('admin:index'))
        self.assertContains(response, 'Перейти в админку')
        self.assertContains(
            response,
            (
                f'{self.data["group"].name} — '
                f'{self.data["solfeggio"].name} — '
                f'{self.data["year"].name}'
            ),
        )

    def test_student_without_group_can_open_journal(self):
        self.data['student'].group = None
        self.data['student'].save(update_fields=['group'])
        self.client.login(username='student_sidorov', password='Pass12345!')

        response = self.client.get(reverse('journal'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-filter-auto-submit="1"')
        self.assertNotContains(response, 'data-grade-dependency-mode="journal_filter"')
        self.assertContains(response, 'Мои оценки')
        self.assertContains(response, 'Нет данных по выбранным фильтрам.')

    def test_user_with_temporary_password_is_redirected_without_warning(self):
        TemporaryCredential.objects.create(
            user=self.data['teacher'].user,
            login=self.data['teacher'].user.username,
            temporary_password='abc234de',
        )
        self.data['teacher'].user.username = 'teacher_renamed'
        self.data['teacher'].user.save(update_fields=['username'])
        self.client.login(username='teacher_ivanov', password='Pass12345!')
        self.client.force_login(self.data['teacher'].user)

        response = self.client.get(reverse('journal'), follow=True)

        self.assertEqual(
            response.redirect_chain,
            [(reverse('password_change'), 302)],
        )
        self.assertNotContains(response, 'Смените временный пароль')

    def test_academic_year_filter_limits_admin_groups(self):
        next_year = AcademicYear.objects.create(
            name='2026/2027',
            starts_on=date(2026, 9, 1),
            ends_on=date(2027, 8, 31),
            is_active=False,
        )
        next_group = StudyGroup.objects.create(
            name='Группа следующего года',
            academic_year=next_year,
        )
        next_subject = self.create_subject(name='Предмет следующего года')
        next_student = self.create_student(
            full_name='Ученик Следующего Года',
            group=next_group,
            instrument=self.data['instrument'],
            username='student_next_year',
        )
        GroupSubject.objects.create(
            group=next_group,
            subject=next_subject,
            teacher=self.data['teacher'],
        )
        Grade.objects.create(
            student=next_student,
            subject=next_subject,
            teacher=self.data['teacher'],
            academic_year=next_year,
            date=date(2026, 10, 1),
            value='5',
        )
        self.client.login(username='admin_test', password='Pass12345!')

        response = self.client.get(
            reverse('journal'),
            {'academic_year': self.data['year'].pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.data['group'].name)
        self.assertNotContains(response, next_group.name)

    def test_inline_edit_rolls_back_when_later_value_is_invalid(self):
        grade = Grade.objects.create(
            student=self.data['student'],
            subject=self.data['solfeggio'],
            teacher=self.data['teacher'],
            academic_year=self.data['year'],
            date=date(2025, 10, 15),
            value='5',
        )
        self.client.login(username='admin_test', password='Pass12345!')

        response = self.client.post(
            f'{reverse("journal")}?group={self.data["group"].pk}&academic_year={self.data["year"].pk}',
            data={
                'action': 'inline_edit',
                (
                    f'grade__{self.data["solfeggio"].pk}__'
                    f'{self.data["student"].pk}__2025-10-15'
                ): '4',
                f'final__{self.data["solfeggio"].pk}__{self.data["student"].pk}': 'bad',
            },
        )

        self.assertEqual(response.status_code, 200)
        grade.refresh_from_db()
        self.assertEqual(grade.value, '5')

    def test_final_grade_controls_are_visible_without_regular_grades(self):
        self.client.login(username='admin_test', password='Pass12345!')

        response = self.client.get(
            reverse('journal'),
            {
                'group': self.data['group'].pk,
                'subject': self.data['solfeggio'].pk,
                'academic_year': self.data['year'].pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            f'name="final__{self.data["solfeggio"].pk}__{self.data["student"].pk}"',
        )

    def test_pass_fail_subject_uses_matching_exam_and_final_grade_options(self):
        pass_fail_subject = self.create_subject(
            name='Зачетный предмет',
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )
        GroupSubject.objects.create(
            group=self.data['group'],
            subject=pass_fail_subject,
            teacher=self.data['teacher'],
        )
        self.client.force_login(self.admin_user)
        journal_url = (
            f'{reverse("journal")}?group={self.data["group"].pk}'
            f'&subject={pass_fail_subject.pk}'
            f'&academic_year={self.data["year"].pk}'
        )

        response = self.client.get(journal_url)

        self.assertEqual(response.status_code, 200)
        html = response.content.decode()
        exam_name = f'exam__{pass_fail_subject.pk}__{self.data["student"].pk}'
        final_name = f'final__{pass_fail_subject.pk}__{self.data["student"].pk}'
        exam_start = html.index(f'<select name="{exam_name}"')
        exam_end = html.index('</select>', exam_start)
        exam_control = html[exam_start:exam_end]
        final_start = html.index(f'<select name="{final_name}"')
        final_end = html.index('</select>', final_start)
        final_control = html[final_start:final_end]

        for control in (exam_control, final_control):
            self.assertIn('<option value="Зачет">', control)
            self.assertIn('<option value="Незачет">', control)
            self.assertIn('<option value="Не аттестован">', control)
            self.assertNotIn('<option value="5">', control)

        invalid_response = self.client.post(journal_url, {
            'action': 'inline_edit',
            exam_name: '5',
        })
        self.assertEqual(invalid_response.status_code, 200)
        self.assertFalse(
            SubjectResult.objects.filter(
                student=self.data['student'],
                subject=pass_fail_subject,
                academic_year=self.data['year'],
            ).exists(),
        )

        valid_response = self.client.post(journal_url, {
            'action': 'inline_edit',
            exam_name: 'Зачет',
        })
        self.assertEqual(valid_response.status_code, 302)
        result = SubjectResult.objects.get(
            student=self.data['student'],
            subject=pass_fail_subject,
            academic_year=self.data['year'],
        )
        self.assertEqual(result.exam_grade, 'Зачет')

    def test_archived_academic_year_is_read_only_in_journal(self):
        grade = Grade.objects.create(
            student=self.data['student'],
            subject=self.data['solfeggio'],
            teacher=self.data['teacher'],
            academic_year=self.data['year'],
            date=date(2025, 10, 15),
            value='5',
        )
        AcademicYear.objects.create(
            name='2026/2027',
            starts_on=date(2026, 9, 1),
            ends_on=date(2027, 8, 31),
        )
        self.client.login(username='admin_test', password='Pass12345!')

        response = self.client.get(
            reverse('journal'),
            {
                'group': self.data['group'].pk,
                'subject': self.data['solfeggio'].pk,
                'academic_year': self.data['year'].pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, grade.value)
        self.assertNotContains(
            response,
            (
                f'name="grade__{self.data["solfeggio"].pk}__'
                f'{self.data["student"].pk}__2025-10-15"'
            ),
        )
        self.assertNotContains(response, '<button class="table-save-button"')
        self.assertNotContains(response, 'name="action" value="add_grade"')

    def test_archived_journal_uses_snapshots_after_current_records_are_renamed(self):
        old_student_name = self.data['student'].full_name
        old_subject_name = self.data['solfeggio'].name
        Grade.objects.create(
            student=self.data['student'],
            subject=self.data['solfeggio'],
            teacher=self.data['teacher'],
            academic_year=self.data['year'],
            date=date(2025, 10, 15),
            value='5',
        )
        AcademicYear.objects.create(
            name='2026/2027',
            starts_on=date(2026, 9, 1),
            ends_on=date(2027, 8, 31),
        )
        self.data['student'].refresh_from_db()
        self.data['student'].full_name = 'Текущее имя ученика'
        self.data['student'].save()
        self.data['solfeggio'].name = 'Текущее название предмета'
        self.data['solfeggio'].save()
        self.client.login(username='admin_test', password='Pass12345!')

        response = self.client.get(
            reverse('journal'),
            {
                'group': self.data['group'].pk,
                'subject': self.data['solfeggio'].pk,
                'academic_year': self.data['year'].pk,
            },
        )

        self.assertContains(response, ' '.join(old_student_name.split()[:2]))
        self.assertContains(response, old_subject_name)

    def test_post_to_archived_academic_year_does_not_change_grade(self):
        grade = Grade.objects.create(
            student=self.data['student'],
            subject=self.data['solfeggio'],
            teacher=self.data['teacher'],
            academic_year=self.data['year'],
            date=date(2025, 10, 15),
            value='5',
        )
        AcademicYear.objects.create(
            name='2026/2027',
            starts_on=date(2026, 9, 1),
            ends_on=date(2027, 8, 31),
        )
        self.client.login(username='admin_test', password='Pass12345!')

        response = self.client.post(
            (
                f'{reverse("journal")}?group={self.data["group"].pk}'
                f'&subject={self.data["solfeggio"].pk}'
                f'&academic_year={self.data["year"].pk}'
            ),
            data={
                'action': 'inline_edit',
                (
                    f'grade__{self.data["solfeggio"].pk}__'
                    f'{self.data["student"].pk}__2025-10-15'
                ): '4',
            },
        )

        self.assertEqual(response.status_code, 302)
        grade.refresh_from_db()
        self.assertEqual(grade.value, '5')
        self.assertEqual(grade.teacher, self.data['teacher'])

    def test_journal_table_builder_batches_assignment_queries(self):
        second_group = self.create_group(
            name='Вторая группа',
            academic_year=self.data['year'],
        )
        second_student = self.create_student(
            full_name='Ученик Второй Группы',
            group=second_group,
            instrument=self.data['instrument'],
            username='student_second_group',
        )
        extra_subject = self.create_subject(name='Хор')
        GroupSubject.objects.create(
            group=second_group,
            subject=extra_subject,
            teacher=self.data['teacher'],
        )

        groups = [self.data['group'], second_group]
        subjects = [
            self.data['solfeggio'],
            self.data['literature'],
            self.data['specialty'],
            extra_subject,
        ]
        enrollments = StudentEnrollment.objects.filter(
            academic_year=self.data['year'],
            student__in=[self.data['student'], second_student],
        ).select_related('student', 'group')

        with CaptureQueriesContext(connection) as captured_queries:
            journal_tables = _build_journal_tables(
                groups=groups,
                subjects=subjects,
                enrollments=enrollments,
                grade_qs=Grade.objects.none(),
                results_qs=SubjectResult.objects.none(),
                selected_academic_year=self.data['year'],
            )

        self.assertGreaterEqual(len(journal_tables), 4)
        self.assertLessEqual(
            len(captured_queries),
            3,
            [query['sql'] for query in captured_queries],
        )

    def test_student_cannot_edit_inline_grades(self):
        self.client.login(username='student_sidorov', password='Pass12345!')

        response = self.client.post(
            reverse('journal'),
            data={
                'action': 'inline_edit',
                (
                    f'grade__{self.data["solfeggio"].pk}__'
                    f'{self.data["student"].pk}__2025-10-15'
                ): '5',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Grade.objects.exists())

    def test_admin_can_add_grade_by_form(self):
        self.client.login(username='admin_test', password='Pass12345!')
        current_url = (
            f'{reverse("journal")}?group={self.data["group"].pk}'
            f'&subject={self.data["solfeggio"].pk}'
            f'&academic_year={self.data["year"].pk}'
        )

        response = self.client.post(
            current_url,
            data={
                'action': 'add_grade',
                'student': self.data['student'].pk,
                'subject': self.data['solfeggio'].pk,
                'teacher': self.data['teacher'].pk,
                'academic_year': self.data['year'].pk,
                'date': '2025-10-16',
                'value': '5',
                'comment': '',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], current_url)
        self.assertTrue(
            Grade.objects.filter(
                student=self.data['student'],
                subject=self.data['solfeggio'],
                teacher=self.data['teacher'],
                date=date(2025, 10, 16),
                value='5',
            ).exists(),
        )

    def test_blank_group_keeps_individual_grade_outside_filtered_group(self):
        other_group = self.create_group(
            name='Другая группа',
            academic_year=self.data['year'],
        )
        other_student = self.create_student(
            full_name='Индивидуальный Ученик',
            group=other_group,
            instrument=self.data['instrument'],
            username='individual_grade_student',
        )
        StudentSubject.objects.create(
            student=other_student,
            subject=self.data['specialty'],
            teacher=self.data['other_teacher'],
        )
        self.client.login(username='admin_test', password='Pass12345!')

        response = self.client.post(
            f'{reverse("journal")}?group={self.data["group"].pk}',
            data={
                'action': 'add_grade',
                'group': '',
                'student': other_student.pk,
                'subject': self.data['specialty'].pk,
                'teacher': self.data['other_teacher'].pk,
                'academic_year': self.data['year'].pk,
                'date': '2025-10-18',
                'value': '5',
                'comment': '',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            Grade.objects.filter(
                student=other_student,
                subject=self.data['specialty'],
                teacher=self.data['other_teacher'],
            ).exists(),
        )

    def test_grade_form_error_is_rendered_only_once_next_to_its_field(self):
        self.client.force_login(self.admin_user)

        response = self.client.post(
            f'{reverse("journal")}?academic_year={self.data["year"].pk}&group={self.data["group"].pk}',
            data={
                'action': 'add_grade',
                'group': self.data['group'].pk,
                'student': self.data['student'].pk,
                'subject': self.data['solfeggio'].pk,
                'teacher': self.data['teacher'].pk,
                'date': '2025-10-16',
                'value': '9',
                'comment': '',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            'Допустимы оценки от 1 до 5 со знаком +/− либо Н.',
            count=1,
        )

    def test_grade_post_uses_form_selection_instead_of_page_subject_filter(self):
        second_student = self.create_student(
            full_name='Второй Ученик',
            group=self.data['group'],
            instrument=self.data['instrument'],
            username='student_second_grade_form',
        )
        self.client.login(username='admin_test', password='Pass12345!')

        response = self.client.post(
            (
                f'{reverse("journal")}?group={self.data["group"].pk}'
                f'&subject={self.data["specialty"].pk}'
                f'&academic_year={self.data["year"].pk}'
            ),
            data={
                'action': 'add_grade',
                'group': self.data['group'].pk,
                'student': second_student.pk,
                'subject': self.data['solfeggio'].pk,
                'teacher': self.data['teacher'].pk,
                'academic_year': self.data['year'].pk,
                'date': '2025-10-17',
                'value': '4+',
                'comment': '',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            Grade.objects.filter(
                student=second_student,
                subject=self.data['solfeggio'],
                teacher=self.data['teacher'],
                value='4+',
            ).exists(),
        )

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_password_change_view_uses_set_password_form_without_old_password(
        self,
    ):
        TemporaryCredential.objects.create(
            user=self.data['teacher'].user,
            login=self.data['teacher'].user.username,
            temporary_password='abc234de',
        )
        self.data['teacher'].user.username = 'teacher_renamed'
        self.data['teacher'].user.save(update_fields=['username'])
        self.client.login(username='teacher_ivanov', password='Pass12345!')
        self.client.force_login(self.data['teacher'].user)

        get_response = self.client.get(reverse('password_change'))

        self.assertEqual(get_response.status_code, 200)
        self.assertNotContains(get_response, 'name="old_password"')
        self.assertContains(get_response, 'name="username"')
        self.assertContains(get_response, 'value="teacher_renamed"')
        self.assertContains(get_response, 'autocomplete="username"')
        self.assertContains(get_response, 'autocomplete="new-password"', count=2)
        self.assertContains(get_response, 'Текущий логин:')
        self.assertContains(get_response, 'name="new_password1"')
        self.assertContains(get_response, 'name="new_password2"')

        post_response = self.client.post(
            reverse('password_change'),
            data={
                'new_password1': 'NewPass12345!',
                'new_password2': 'NewPass12345!',
            },
        )

        self.assertEqual(post_response.status_code, 302)
        self.assertFalse(
            TemporaryCredential.objects.filter(
                user=self.data['teacher'].user,
            ).exists(),
        )

        journal_response = self.client.get(reverse('journal'))

        self.assertEqual(journal_response.status_code, 200)
        self.assertNotContains(journal_response, 'Смените временный пароль')

        self.client.logout()

        self.assertTrue(
            self.client.login(
                username='teacher_renamed',
                password='NewPass12345!',
            ),
        )


class CourseApplicationDuplicateErrorTests(TestCase):
    def test_recognizes_late_model_validation_duplicate(self):
        error = ValidationError({
            'student_phone': ValidationError(
                'duplicate',
                code='duplicate_phone_for_year',
            ),
        })
        self.assertTrue(_is_duplicate_course_application_phone_error(error))

    def test_does_not_mask_unrelated_validation_error(self):
        error = ValidationError({'birth_date': 'invalid'})
        self.assertFalse(_is_duplicate_course_application_phone_error(error))

    def test_recognizes_sqlite_unique_error_shape(self):
        error = IntegrityError(
            'UNIQUE constraint failed: '
            'journal_courseapplication.academic_year_id, '
            'journal_courseapplication.student_phone'
        )
        self.assertTrue(_is_duplicate_course_application_phone_error(error))

    def test_does_not_mask_unrelated_integrity_error(self):
        error = IntegrityError('NOT NULL constraint failed: other.field')
        self.assertFalse(_is_duplicate_course_application_phone_error(error))


@tag('slow', 'concurrency')
@skipUnless(connection.vendor == 'postgresql', 'PostgreSQL concurrency test')
class CourseApplicationConcurrencyTests(JournalTestDataMixin, TransactionTestCase):
    reset_sequences = True

    def test_concurrent_same_phone_creates_exactly_one_application(self):
        year = self.create_academic_year()
        barrier = Barrier(2)
        result_lock = Lock()
        results = []
        original_full_clean = CourseApplication.full_clean

        def synchronized_full_clean(instance, *args, **kwargs):
            original_full_clean(instance, *args, **kwargs)
            barrier.wait(timeout=10)

        def create_application(last_name):
            close_old_connections()
            try:
                payload = self.application_payload(
                    last_name=last_name,
                    academic_year_id=year.pk,
                )
                with transaction.atomic():
                    CourseApplication.objects.create(**payload)
            except IntegrityError:
                outcome = 'duplicate'
            else:
                outcome = 'created'
            finally:
                close_old_connections()
            with result_lock:
                results.append(outcome)

        with patch.object(CourseApplication, 'full_clean', synchronized_full_clean):
            threads = [
                Thread(target=create_application, args=('Иванов',), daemon=True),
                Thread(target=create_application, args=('Петров',), daemon=True),
            ]
            for thread in threads:
                thread.start()
            for thread in threads:
                thread.join(timeout=15)

        self.assertFalse(any(thread.is_alive() for thread in threads))
        self.assertCountEqual(results, ['created', 'duplicate'])
        self.assertEqual(
            CourseApplication.objects.filter(
                academic_year=year,
                student_phone='+7 (999) 123-45-67',
            ).count(),
            1,
        )

    def test_concurrent_same_identity_reuses_one_student_and_account(self):
        year = self.create_academic_year()
        barrier = Barrier(2)
        result_lock = Lock()
        results = []
        original_full_clean = CourseApplication.full_clean

        def synchronized_full_clean(instance, *args, **kwargs):
            original_full_clean(instance, *args, **kwargs)
            barrier.wait(timeout=10)

        def create_application(phone):
            close_old_connections()
            try:
                application = CourseApplication.objects.create(
                    **self.application_payload(
                        student_phone=phone,
                        academic_year_id=year.pk,
                    ),
                )
                outcome = (
                    application.pk,
                    application.student_id,
                    application.user_id,
                )
            except Exception as exc:  # Stored for an explicit assertion in the main thread.
                outcome = exc
            finally:
                close_old_connections()
            with result_lock:
                results.append(outcome)

        with patch.object(CourseApplication, 'full_clean', synchronized_full_clean):
            threads = [
                Thread(
                    target=create_application,
                    args=(phone,),
                    daemon=True,
                )
                for phone in ('+7 (999) 123-45-67', '+7 (999) 765-43-21')
            ]
            for thread in threads:
                thread.start()
            for thread in threads:
                thread.join(timeout=20)

        self.assertFalse(any(thread.is_alive() for thread in threads))
        self.assertEqual(len(results), 2)
        self.assertFalse([result for result in results if isinstance(result, Exception)], results)
        student_ids = {result[1] for result in results}
        user_ids = {result[2] for result in results}
        self.assertEqual(len(student_ids), 1)
        self.assertEqual(len(user_ids), 1)
        self.assertNotIn(None, student_ids)
        self.assertNotIn(None, user_ids)
        self.assertEqual(CourseApplication.objects.filter(academic_year=year).count(), 2)
        self.assertEqual(Student.objects.filter(pk__in=student_ids).count(), 1)
        self.assertEqual(TemporaryCredential.objects.filter(user_id__in=user_ids).count(), 1)


class JournalCategoryVisibilityTests(JournalTestDataMixin, TestCase):
    def test_student_without_assigned_subjects_does_not_see_subject_or_assessment_categories(self):
        year = self.create_academic_year()
        group = self.create_group(academic_year=year)
        student = self.create_student(group=group, username='student_without_subjects')
        self.client.force_login(student.user)

        response = self.client.get(reverse('journal'), {'academic_year': year.pk})

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context['has_subjects'])
        self.assertNotContains(response, 'data-filter-field="subject"')
        self.assertNotContains(response, 'data-summary-category="subjects"')
        self.assertNotContains(response, 'data-journal-category="assessments"')

    def test_teacher_with_subject_but_without_assigned_works_sees_only_subject_category(self):
        data = self.create_base_journal()
        self.client.force_login(data['teacher'].user)

        response = self.client.get(
            reverse('journal'),
            {'academic_year': data['year'].pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['has_subjects'])
        self.assertFalse(response.context['assessment_filter_enabled'])
        self.assertContains(response, 'data-summary-category="subjects"')
        self.assertNotContains(response, 'data-journal-category="assessments"')

    def test_teacher_with_assigned_work_sees_assessment_category(self):
        data = self.create_base_journal()
        assessment_subject = Subject.objects.create(
            name='Оркестр для личного кабинета',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
        )
        GroupSubject.objects.create(
            group=data['group'],
            subject=assessment_subject,
            teacher=data['teacher'],
        )
        assessment_group = AssessmentGroup.objects.create(
            name='Назначенная программа',
            subject=assessment_subject,
            academic_year=data['year'],
        )
        AssessmentItem.objects.create(
            title='Назначенное произведение',
            subject=assessment_subject,
            academic_year=data['year'],
            group=assessment_group,
            responsible_teacher=data['teacher'],
        )
        self.client.force_login(data['teacher'].user)

        response = self.client.get(
            reverse('journal'),
            {'academic_year': data['year'].pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['assessment_filter_enabled'])
        self.assertContains(response, 'data-journal-category="assessments"')


class AcademicYearJournalAccessTests(JournalTestDataMixin, TestCase):
    def test_student_can_select_only_years_with_enrollment(self):
        old_year = self.create_academic_year(name='2025/2026')
        old_group = self.create_group(academic_year=old_year)
        student = self.create_student(group=old_group)
        new_year = self.create_academic_year(name='2026/2027')
        self.client.force_login(student.user)

        response = self.client.get(reverse('journal'), {'academic_year': new_year.pk})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['selected_academic_year'], old_year)
        self.assertEqual(list(response.context['academic_years']), [old_year])

    def test_later_application_adds_existing_student_year_without_duplicate_account(self):
        first_year = self.create_academic_year(name='2025/2026')
        first_application = CourseApplication.objects.create(**self.application_payload())
        second_year = self.create_academic_year(name='2026/2027')
        second_application = CourseApplication.objects.create(
            **self.application_payload(student_phone='+7 (999) 765-43-21'),
        )
        TemporaryCredential.objects.filter(user=first_application.user).delete()
        self.client.force_login(first_application.user)

        response = self.client.get(reverse('journal'), {'academic_year': first_year.pk})
        available_ids = set(response.context['academic_years'].values_list('pk', flat=True))

        self.assertEqual(first_application.student_id, second_application.student_id)
        self.assertEqual(first_application.user_id, second_application.user_id)
        self.assertEqual(available_ids, {first_year.pk, second_year.pk})
        self.assertEqual(response.context['selected_academic_year'], first_year)

    def test_teacher_sees_only_membership_years(self):
        old_year = self.create_academic_year(name='2025/2026')
        teacher = self.create_teacher()
        new_year = self.create_academic_year(name='2026/2027')
        self.client.force_login(teacher.user)

        response = self.client.get(reverse('journal'), {'academic_year': new_year.pk})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['selected_academic_year'], old_year)
        self.assertEqual(list(response.context['academic_years']), [old_year])


    def test_inactive_teacher_membership_remains_viewable_but_not_editable(self):
        year = self.create_academic_year()
        teacher = self.create_teacher()
        membership = TeacherEnrollment.objects.get(teacher=teacher, academic_year=year)
        membership.is_active = False
        membership.save()
        teacher.refresh_from_db()
        self.client.force_login(teacher.user)

        response = self.client.get(reverse('journal'), {'academic_year': year.pk})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['selected_academic_year'], year)
        self.assertFalse(response.context['can_edit_journal'])


    def test_inactive_teacher_cannot_bypass_read_only_mode_with_post(self):
        data = self.create_base_journal()
        grade = Grade.objects.create(
            student=data['student'],
            subject=data['solfeggio'],
            teacher=data['teacher'],
            academic_year=data['year'],
            date=date(2025, 10, 15),
            value='5',
        )
        membership = TeacherEnrollment.objects.get(
            teacher=data['teacher'],
            academic_year=data['year'],
        )
        membership.is_active = False
        membership.save(update_fields=['is_active'])
        self.client.force_login(data['teacher'].user)

        response = self.client.post(
            (
                f'{reverse("journal")}?group={data["group"].pk}'
                f'&subject={data["solfeggio"].pk}'
                f'&academic_year={data["year"].pk}'
            ),
            data={
                'action': 'inline_edit',
                (
                    f'grade__{data["solfeggio"].pk}__'
                    f'{data["student"].pk}__2025-10-15'
                ): '2',
            },
        )

        self.assertEqual(response.status_code, 302)
        grade.refresh_from_db()
        self.assertEqual(grade.value, '5')


class AcademicYearAdminContextTests(JournalTestDataMixin, TestCase):
    def setUp(self):
        self.superuser = User.objects.create_superuser(
            username='year_admin',
            email='admin@example.com',
            password='AdminPass123!',
        )
        self.factory = RequestFactory()

    def admin_request(self, academic_year):
        request = self.factory.get('/admin/', {'academic_year': academic_year.pk})
        request.user = self.superuser
        request.session = {}
        return request

    def test_administrator_can_be_assigned_to_multiple_academic_years(self):
        old_year = self.create_academic_year(name='2025/2026')
        new_year = self.create_academic_year(name='2026/2027')
        UserAcademicYearMembership.objects.bulk_create([
            UserAcademicYearMembership(user=self.superuser, academic_year=old_year),
            UserAcademicYearMembership(user=self.superuser, academic_year=new_year),
        ])

        self.assertEqual(
            set(academic_year_ids_for_user(self.superuser)),
            {old_year.pk, new_year.pk},
        )

    def test_administrator_sees_only_manually_assigned_academic_years(self):
        assigned_year = self.create_academic_year(name='2025/2026')
        hidden_year = self.create_academic_year(name='2026/2027')
        UserAcademicYearMembership.objects.create(
            user=self.superuser,
            academic_year=assigned_year,
        )
        request = self.admin_request(hidden_year)

        context = get_admin_academic_year_context(request)

        self.assertEqual(context['admin_selected_academic_year'], assigned_year)
        self.assertEqual(list(context['admin_academic_years']), [assigned_year])

    def test_teacher_enrollments_create_memberships_for_every_academic_year(self):
        old_year = self.create_academic_year(name='2025/2026')
        teacher = self.create_teacher(username='multi_year_teacher')
        self.assertTrue(
            TeacherEnrollment.objects.filter(teacher=teacher, academic_year=old_year).exists()
        )
        new_year = self.create_academic_year(name='2026/2027')
        TeacherEnrollment.objects.create(teacher=teacher, academic_year=new_year)

        self.assertEqual(
            set(academic_year_ids_for_user(teacher.user)),
            {old_year.pk, new_year.pk},
        )

    def test_admin_save_keeps_the_same_change_page(self):
        self.create_academic_year()
        contact = PasswordRecoveryContact.objects.create(
            name='Администратор',
            phone='+7 (999) 123-45-67',
            messengers='Telegram',
        )
        self.client.force_login(self.superuser)
        change_url = reverse(
            'admin:journal_passwordrecoverycontact_change',
            args=[contact.pk],
        )

        response = self.client.post(change_url, data={
            'name': 'Администратор',
            'phone': '+7 (999) 123-45-67',
            'messengers': 'Telegram, MAX',
            'is_active': 'on',
            'display_order': '0',
            '_save': 'Сохранить',
        })

        self.assertRedirects(response, change_url, fetch_redirect_response=False)

    def test_student_queryset_contains_only_people_from_selected_year(self):
        old_year = self.create_academic_year(name='2025/2026')
        old_group = self.create_group(name='Старая группа', academic_year=old_year)
        instrument = self.create_instrument()
        old_student = self.create_student(
            full_name='Старый Ученик',
            group=old_group,
            instrument=instrument,
            username='old_student',
        )
        new_year = self.create_academic_year(name='2026/2027')
        new_group = self.create_group(name='Новая группа', academic_year=new_year)
        new_student = self.create_student(
            full_name='Сидоров Семён Сергеевич',
            group=new_group,
            instrument=instrument,
            username='new_student',
        )

        model_admin = StudentAdmin(Student, django_admin.site)
        queryset = model_admin.get_queryset(self.admin_request(old_year))

        self.assertIn(old_student, queryset)
        self.assertNotIn(new_student, queryset)

    def test_archived_year_blocks_edits_but_allows_explicit_deletion(self):
        old_year = self.create_academic_year(name='2025/2026')
        old_group = self.create_group(academic_year=old_year)
        instrument = self.create_instrument()
        student = self.create_student(group=old_group, instrument=instrument)
        self.create_academic_year(name='2026/2027')
        request = self.admin_request(old_year)

        group_admin = StudyGroupAdmin(StudyGroup, django_admin.site)
        student_admin = StudentAdmin(Student, django_admin.site)
        temporary_admin = TemporaryCredentialAdmin(TemporaryCredential, django_admin.site)
        student_inline = StudentInline(StudyGroup, django_admin.site)

        self.assertFalse(group_admin.has_add_permission(request))
        self.assertFalse(group_admin.has_change_permission(request, old_group))
        self.assertTrue(group_admin.has_delete_permission(request, old_group))
        self.assertTrue(student_admin.has_change_permission(request, student))
        self.assertFalse(student_admin.has_add_permission(request))
        self.assertTrue(student_admin.has_delete_permission(request, student))
        self.assertFalse(temporary_admin.has_add_permission(request))
        self.assertFalse(temporary_admin.has_change_permission(request))
        self.assertFalse(student_inline.has_add_permission(request, old_group))
        self.assertFalse(student_inline.has_change_permission(request, old_group))
        self.assertFalse(student_inline.has_delete_permission(request, old_group))

    def test_archived_student_form_keeps_profile_tabs_but_replaces_year_fields(self):
        old_year = self.create_academic_year(name='2025/2026')
        old_group = self.create_group(academic_year=old_year)
        student = self.create_student(group=old_group)
        self.create_academic_year(name='2026/2027')
        model_admin = StudentAdmin(Student, django_admin.site)

        fieldsets = model_admin.get_fieldsets(self.admin_request(old_year), student)
        flattened_fields = {
            field
            for _title, options in fieldsets
            for field in options['fields']
        }

        self.assertIn('full_name', flattened_fields)
        self.assertIn('city_church', flattened_fields)
        self.assertIn('user', flattened_fields)
        self.assertIn('selected_year_group_display', flattened_fields)
        self.assertIn('selected_year_active_display', flattened_fields)
        self.assertNotIn('group', flattened_fields)
        self.assertNotIn('is_active', flattened_fields)

    def test_archived_course_application_change_page_does_not_raise_key_error(self):
        old_year = self.create_academic_year(name='2025/2026')
        application = CourseApplication.objects.create(**self.application_payload())
        self.create_academic_year(name='2026/2027')
        self.client.force_login(self.superuser)

        response = self.client.get(
            reverse('admin:journal_courseapplication_change', args=[application.pk]),
            {'academic_year': old_year.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, application.full_name)

    def test_student_change_page_renders_assessment_group_inline_without_key_error(self):
        year = self.create_academic_year()
        study_group = self.create_group(academic_year=year)
        student = self.create_student(group=study_group)
        teacher = self.create_teacher(username='assessment_teacher')
        subject = Subject.objects.create(
            name='Оркестр',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )
        GroupSubject.objects.create(
            group=study_group,
            subject=subject,
            teacher=teacher,
        )
        assessment_group = AssessmentGroup.objects.create(
            name='Старший оркестр',
            subject=subject,
            academic_year=year,
        )
        self.client.force_login(self.superuser)

        response = self.client.get(
            reverse('admin:journal_student_change', args=[student.pk]),
            {'academic_year': year.pk},
        )

        self.assertEqual(response.status_code, 200)

        inline = StudentAssessmentGroupInline(Student, django_admin.site)
        formset = inline.get_formset(self.admin_request(year), student)(instance=student)
        self.assertIn('assessment_group', formset.empty_form.fields)
        self.assertIn(
            assessment_group,
            formset.empty_form.fields['assessment_group'].queryset,
        )

    def test_group_student_inline_exposes_related_student_controls_and_card_link(self):
        year = self.create_academic_year()
        group = self.create_group(academic_year=year)
        student = self.create_student(group=group)
        enrollment = StudentEnrollment.objects.get(student=student, academic_year=year)
        request = self.admin_request(year)
        inline = StudentInline(StudyGroup, django_admin.site)
        formset_class = inline.get_formset(request, group)
        form = formset_class(instance=group).forms[0]

        self.assertIn('student_card_link', inline.fields)
        self.assertTrue(form.fields['student'].widget.can_add_related)
        self.assertTrue(form.fields['student'].widget.can_change_related)
        # Django intentionally hides the related-object delete icon for a
        # CASCADE relation. The student remains deletable from its own table,
        # where the complete cascade is shown before confirmation.
        self.assertFalse(form.fields['student'].widget.can_delete_related)
        self.assertIn(student.full_name, str(inline.student_card_link(enrollment)))

    def test_active_group_can_enroll_student_from_previous_year(self):
        old_year = self.create_academic_year(name='2025/2026')
        old_group = self.create_group(academic_year=old_year)
        student = self.create_student(group=old_group)
        new_year = self.create_academic_year(name='2026/2027')
        new_group = self.create_group(name='Новая группа', academic_year=new_year)
        student.refresh_from_db()
        self.assertIsNone(student.group_id)
        self.assertFalse(student.is_active)

        request = self.admin_request(new_year)
        inline = StudentInline(StudyGroup, django_admin.site)
        formset = inline.get_formset(request, new_group)(instance=new_group)
        enrollment = formset._move_student_enrollment(student, commit=True)

        student.refresh_from_db()
        self.assertEqual(enrollment.academic_year, new_year)
        self.assertEqual(enrollment.group, new_group)
        self.assertTrue(enrollment.is_active)
        self.assertEqual(student.group, new_group)
        self.assertTrue(student.is_active)

    def test_old_teacher_can_be_reused_in_active_year_assignment(self):
        self.create_academic_year(name='2025/2026')
        teacher = self.create_teacher()
        new_year = self.create_academic_year(name='2026/2027')
        new_group = self.create_group(academic_year=new_year)
        subject = self.create_subject()
        teacher.refresh_from_db()
        self.assertFalse(teacher.is_active)
        self.assertIn(teacher, assignment_teacher_queryset())

        GroupSubject.objects.create(group=new_group, subject=subject, teacher=teacher)

        teacher.refresh_from_db()
        self.assertTrue(teacher.is_active)
        self.assertTrue(
            TeacherEnrollment.objects.filter(
                teacher=teacher,
                academic_year=new_year,
                is_active=True,
            ).exists(),
        )

    def test_every_non_global_admin_disables_add_but_allows_delete_in_archive_mode(self):
        old_year = self.create_academic_year(name='2025/2026')
        self.create_academic_year(name='2026/2027')
        request = self.admin_request(old_year)
        global_models = {PasswordRecoveryContact}

        for model, model_admin in django_admin.site._registry.items():
            if model in global_models:
                continue
            if model._meta.app_label not in {'journal', 'auth'}:
                continue
            with self.subTest(model=model._meta.label):
                self.assertFalse(model_admin.has_add_permission(request))
                self.assertTrue(model_admin.has_delete_permission(request))

    def test_registration_settings_are_read_only_and_recovery_settings_editable_in_archive_mode(self):
        old_year = self.create_academic_year(name='2025/2026')
        self.create_academic_year(name='2026/2027')
        request = self.admin_request(old_year)
        registration_admin = CourseRegistrationSettingsAdmin(
            CourseRegistrationSettings,
            django_admin.site,
        )
        recovery_admin = PasswordRecoveryContactAdmin(
            PasswordRecoveryContact,
            django_admin.site,
        )

        self.assertFalse(registration_admin.has_change_permission(request))
        self.assertTrue(recovery_admin.has_add_permission(request))

    def test_inactive_academic_year_is_not_editable_but_is_deletable(self):
        old_year = self.create_academic_year(name='2025/2026')
        active_year = self.create_academic_year(name='2026/2027')
        old_year.refresh_from_db()
        request = self.admin_request(active_year)
        year_admin = AcademicYearAdmin(AcademicYear, django_admin.site)

        self.assertFalse(year_admin.has_change_permission(request, old_year))
        self.assertTrue(year_admin.has_delete_permission(request, old_year))

    def test_city_church_fields_are_wide_in_admin_and_registration_forms(self):
        year = self.create_academic_year()
        group = self.create_group(academic_year=year)
        student_form = StudentAdminForm()
        application_form = CourseApplicationPublicForm()
        inline = StudentInline(StudyGroup, django_admin.site)
        inline_form = inline.get_formset(self.admin_request(year), group)(instance=group).empty_form

        self.assertIn('city-church-field', student_form.fields['city_church'].widget.attrs.get('class', ''))
        self.assertEqual(application_form.fields['city_church'].widget.attrs.get('size'), '80')
        self.assertEqual(inline_form.fields['city_church'].widget.attrs.get('size'), '80')

        css = Path('journal/static/journal/admin_dashboard.css').read_text(encoding='utf-8')
        javascript = Path('journal/static/journal/admin_responsive.js').read_text(encoding='utf-8')
        self.assertIn('@import url("layout-tablet.css");', css)
        self.assertIn('@import url("layout-mobile.css");', css)
        self.assertIn('@media (max-width: 767.98px)', css)
        self.assertIn('responsive-admin-modal', css)
        self.assertIn('body.change-form .field-city_church', css)
        self.assertIn('min-width: 0 !important;', css)
        self.assertIn('window.open', javascript)
        self.assertIn('matchMedia', javascript)
        self.assertIn('journal-admin-form-state', javascript)
        self.assertIn('scrollY', javascript)
        self.assertIn('scrollLeft', javascript)
        self.assertIn('activeTab', javascript)
        self.assertIn('#content-main form[method="post"]', javascript)
        self.assertIn("state.path.endsWith('/add/')", javascript)
        self.assertIn('showAdminSaveNotification', javascript)
        self.assertIn('adminValidationMessage', javascript)
        self.assertIn('revealAdminError', javascript)
        self.assertIn('Не удалось сохранить запись', javascript)
        self.assertIn('journal-save-toast journal-save-toast--', javascript)
        self.assertIn("isSuccess ? 'success' : 'error'", javascript)
        self.assertIn('.journal-save-toast--success', css)
        self.assertIn('.journal-save-toast--error', css)

        admin_source = Path('journal/admin.py').read_text(encoding='utf-8')
        self.assertNotRegex(admin_source, r'extra\s*=\s*[1-9]')
        change_form_template = Path(
            'templates/admin/change_form.html'
        ).read_text(encoding='utf-8')
        self.assertIn('точная причина указана рядом с каждым полем', change_form_template)
        self.assertNotIn('Please correct the errors below', change_form_template)

        form_state = Path(
            'journal/static/journal/form_state.js'
        ).read_text(encoding='utf-8')
        self.assertIn('journal-form-state:', form_state)
        self.assertIn("querySelectorAll('.table-scroll')", form_state)
        self.assertIn('scrollers', form_state)
        self.assertIn('data-flash-message', form_state)
        self.assertIn('save-toast save-toast--', form_state)
        self.assertIn("isSuccess ? 'success' : 'error'", form_state)
        self.assertIn('main form[data-preserve-scroll]', form_state)
        self.assertIn('main a[data-preserve-scroll]', form_state)
        self.assertIn('data-filter-auto-submit', form_state)

        journal_template = Path('templates/journal.html').read_text(encoding='utf-8')
        self.assertIn('data-save-context="journal-filters"', journal_template)
        self.assertIn('data-save-context="assessment-filters"', journal_template)
        self.assertNotIn('onchange="this.form.submit()"', journal_template)

        mobile_css = Path('journal/static/journal/layout-mobile.css').read_text(encoding='utf-8')
        self.assertIn('.filter-form > *', mobile_css)
        self.assertIn('.grade-form > *', mobile_css)
        self.assertIn('.field > .select-search-input', mobile_css)
        self.assertIn('min-height: 100dvh;', mobile_css)
        self.assertIn('env(safe-area-inset-bottom)', mobile_css)
        self.assertIn('(max-width: 767.98px) and (orientation: portrait)', mobile_css)
        self.assertIn('grid-template-columns: repeat(2, minmax(0, 1fr)) !important;', mobile_css)
        self.assertIn('body.change-form .submit-row', mobile_css)
        self.assertIn('.related-widget-wrapper-link', mobile_css)
        self.assertNotIn('max-height: 500px', mobile_css)
        self.assertIn(
            '(max-width: 1023.98px) and (orientation: landscape) and (pointer: coarse)',
            mobile_css,
        )

        device_styles = Path('templates/journal/device_styles.html').read_text(encoding='utf-8')
        self.assertNotIn(' media="', device_styles)
        self.assertEqual(device_styles.count('data-layout-styles='), 5)
        self.assertNotIn('?v=', device_styles)
        self.assertIn('journal/responsive_overflow.css', device_styles)

        admin_base_template = Path(
            'templates/admin/base_site.html'
        ).read_text(encoding='utf-8')
        self.assertIn('journal/responsive_overflow.css', admin_base_template)
        self.assertIn('journal/responsive_tables.js', admin_base_template)
        self.assertIn('journal/responsive_tables.js', journal_template)

        tablet_css = Path('journal/static/journal/layout-tablet.css').read_text(encoding='utf-8')
        self.assertIn('(min-width: 768px) and (max-width: 1023.98px)', tablet_css)
        self.assertIn('grid-template-columns: minmax(0, 1fr) !important;', tablet_css)
        self.assertIn('.filter-form > *', tablet_css)
        self.assertIn('(orientation: portrait)', tablet_css)
        self.assertIn('(orientation: landscape)', tablet_css)
        self.assertIn('repeat(3, minmax(0, 1fr))', tablet_css)
        self.assertIn('env(safe-area-inset-bottom)', tablet_css)

        overflow_css = Path(
            'journal/static/journal/responsive_overflow.css'
        ).read_text(encoding='utf-8')
        self.assertIn('overflow-x: auto !important;', overflow_css)
        self.assertIn('overflow-y: auto !important;', overflow_css)
        self.assertIn('touch-action: pan-x pan-y;', overflow_css)
        self.assertIn('max-height: 72dvh !important;', overflow_css)
        self.assertNotIn('font-size:', overflow_css)
        self.assertNotIn('padding:', overflow_css)

        responsive_tables = Path(
            'journal/static/journal/responsive_tables.js'
        ).read_text(encoding='utf-8')
        self.assertIn("querySelectorAll('table')", responsive_tables)
        self.assertIn("overflow-y", overflow_css)
        self.assertIn('MutationObserver', responsive_tables)
        self.assertIn("document.addEventListener('formset:added'", responsive_tables)
        self.assertIn("role', 'region'", responsive_tables)

    def test_all_project_admins_keep_the_current_form_or_list_after_save(self):
        for model, model_admin in django_admin.site._registry.items():
            if model._meta.app_label not in {'journal', 'auth'}:
                continue
            with self.subTest(model=model._meta.label):
                self.assertIsInstance(model_admin, JournalAdminDescriptionMixin)

        default_save = RequestFactory().post('/admin/example/', {'_save': 'Сохранить'})
        explicit_navigation = RequestFactory().post(
            '/admin/example/',
            {'_addanother': 'Сохранить и добавить другой объект'},
        )
        self.assertTrue(JournalAdminDescriptionMixin._keep_change_form_open(default_save))
        self.assertFalse(JournalAdminDescriptionMixin._keep_change_form_open(explicit_navigation))

    def test_archived_admin_lists_use_assignment_snapshots(self):
        old_year = self.create_academic_year(name='2025/2026')
        old_group = self.create_group(academic_year=old_year)
        teacher = self.create_teacher(full_name='Старое имя преподавателя')
        student = self.create_student(group=old_group)
        group_subject = GroupSubject.objects.create(
            group=old_group,
            subject=self.create_subject(name='Старое название предмета'),
            teacher=teacher,
        )
        specialty_subject = self.create_subject(
            name='Старая специальность',
            is_specialty=True,
        )
        specialty = StudentSubject.objects.create(
            student=student,
            subject=specialty_subject,
            teacher=teacher,
            academic_year=old_year,
        )
        self.create_academic_year(name='2026/2027')

        group_subject.subject.name = 'Новое название предмета'
        group_subject.subject.save()
        specialty_subject.name = 'Новая специальность'
        specialty_subject.save()
        teacher.full_name = 'Новое имя преподавателя'
        teacher.save()

        group_admin = StudyGroupAdmin(StudyGroup, django_admin.site)
        teacher_admin = TeacherAdmin(Teacher, django_admin.site)
        student_admin = StudentAdmin(Student, django_admin.site)
        group = group_admin.get_queryset(self.admin_request(old_year)).get(pk=old_group.pk)
        selected_teacher = teacher_admin.get_queryset(self.admin_request(old_year)).get(pk=teacher.pk)
        selected_student = student_admin.get_queryset(self.admin_request(old_year)).get(pk=student.pk)

        self.assertIn('Старое название предмета', str(group_admin.subjects_display_short(group)))
        self.assertIn('Старое имя преподавателя', str(group_admin.teachers_display_short(group)))
        self.assertIn('Старое название предмета', str(teacher_admin.group_subjects_short(selected_teacher)))
        self.assertEqual(
            str(student_admin.specialty_subject_display(selected_student)),
            specialty.subject_name_snapshot,
        )
        self.assertEqual(
            str(student_admin.specialty_teacher_display(selected_student)),
            specialty.teacher_name_snapshot,
        )

    def test_repeat_student_temporary_credential_belongs_only_to_creation_year(self):
        old_year = self.create_academic_year(name='2025/2026')
        first_application = CourseApplication.objects.create(**self.application_payload())
        active_year = self.create_academic_year(name='2026/2027')
        CourseApplication.objects.create(
            **self.application_payload(student_phone='+7 (999) 765-43-21'),
        )
        credential = TemporaryCredential.objects.get(user_id=first_application.user_id)

        old_credentials = filter_temporary_credentials_for_year(
            TemporaryCredential.objects.all(),
            old_year,
        )
        active_credentials = filter_temporary_credentials_for_year(
            TemporaryCredential.objects.all(),
            active_year,
        )

        self.assertIn(credential, old_credentials)
        self.assertNotIn(credential, active_credentials)

    def test_temporary_credentials_are_scoped_to_selected_academic_year(self):
        old_year = self.create_academic_year(name='2025/2026')
        old_application = CourseApplication.objects.create(**self.application_payload())
        active_year = self.create_academic_year(name='2026/2027')
        active_application = CourseApplication.objects.create(
            **self.application_payload(
                last_name='Петров',
                student_phone='+7 (999) 765-43-21',
            ),
        )
        staff_user = User.objects.create_user(
            username='yearless_staff',
            password='Pass12345!',
            is_staff=True,
        )
        staff_credential = TemporaryCredential.objects.create(
            user=staff_user,
            login=staff_user.username,
            temporary_password='StaffTemp123!',
        )

        old_credentials = filter_temporary_credentials_for_year(
            TemporaryCredential.objects.all(),
            old_year,
        )
        active_credentials = filter_temporary_credentials_for_year(
            TemporaryCredential.objects.all(),
            active_year,
        )

        self.assertIn(old_application.temporary_credential, old_credentials)
        self.assertNotIn(active_application.temporary_credential, old_credentials)
        self.assertNotIn(staff_credential, old_credentials)
        self.assertIn(active_application.temporary_credential, active_credentials)
        self.assertIn(staff_credential, active_credentials)


class AdminDashboardTests(JournalTestDataMixin, TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            username='dashboard_admin',
            password='Pass12345!',
            email='dashboard-admin@example.com',
        )
        self.superuser = self.admin_user

    def test_dashboard_reference_section_contains_orchestra_parts(self):
        self.client.force_login(self.superuser)

        response = self.client.get(reverse('admin:index'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Партии оркестра')
        self.assertContains(response, reverse('admin:journal_orchestrapart_changelist'))

    def test_student_change_card_uses_bounded_query_count(self):
        data = self.create_base_journal()
        self.client.force_login(self.admin_user)

        with CaptureQueriesContext(connection) as captured_queries:
            response = self.client.get(
                reverse('admin:journal_student_change', args=[data['student'].pk]),
                {'academic_year': data['year'].pk},
            )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'journal/select_search.js')
        self.assertNotContains(response, 'data-searchable-select')
        self.assertNotContains(response, 'grades-TOTAL_FORMS')
        self.assertLessEqual(
            len(captured_queries),
            85,
            'Карточка ученика снова выполняет слишком много SQL-запросов.',
        )

    def test_student_changelist_does_not_issue_queries_per_student(self):
        data = self.create_base_journal()
        for index in range(12):
            self.create_student(
                full_name=f'Ученик Производительности {index:02d}',
                group=data['group'],
                instrument=data['instrument'],
                username=f'performance_student_{index:02d}',
            )
        self.client.force_login(self.admin_user)

        with CaptureQueriesContext(connection) as captured_queries:
            response = self.client.get(reverse('admin:journal_student_changelist'))

        self.assertEqual(response.status_code, 200)
        self.assertLessEqual(
            len(captured_queries),
            30,
            [query['sql'] for query in captured_queries],
        )

    def test_archived_academic_year_records_are_read_only_in_admin(self):
        data = self.create_base_journal()
        AcademicYear.objects.create(
            name='2026/2027',
            starts_on=date(2026, 9, 1),
            ends_on=date(2027, 8, 31),
        )
        data['group'].refresh_from_db()

        request = type('Request', (), {'user': self.admin_user})()
        model_admin = django_admin.site._registry[StudyGroup]

        self.assertFalse(model_admin.has_change_permission(request, data['group']))
        self.assertTrue(model_admin.has_delete_permission(request, data['group']))

    def test_active_models_have_activate_and_deactivate_actions(self):
        subject = self.create_subject(name='Предмет для смены статуса')
        self.client.force_login(self.admin_user)
        changelist_url = reverse('admin:journal_subject_changelist')

        response = self.client.post(changelist_url, {
            'action': 'deactivate_selected_records',
            '_selected_action': [subject.pk],
        })
        self.assertEqual(response.status_code, 302)
        subject.refresh_from_db()
        self.assertFalse(subject.is_active)

        response = self.client.post(changelist_url, {
            'action': 'activate_selected_records',
            '_selected_action': [subject.pk],
        })
        self.assertEqual(response.status_code, 302)
        subject.refresh_from_db()
        self.assertTrue(subject.is_active)

    def test_archived_group_page_uses_enrollment_and_assignment_snapshots(self):
        data = self.create_base_journal()
        archived_student_name = data['student'].full_name
        archived_subject_name = data['solfeggio'].name
        AcademicYear.objects.create(
            name='2026/2027',
            starts_on=date(2026, 9, 1),
            ends_on=date(2027, 8, 31),
        )
        data['student'].refresh_from_db()
        data['student'].full_name = 'Текущее имя ученика'
        data['student'].save()
        data['solfeggio'].name = 'Текущее название предмета'
        data['solfeggio'].save()
        self.client.login(username='dashboard_admin', password='Pass12345!')

        response = self.client.get(
            reverse('admin:journal_studygroup_change', args=[data['group'].pk]),
            {'academic_year': data['year'].pk},
        )

        self.assertContains(response, archived_student_name)
        self.assertContains(response, archived_subject_name)
        self.assertNotContains(response, 'name="student_enrollments-0-student"')
        self.assertNotContains(response, 'name="group_subjects-0-subject"')

    def test_admin_dashboard_counts_only_subjects_assigned_in_selected_year(self):
        data = self.create_base_journal()
        self.create_subject(name='Не назначенный предмет')
        request = RequestFactory().get('/admin/')
        request.user = self.admin_user
        request.session = {
            'journal_admin_academic_year_id': data['year'].pk,
        }

        dashboard = journal_admin_dashboard({'request': request})
        subject_stat = next(
            stat for stat in dashboard['stats']
            if stat['label'] == 'Предметы выбранного года'
        )

        self.assertEqual(subject_stat['value'], 3)

    def test_admin_dashboard_links_recovery_settings_and_related_data(self):
        self.create_base_journal()
        PasswordRecoveryContact.objects.create(
            name='Администратор',
            phone='+7 (999) 123-45-67',
            messengers='Telegram',
        )
        self.client.login(username='dashboard_admin', password='Pass12345!')

        response = self.client.get(reverse('admin:index'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Настройки восстановления')
        self.assertContains(response, reverse('admin:journal_passwordrecoverycontact_changelist'))
        self.assertContains(response, 'Связанные данные')
        self.assertContains(response, 'Групповые предметы')
        self.assertContains(response, reverse('admin:journal_groupsubject_changelist'))
        self.assertContains(response, 'Индивидуальные предметы')
        self.assertContains(response, reverse('admin:journal_studentsubject_changelist'))
        self.assertContains(response, 'Квалификации преподавателей')
        self.assertContains(response, reverse('admin:journal_teachersubject_changelist'))
        self.assertContains(response, 'Инструкция')
        self.assertContains(response, reverse('admin_guide'))

    def test_admin_dashboard_shows_assessment_data_and_selected_year_result_count(self):
        data = self.create_base_journal()
        assessment_subject = Subject.objects.create(
            name='Оркестр',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )
        GroupSubject.objects.create(
            group=data['group'],
            subject=assessment_subject,
            teacher=data['teacher'],
        )
        assessment_group = AssessmentGroup.objects.create(
            name='Оркестровые партии',
            subject=assessment_subject,
            academic_year=data['year'],
        )
        item = AssessmentItem.objects.create(
            title='Первая партия',
            subject=assessment_subject,
            academic_year=data['year'],
            group=assessment_group,
            responsible_teacher=data['teacher'],
        )
        assignment = StudentAssessmentGroup.objects.create(
            student=data['student'],
            assessment_group=assessment_group,
            academic_year=data['year'],
        )
        AssessmentResult.objects.create(
            enrollment=assignment.enrollment,
            item=item,
            status=AssessmentResult.STATUS_PASSED,
            assessed_by=data['teacher'],
        )
        request = RequestFactory().get('/admin/')
        request.user = self.admin_user
        request.session = {'journal_admin_academic_year_id': data['year'].pk}

        dashboard = journal_admin_dashboard({'request': request})
        result_stat = next(
            stat for stat in dashboard['stats']
            if stat['label'] == 'Результаты сдачи'
        )

        self.assertEqual(result_stat['value'], 1)
        self.client.force_login(self.admin_user)
        response = self.client.get(reverse('admin:index'))
        self.assertContains(response, 'Сдача произведений')
        self.assertContains(response, reverse('admin:journal_assessmentgroup_changelist'))
        self.assertContains(response, reverse('admin:journal_assessmentitem_changelist'))
        self.assertContains(response, reverse('admin:journal_studentassessmentgroup_changelist'))
        self.assertContains(response, reverse('admin:journal_assessmentresult_changelist'))
        self.assertContains(response, reverse('admin:journal_finalgraderule_changelist'))

    def test_admin_sidebar_orders_tables_by_workflow(self):
        ordering = settings.JAZZMIN_SETTINGS['order_with_respect_to']
        journal_models = [
            item
            for item in ordering
            if item.startswith('journal.')
        ]

        self.assertEqual(
            journal_models,
            [
                'journal.StudyGroup',
                'journal.Student',
                'journal.Teacher',
                'journal.Subject',
                'journal.AcademicYear',
                'journal.Instrument',
                'journal.OrchestraPart',
                'journal.Grade',
                'journal.SubjectResult',
                'journal.AssessmentGroup',
                'journal.AssessmentItem',
                'journal.StudentAssessmentGroup',
                'journal.AssessmentResult',
                'journal.FinalGradeRule',
                'journal.GroupSubject',
                'journal.StudentSubject',
                'journal.TeacherSubject',
                'journal.CourseApplication',
                'journal.CourseRegistrationSettings',
                'journal.TemporaryCredential',
                'journal.ErrorLog',
                'journal.PasswordRecoveryContact',
            ],
        )

    def test_admin_copyright_uses_current_year_without_duplicate_text(self):
        copyright_name = settings.JAZZMIN_SETTINGS['copyright']

        self.assertEqual(copyright_name, 'Электронный журнал музыкальных курсов')
        self.assertNotIn('©', copyright_name)
        self.assertNotIn(str(date.today().year), copyright_name)
        self.assertNotIn('Все права защищены', copyright_name)
        self.assertEqual(
            f'Copyright © {date.today().year} {copyright_name}. Все права защищены.',
            (
                f'Copyright © {date.today().year} Электронный журнал '
                'музыкальных курсов. Все права защищены.'
            ),
        )

    def test_admin_guide_is_visible_only_for_superuser(self):
        User.objects.create_user(
            username='guide_staff',
            password='Pass12345!',
            is_staff=True,
        )

        self.client.login(username='guide_staff', password='Pass12345!')
        staff_response = self.client.get(reverse('admin_guide'))
        self.assertEqual(staff_response.status_code, 302)

        self.client.login(username='dashboard_admin', password='Pass12345!')
        admin_response = self.client.get(reverse('admin_guide'))

        self.assertEqual(admin_response.status_code, 200)
        self.assertContains(admin_response, 'Как работать с журналом')
        self.assertContains(admin_response, reverse('admin:journal_academicyear_changelist'))
        self.assertContains(admin_response, 'Архивный год предназначен для просмотра')
        self.assertContains(admin_response, reverse('admin:journal_student_changelist'))
        self.assertContains(admin_response, reverse('admin:journal_assessmentelement_changelist'))
        self.assertContains(admin_response, reverse('admin:journal_assessmentgroup_changelist'))
        self.assertContains(admin_response, reverse('admin:journal_errorlog_changelist'))
        self.assertContains(admin_response, reverse('admin_export_all_data_excel'))
        self.assertContains(admin_response, 'Любую таблицу можно прокручивать пальцем')
        self.assertContains(admin_response, 'два непересекающихся учебных периода по 14 дней')
        self.assertContains(admin_response, 'Одни и те же ученики зачисляются в оба периода')
        self.assertContains(admin_response, reverse('admin_data_tools'))

    def test_admin_changelist_add_button_is_ordered_before_search(self):
        css = Path('journal/static/journal/admin_dashboard.css').read_text(encoding='utf-8')

        self.assertIn('body.change-list #change-list-filters .object-tools', css)
        self.assertIn('order: 1;', css)
        self.assertIn('body.change-list #changelist-search', css)
        self.assertIn('order: 2;', css)

    def test_admin_changelists_show_table_descriptions(self):
        models = (
            User,
            Group,
            AcademicYear,
            Instrument,
            OrchestraPart,
            Subject,
            StudyGroup,
            Teacher,
            Student,
            TeacherSubject,
            GroupSubject,
            StudentSubject,
            Grade,
            SubjectResult,
            CourseApplication,
            TemporaryCredential,
            CourseRegistrationSettings,
            PasswordRecoveryContact,
        )

        for model in models:
            with self.subTest(model=model.__name__):
                model_admin = django_admin.site._registry[model]
                self.assertTrue(model_admin.changelist_description)
                self.assertEqual(
                    model_admin.change_list_template,
                    'admin/journal/change_list_with_description.html',
                )

        self.client.login(username='dashboard_admin', password='Pass12345!')
        response = self.client.get(reverse('admin:journal_studygroup_changelist'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, django_admin.site._registry[StudyGroup].changelist_description)
        self.assertContains(response, 'journal-changelist-description')

    def test_related_models_are_visible_in_admin_without_teacher_qualification_inlines(self):
        request = type('Request', (), {'user': self.admin_user})()

        for model in (GroupSubject, StudentSubject, TeacherSubject):
            with self.subTest(model=model.__name__):
                model_admin = django_admin.site._registry[model]
                self.assertTrue(model_admin.get_model_perms(request).get('view'))

        self.assertEqual(
            [inline.model for inline in django_admin.site._registry[Subject].inlines],
            [GroupSubject, StudentSubject, AssessmentGroup, FinalGradeRule],
        )
        self.assertEqual(
            [inline.model for inline in django_admin.site._registry[Teacher].inlines],
            [GroupSubject, StudentSubject, AssessmentItem],
        )
        self.assertEqual(
            [inline.model for inline in django_admin.site._registry[Instrument].inlines],
            [OrchestraPart],
        )

    def test_subject_admin_shows_assignment_inline_for_subject_type(self):
        group_subject = self.create_subject(name='Групповой предмет')
        individual_subject = self.create_subject(name='Индивидуальный предмет', is_specialty=True)
        self.client.login(username='dashboard_admin', password='Pass12345!')

        group_response = self.client.get(reverse('admin:journal_subject_change', args=[group_subject.pk]))
        self.assertContains(group_response, 'Индивидуальный предмет')
        self.assertContains(group_response, 'Группы, где есть этот предмет')
        self.assertNotContains(group_response, 'Индивидуальные ученики по этому предмету')

        individual_response = self.client.get(
            reverse('admin:journal_subject_change', args=[individual_subject.pk])
        )
        self.assertContains(individual_response, 'Индивидуальный предмет')
        self.assertContains(individual_response, 'Индивидуальные ученики по этому предмету')
        self.assertNotContains(individual_response, 'Группы, где есть этот предмет')

    def test_subject_change_page_renders_group_assignment_rows_and_related_controls(self):
        year = self.create_academic_year()
        group = self.create_group(name='Связанная группа из таблицы', academic_year=year)
        teacher = self.create_teacher(username='subject_inline_render_teacher')
        subject = self.create_subject(name='Предмет с группой в карточке')
        self.create_group_assignment(group=group, subject=subject, teacher=teacher)
        self.client.force_login(self.admin_user)

        response = self.client.get(
            reverse('admin:journal_subject_change', args=[subject.pk]),
            {'academic_year': year.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Группы, где есть этот предмет')
        self.assertContains(response, group.name)
        self.assertContains(response, 'name="group_subjects-TOTAL_FORMS"')
        self.assertContains(response, 'add-related')
        self.assertContains(response, 'change-related')
        self.assertContains(response, 'view-related')
        self.assertContains(response, 'delete-related')

    def test_subject_change_page_lists_all_groups_for_selected_subject(self):
        year = self.create_academic_year()
        subject = self.create_subject(name='Предмет для списка групп')
        teacher = self.create_teacher(username='subject_all_groups_teacher')
        groups = [
            self.create_group(name='Альфа-группа', academic_year=year),
            self.create_group(name='Бета-группа', academic_year=year),
        ]
        for group in groups:
            self.create_group_assignment(
                group=group,
                subject=subject,
                teacher=teacher,
            )
        self.client.force_login(self.admin_user)

        response = self.client.get(
            reverse('admin:journal_subject_change', args=[subject.pk]),
            {'academic_year': year.pk},
        )

        self.assertEqual(response.status_code, 200)
        for group in groups:
            self.assertContains(response, group.name)
        self.assertContains(response, 'Добавить еще один')
        self.assertNotContains(response, 'inline-group collapse')

    def test_subject_change_page_saves_manually_added_group_relation_once(self):
        year = self.create_academic_year()
        group = self.create_group(name='Группа для ручной связи', academic_year=year)
        teacher = self.create_teacher(username='subject_inline_post_teacher')
        subject = self.create_subject(name='Предмет для ручной связи')
        TeacherSubject.objects.create(teacher=teacher, subject=subject)
        self.client.force_login(self.admin_user)
        url = reverse('admin:journal_subject_change', args=[subject.pk])

        get_response = self.client.get(url, {'academic_year': year.pk})

        self.assertEqual(get_response.status_code, 200)
        self.assertEqual(
            get_response.content.count(b'name="group_subjects-TOTAL_FORMS"'),
            1,
        )

        response = self.client.post(
            f'{url}?academic_year={year.pk}',
            {
                'name': subject.name,
                'assessment_mode': Subject.ASSESSMENT_MODE_STANDARD,
                'final_grade_type': Subject.FINAL_GRADE_TYPE_NUMERIC,
                'is_active': 'on',
                'group_subjects-TOTAL_FORMS': '1',
                'group_subjects-INITIAL_FORMS': '0',
                'group_subjects-MIN_NUM_FORMS': '0',
                'group_subjects-MAX_NUM_FORMS': '1000',
                'group_subjects-0-group': str(group.pk),
                'group_subjects-0-teacher': str(teacher.pk),
                'group_subjects-0-sort_order': '10',
                'group_subjects-0-is_active': 'on',
                '_save': 'Сохранить',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            GroupSubject.objects.filter(
                group=group,
                subject=subject,
                teacher=teacher,
            ).exists(),
        )

    def test_subject_change_page_exposes_non_inline_reverse_relations(self):
        data = self.create_base_journal()
        Grade.objects.create(
            student=data['student'],
            subject=data['solfeggio'],
            teacher=data['teacher'],
            academic_year=data['year'],
            date=date(2025, 10, 10),
            value='5',
        )
        self.client.force_login(self.admin_user)

        response = self.client.get(
            reverse('admin:journal_subject_change', args=[data['solfeggio'].pk]),
            {'academic_year': data['year'].pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Связанные данные')
        self.assertContains(response, 'Оценки')
        self.assertContains(response, 'Просмотреть / редактировать')
        self.assertContains(response, 'Удалить')
        self.assertContains(response, 'subject__id__exact')

    def test_subject_group_inline_exposes_full_related_group_controls(self):
        year = self.create_academic_year()
        group = self.create_group(name='Связанная группа', academic_year=year)
        form = GroupSubjectForSubjectAdminForm()
        group_field = form.fields['group']
        widget = group_field.widget

        self.assertEqual(list(group_field.queryset), [group])
        self.assertEqual(widget.__class__.__name__, 'RelatedFieldWidgetWrapper')
        self.assertTrue(widget.can_add_related)
        self.assertTrue(widget.can_change_related)
        self.assertTrue(widget.can_delete_related)
        self.assertTrue(widget.can_view_related)
        self.assertIn('удаляется флажком', group_field.help_text)

        inline = GroupSubjectForSubjectInline(Subject, django_admin.site)
        self.assertTrue(inline.can_delete)
        self.assertTrue(inline.show_change_link)
        self.assertEqual(inline.form, GroupSubjectForSubjectAdminForm)
        self.assertNotIn('collapse', inline.classes or ())
        self.assertEqual(inline.ordering, ('group__name', 'sort_order', 'pk'))

    def test_subject_admin_keeps_tabs_with_existing_related_data_visible(self):
        year = self.create_academic_year()
        group = self.create_group(academic_year=year)
        teacher = self.create_teacher(username='legacy_subject_tab_teacher')
        subject = self.create_subject(name='Предмет с сохранённой связью')
        self.create_group_assignment(group=group, subject=subject, teacher=teacher)
        Subject.objects.filter(pk=subject.pk).update(is_specialty=True)
        subject.refresh_from_db()

        request = RequestFactory().get('/admin/', {'academic_year': year.pk})
        request.user = self.admin_user
        request.session = {}
        subject_admin = django_admin.site._registry[Subject]
        assignment_inline_names = {
            inline.__name__ for inline in subject_admin.get_inlines(request, subject)
        }

        self.assertIn('GroupSubjectForSubjectInline', assignment_inline_names)
        self.assertIn('StudentSubjectForSubjectInline', assignment_inline_names)

        assessment_subject = Subject.objects.create(
            name='Предмет с сохранёнными произведениями',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )
        AssessmentGroup.objects.create(
            name='Сохранённая группа произведений',
            subject=assessment_subject,
            academic_year=year,
        )
        Subject.objects.filter(pk=assessment_subject.pk).update(
            assessment_mode=Subject.ASSESSMENT_MODE_STANDARD,
        )
        assessment_subject.refresh_from_db()

        assessment_inline_names = {
            inline.__name__
            for inline in subject_admin.get_inlines(request, assessment_subject)
        }
        self.assertIn('AssessmentGroupForSubjectInline', assessment_inline_names)
        self.assertIn('FinalGradeRuleForSubjectInline', assessment_inline_names)

    def test_student_admin_form_embeds_orchestra_parts_by_instrument(self):
        domra = self.create_instrument(name='Домра')
        bayan = self.create_instrument(name='Баян')
        domra_part = OrchestraPart.objects.create(
            instrument=domra,
            name='Малая первая',
        )
        bayan_part = OrchestraPart.objects.create(
            instrument=bayan,
            name='Первый',
        )

        form = StudentAdminForm(data={
            'instrument': domra.pk,
            'custom_instrument': '',
            'orchestra_part': domra_part.pk,
        })

        self.assertEqual(
            set(form.fields['orchestra_part'].queryset),
            {domra_part, bayan_part},
        )
        parts_map = json.loads(
            form.fields['orchestra_part'].widget.attrs['data-orchestra-parts-map'],
        )
        self.assertEqual(
            parts_map[str(domra.pk)],
            [{'id': domra_part.pk, 'name': domra_part.name}],
        )
        self.assertIn(
            'journal/orchestra_part_dependencies_v5.js',
            StudentAdminForm.Media.js,
        )

    def test_student_admin_instrument_fields_follow_reference_availability(self):
        domra = self.create_instrument(name='Домра с партиями')
        piano = self.create_instrument(name='Фортепиано без партий')
        OrchestraPart.objects.create(instrument=domra, name='Первая партия')

        empty_form = StudentAdminForm()
        domra_form = StudentAdminForm(data={
            'instrument': domra.pk,
            'custom_instrument': '',
        })
        piano_form = StudentAdminForm(data={
            'instrument': piano.pk,
            'custom_instrument': '',
        })
        custom_form = StudentAdminForm(data={
            'instrument': '',
            'custom_instrument': 'Гусли',
        })

        self.assertEqual(empty_form.fields['instrument'].empty_label, 'Другой инструмент')
        self.assertEqual(
            empty_form.fields['instrument'].widget.attrs['data-placeholder'],
            'Другой инструмент',
        )
        self.assertNotIn('disabled', empty_form.fields['custom_instrument'].widget.attrs)
        self.assertIn('disabled', domra_form.fields['custom_instrument'].widget.attrs)
        self.assertIn('disabled', piano_form.fields['custom_instrument'].widget.attrs)
        self.assertNotIn('disabled', custom_form.fields['custom_instrument'].widget.attrs)
        self.assertIn('disabled', empty_form.fields['orchestra_part'].widget.attrs)
        self.assertNotIn('disabled', domra_form.fields['orchestra_part'].widget.attrs)
        self.assertIn('disabled', piano_form.fields['orchestra_part'].widget.attrs)
        self.assertIn('disabled', custom_form.fields['orchestra_part'].widget.attrs)

    def test_student_add_page_renders_other_instrument_and_dependency_scripts(self):
        self.create_instrument(name='Баян')
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse('admin:journal_student_add'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Другой инструмент')
        self.assertContains(response, 'data-instrument-reference="1"')
        self.assertContains(response, 'data-custom-instrument="1"')
        self.assertContains(response, 'data-orchestra-part="1"')
        self.assertContains(response, 'journal/orchestra_part_dependencies_v5.js')

    def test_orchestra_part_script_refreshes_jazzmin_select2_after_api_load(self):
        javascript = Path(
            'journal/static/journal/orchestra_part_dependencies_v5.js'
        ).read_text(encoding='utf-8')

        self.assertIn("wrapped.trigger('change')", javascript)
        self.assertIn('select2:select.journalOrchestraParts', javascript)
        self.assertIn('select2:clear.journalOrchestraParts', javascript)
        self.assertIn("new URL(endpoint, window.location.origin)", javascript)
        self.assertIn("parts.forEach(function (part)", javascript)
        self.assertIn(
            'data-orchestra-parts-map',
            Path('journal/forms.py').read_text(encoding='utf-8'),
        )
        self.assertIn("field.removeAttribute('disabled')", javascript)

    def test_used_subject_delete_page_confirms_and_cascades_related_data(self):
        data = self.create_base_journal()
        subject = data['solfeggio']
        assignment = GroupSubject.objects.get(
            group=data['group'],
            subject=subject,
        )
        self.client.force_login(self.admin_user)
        delete_url = reverse('admin:journal_subject_delete', args=[subject.pk])

        confirmation = self.client.get(delete_url)

        self.assertEqual(confirmation.status_code, 200)
        self.assertContains(confirmation, str(assignment))
        self.assertContains(confirmation, 'name="post" value="yes"')

        response = self.client.post(delete_url, {'post': 'yes'})

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Subject.objects.filter(pk=subject.pk).exists())
        self.assertFalse(GroupSubject.objects.filter(pk=assignment.pk).exists())

    def test_unused_subject_can_still_be_deleted_permanently(self):
        subject = self.create_subject(name='Неиспользуемый предмет')
        self.client.force_login(self.admin_user)
        delete_url = reverse('admin:journal_subject_delete', args=[subject.pk])

        confirmation = self.client.get(delete_url)

        self.assertEqual(confirmation.status_code, 200)
        self.assertContains(confirmation, 'name="post" value="yes"')

        response = self.client.post(delete_url, {'post': 'yes'})

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Subject.objects.filter(pk=subject.pk).exists())

    def test_subject_admin_autocomplete_filters_subjects_by_assignment_type(self):
        group_subject = self.create_subject(name='Групповой предмет')
        individual_subject = self.create_subject(name='Индивидуальный предмет', is_specialty=True)
        model_admin = django_admin.site._registry[Subject]

        group_request = type(
            'Request',
            (),
            {
                'user': self.admin_user,
                'GET': {'field_name': 'subject', 'model_name': 'groupsubject'},
            },
        )()
        group_queryset, _ = model_admin.get_search_results(
            group_request,
            Subject.objects.all(),
            '',
        )
        self.assertEqual(set(group_queryset), {group_subject})

        individual_request = type(
            'Request',
            (),
            {
                'user': self.admin_user,
                'GET': {'field_name': 'subject', 'model_name': 'studentsubject'},
            },
        )()
        individual_queryset, _ = model_admin.get_search_results(
            individual_request,
            Subject.objects.all(),
            '',
        )
        self.assertEqual(set(individual_queryset), {individual_subject})

    def test_group_admin_allows_adding_students_inline(self):
        year = self.create_academic_year()
        group = StudyGroup.objects.create(name='Группа с учениками', academic_year=year)
        source_group = StudyGroup.objects.create(name='Исходная группа', academic_year=year)
        instrument = self.create_instrument()
        student = Student.objects.create(
            full_name='Готовый Ученик',
            group=source_group,
            instrument=instrument,
            city_church='Тамбов / Центр',
            is_active=True,
        )
        self.client.login(username='dashboard_admin', password='Pass12345!')

        get_response = self.client.get(reverse('admin:journal_studygroup_change', args=[group.pk]))
        self.assertNotContains(get_response, 'name="student_enrollments-0-student"')
        self.assertContains(get_response, 'name="student_enrollments-__prefix__-student"')
        self.assertContains(get_response, 'name="student_enrollments-__prefix__-city_church"')
        self.assertContains(get_response, f'value="{student.pk}"')
        self.assertContains(get_response, 'data-city-church="Тамбов / Центр"')
        self.assertContains(get_response, 'data-student-city-target="1"')
        self.assertContains(get_response, 'disabled')
        self.assertNotContains(get_response, 'name="student_enrollments-0-full_name"')
        self.assertNotContains(get_response, 'name="student_enrollments-0-instrument"')

        response = self.client.post(
            reverse('admin:journal_studygroup_change', args=[group.pk]),
            data={
                'name': group.name,
                'academic_year': year.pk,
                'is_active': 'on',
                'group_subjects-TOTAL_FORMS': '0',
                'group_subjects-INITIAL_FORMS': '0',
                'group_subjects-MIN_NUM_FORMS': '0',
                'group_subjects-MAX_NUM_FORMS': '1000',
                'student_enrollments-TOTAL_FORMS': '1',
                'student_enrollments-INITIAL_FORMS': '0',
                'student_enrollments-MIN_NUM_FORMS': '0',
                'student_enrollments-MAX_NUM_FORMS': '1000',
                'student_enrollments-0-id': '',
                'student_enrollments-0-group': group.pk,
                'student_enrollments-0-student': student.pk,
                'student_enrollments-0-city_church': 'Воронеж / Север',
                '_save': 'Save',
            },
        )

        self.assertEqual(response.status_code, 302)
        student.refresh_from_db()
        self.assertEqual(Student.objects.count(), 1)
        self.assertEqual(student.group, group)
        self.assertEqual(student.city_church, 'Тамбов / Центр')

    def test_group_admin_shows_inline_error_for_duplicate_group_subject(self):
        year = self.create_academic_year()
        group = StudyGroup.objects.create(name='Группа с дублем предмета', academic_year=year)
        subject = self.create_subject(name='Групповой дубль')
        teacher = self.create_teacher(username='duplicate_group_teacher')
        assignment = GroupSubject.objects.create(
            group=group,
            subject=subject,
            teacher=teacher,
            sort_order=10,
        )
        self.client.login(username='dashboard_admin', password='Pass12345!')

        response = self.client.post(
            reverse('admin:journal_studygroup_change', args=[group.pk]),
            data={
                'name': group.name,
                'academic_year': year.pk,
                'is_active': 'on',
                'group_subjects-TOTAL_FORMS': '2',
                'group_subjects-INITIAL_FORMS': '1',
                'group_subjects-MIN_NUM_FORMS': '0',
                'group_subjects-MAX_NUM_FORMS': '1000',
                'group_subjects-0-id': assignment.pk,
                'group_subjects-0-group': group.pk,
                'group_subjects-0-subject': subject.pk,
                'group_subjects-0-teacher': teacher.pk,
                'group_subjects-0-sort_order': '10',
                'group_subjects-0-is_active': 'on',
                'group_subjects-1-id': '',
                'group_subjects-1-group': group.pk,
                'group_subjects-1-subject': subject.pk,
                'group_subjects-1-teacher': teacher.pk,
                'group_subjects-1-sort_order': '20',
                'group_subjects-1-is_active': 'on',
                'student_enrollments-TOTAL_FORMS': '0',
                'student_enrollments-INITIAL_FORMS': '0',
                'student_enrollments-MIN_NUM_FORMS': '0',
                'student_enrollments-MAX_NUM_FORMS': '1000',
                '_save': 'Save',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            'В этой группе уже есть такой предмет.',
        )

    def test_group_admin_allows_editing_and_deleting_students_inline(self):
        year = self.create_academic_year()
        group = StudyGroup.objects.create(name='Группа с редактированием', academic_year=year)
        fallback_group = StudyGroup.objects.create(
            name=CourseApplication.STUDENT_COURSE_GROUP_NAME,
            academic_year=year,
        )
        instrument = self.create_instrument()
        student_to_edit = Student.objects.create(
            full_name='Старое Имя',
            group=group,
            instrument=instrument,
            city_church='Старый город / церковь',
            is_active=True,
        )
        student_to_delete = Student.objects.create(
            full_name='Удаляемый Ученик',
            group=group,
            instrument=instrument,
            is_active=True,
        )
        enrollment_to_edit = student_to_edit.enrollment_for_year(year)
        enrollment_to_delete = student_to_delete.enrollment_for_year(year)
        self.client.login(username='dashboard_admin', password='Pass12345!')

        response = self.client.post(
            reverse('admin:journal_studygroup_change', args=[group.pk]),
            data={
                'name': group.name,
                'academic_year': year.pk,
                'is_active': 'on',
                'group_subjects-TOTAL_FORMS': '0',
                'group_subjects-INITIAL_FORMS': '0',
                'group_subjects-MIN_NUM_FORMS': '0',
                'group_subjects-MAX_NUM_FORMS': '1000',
                'student_enrollments-TOTAL_FORMS': '2',
                'student_enrollments-INITIAL_FORMS': '2',
                'student_enrollments-MIN_NUM_FORMS': '0',
                'student_enrollments-MAX_NUM_FORMS': '1000',
                'student_enrollments-0-id': enrollment_to_edit.pk,
                'student_enrollments-0-group': group.pk,
                'student_enrollments-0-student': student_to_edit.pk,
                'student_enrollments-0-city_church': 'Новый город / церковь',
                'student_enrollments-1-id': enrollment_to_delete.pk,
                'student_enrollments-1-group': group.pk,
                'student_enrollments-1-student': student_to_delete.pk,
                'student_enrollments-1-city_church': student_to_delete.city_church,
                'student_enrollments-1-DELETE': 'on',
                '_save': 'Save',
            },
        )

        self.assertEqual(response.status_code, 302)
        student_to_edit.refresh_from_db()
        student_to_delete.refresh_from_db()
        self.assertEqual(student_to_edit.city_church, 'Старый город / церковь')
        self.assertEqual(student_to_delete.group, fallback_group)

    def test_admin_assignment_forms_limit_field_choices(self):
        data = self.create_base_journal()
        inactive_group = self.create_group(
            name='Неактивная группа',
            academic_year=data['year'],
        )
        inactive_group.is_active = False
        inactive_group.save()
        inactive_teacher = self.create_teacher(username='inactive_teacher')
        inactive_teacher.is_active = False
        inactive_teacher.save()
        inactive_subject = self.create_subject(name='Неактивный предмет')
        inactive_subject.is_active = False
        inactive_subject.save()

        group_form = GroupSubjectAdminForm()
        student_form = StudentSubjectAdminForm()
        group_admin = django_admin.site._registry[GroupSubject]
        student_admin = django_admin.site._registry[StudentSubject]

        self.assertIs(group_admin.form, GroupSubjectAdminForm)
        self.assertEqual(group_admin.autocomplete_fields, ())
        self.assertIn('journal/admin_assignment_dependencies.js', GroupSubjectAdminForm.Media.js)
        self.assertNotIn(inactive_group, group_form.fields['group'].queryset)
        self.assertNotIn(inactive_teacher, group_form.fields['teacher'].queryset)
        self.assertNotIn(inactive_subject, group_form.fields['subject'].queryset)
        self.assertIn(data['solfeggio'], group_form.fields['subject'].queryset)
        self.assertNotIn(data['specialty'], group_form.fields['subject'].queryset)
        self.assertIs(student_admin.form, StudentSubjectAdminForm)
        self.assertEqual(student_admin.autocomplete_fields, ())
        self.assertIn('journal/admin_assignment_dependencies.js', StudentSubjectAdminForm.Media.js)
        self.assertIn(data['specialty'], student_form.fields['subject'].queryset)
        self.assertNotIn(data['solfeggio'], student_form.fields['subject'].queryset)

    def test_admin_table_forms_do_not_add_per_field_search_inputs(self):
        forms_to_check = (
            SubjectResultAdminForm(),
            GroupSubjectAdminForm(),
            StudentSubjectAdminForm(),
            AssessmentItemAdminForm(),
            StudentAssessmentGroupAdminForm(),
            AssessmentResultAdminForm(),
            FinalGradeRuleAdminForm(),
        )

        for form in forms_to_check:
            with self.subTest(form=form.__class__.__name__):
                self.assertNotIn('journal/select_search.js', form.media._js)
                for field in form.fields.values():
                    self.assertNotIn('data-searchable-select', field.widget.attrs)

    def test_student_subject_admin_form_uses_subject_classification_without_duplicate_flag(self):
        data = self.create_base_journal()
        extra_subject = self.create_subject(
            name='Индивидуальная импровизация',
            is_specialty=True,
        )
        student = self.create_student(
            full_name='Ученик без специальности',
            group=data['group'],
            instrument=data['instrument'],
            username='student_without_specialty',
        )

        extra_form = StudentSubjectAdminForm(
            data={
                'student': student.pk,
                'subject': extra_subject.pk,
                'teacher': data['teacher'].pk,
                'is_active': 'on',
            },
        )
        specialty_form = StudentSubjectAdminForm(
            data={
                'student': student.pk,
                'subject': data['specialty'].pk,
                'teacher': data['other_teacher'].pk,
                'is_active': 'on',
            },
        )

        self.assertTrue(extra_form.is_valid(), extra_form.errors)
        self.assertTrue(specialty_form.is_valid(), specialty_form.errors)
        self.assertNotIn('is_specialty', extra_form.fields)
        self.assertNotIn('is_specialty', specialty_form.fields)

    def test_subject_result_admin_form_limits_subjects_by_student_assignments(self):
        data = self.create_base_journal()
        unassigned_subject = self.create_subject(name='Неназначенный предмет')

        form = SubjectResultAdminForm(
            data={
                'student': data['student'].pk,
                'academic_year': data['year'].pk,
                'subject': '',
                'exam_grade': '',
                'final_grade': '',
            },
        )

        subject_queryset = form.fields['subject'].queryset
        self.assertIn(data['solfeggio'], subject_queryset)
        self.assertIn(data['specialty'], subject_queryset)
        self.assertNotIn(unassigned_subject, subject_queryset)

    def test_grade_admin_change_form_has_single_group_field(self):
        model_admin = django_admin.site._registry[Grade]
        fieldsets = model_admin.get_fieldsets(type('Request', (), {'user': self.admin_user})())
        fields = [
            field
            for _name, options in fieldsets
            for field in options.get('fields', ())
        ]

        self.assertIn('group', fields)
        self.assertNotIn('student_group_display', fields)

    def test_group_admin_detaches_student_when_no_fallback_group_exists(self):
        year = self.create_academic_year()
        group = StudyGroup.objects.create(name='Ученики курсов', academic_year=year)
        instrument = self.create_instrument()
        student = Student.objects.create(
            full_name='Открепляемый Ученик',
            group=group,
            instrument=instrument,
            is_active=True,
        )
        enrollment = student.enrollment_for_year(year)
        self.client.login(username='dashboard_admin', password='Pass12345!')

        response = self.client.post(
            reverse('admin:journal_studygroup_change', args=[group.pk]),
            data={
                'name': group.name,
                'academic_year': year.pk,
                'is_active': 'on',
                'group_subjects-TOTAL_FORMS': '0',
                'group_subjects-INITIAL_FORMS': '0',
                'group_subjects-MIN_NUM_FORMS': '0',
                'group_subjects-MAX_NUM_FORMS': '1000',
                'student_enrollments-TOTAL_FORMS': '1',
                'student_enrollments-INITIAL_FORMS': '1',
                'student_enrollments-MIN_NUM_FORMS': '0',
                'student_enrollments-MAX_NUM_FORMS': '1000',
                'student_enrollments-0-id': enrollment.pk,
                'student_enrollments-0-group': group.pk,
                'student_enrollments-0-student': student.pk,
                'student_enrollments-0-city_church': student.city_church,
                'student_enrollments-0-DELETE': 'on',
                '_save': 'Save',
            },
        )

        self.assertEqual(response.status_code, 302)
        student.refresh_from_db()
        self.assertIsNone(student.group)

    def test_student_admin_shows_inline_error_for_duplicate_individual_subject(self):
        data = self.create_base_journal()
        student = data['student']
        assignment = StudentSubject.objects.get(
            student=student,
            subject=data['specialty'],
        )
        self.client.login(username='dashboard_admin', password='Pass12345!')

        response = self.client.post(
            reverse('admin:journal_student_change', args=[student.pk]),
            data={
                'full_name': student.full_name,
                'gender': student.gender,
                'birth_date': student.birth_date.isoformat() if student.birth_date else '',
                'group': student.group_id,
                'instrument': student.instrument_id,
                'is_active': 'on',
                'student_phone': student.student_phone,
                'parent_contacts': student.parent_contacts,
                'city_church': student.city_church,
                'music_education': student.music_education,
                'comments': student.comments,
                'user': student.user_id,
                'individual_subjects-TOTAL_FORMS': '2',
                'individual_subjects-INITIAL_FORMS': '1',
                'individual_subjects-MIN_NUM_FORMS': '0',
                'individual_subjects-MAX_NUM_FORMS': '1000',
                'individual_subjects-0-id': assignment.pk,
                'individual_subjects-0-student': student.pk,
                'individual_subjects-0-subject': data['specialty'].pk,
                'individual_subjects-0-teacher': data['other_teacher'].pk,
                'individual_subjects-0-is_active': 'on',
                'individual_subjects-1-id': '',
                'individual_subjects-1-student': student.pk,
                'individual_subjects-1-subject': data['specialty'].pk,
                'individual_subjects-1-teacher': data['other_teacher'].pk,
                'individual_subjects-1-is_active': 'on',
                'subject_results-TOTAL_FORMS': '0',
                'subject_results-INITIAL_FORMS': '0',
                'subject_results-MIN_NUM_FORMS': '0',
                'subject_results-MAX_NUM_FORMS': '1000',
                '_save': 'Save',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            'У ученика уже есть такой индивидуальный предмет.',
        )

    def test_student_admin_uses_prefetched_specialty_without_row_queries(self):
        data = self.create_base_journal()
        request = type('Request', (), {'user': self.admin_user})()
        model_admin = django_admin.site._registry[Student]
        student = model_admin.get_queryset(request).get(pk=data['student'].pk)

        with CaptureQueriesContext(connection) as captured_queries:
            self.assertEqual(model_admin.specialty_teacher_display(student), data['other_teacher'])
            self.assertEqual(model_admin.specialty_subject_display(student), data['specialty'])

        self.assertEqual(
            len(captured_queries),
            0,
            [query['sql'] for query in captured_queries],
        )

    def test_grade_admin_uses_assignment_annotations_without_row_queries(self):
        data = self.create_base_journal()
        grade = Grade.objects.create(
            student=data['student'],
            subject=data['solfeggio'],
            teacher=data['teacher'],
            date=date(2025, 10, 8),
            value='5',
        )
        request = type('Request', (), {'user': self.admin_user})()
        model_admin = django_admin.site._registry[Grade]
        grade_from_queryset = model_admin.get_queryset(request).get(pk=grade.pk)

        with CaptureQueriesContext(connection) as captured_queries:
            self.assertEqual(model_admin.source_type_display(grade_from_queryset), 'Групповой предмет')

        self.assertEqual(
            len(captured_queries),
            0,
            [query['sql'] for query in captured_queries],
        )

    def test_admin_password_reset_removes_temporary_credentials(self):
        self.create_academic_year()
        user = User.objects.create_user(
            username='temporary reset user',
            password='OldPass123!',
            is_staff=True,
        )
        TemporaryCredential.objects.create(
            user=user,
            login=user.username,
            temporary_password='OldPass123!',
        )
        self.client.force_login(self.admin_user)

        response = self.client.post(
            reverse('admin:auth_user_password_change', args=[user.pk]),
            {
                'usable_password': 'true',
                'password1': 'NewSecurePass456!',
                'password2': 'NewSecurePass456!',
            },
        )

        self.assertEqual(response.status_code, 302)
        user.refresh_from_db()
        self.assertTrue(user.check_password('NewSecurePass456!'))
        self.assertFalse(TemporaryCredential.objects.filter(user=user).exists())

    def test_invalid_admin_password_reset_keeps_temporary_credentials(self):
        self.create_academic_year()
        user = User.objects.create_user(
            username='temporary invalid reset user',
            password='OldPass123!',
            is_staff=True,
        )
        TemporaryCredential.objects.create(
            user=user,
            login=user.username,
            temporary_password='OldPass123!',
        )
        self.client.force_login(self.admin_user)

        response = self.client.post(
            reverse('admin:auth_user_password_change', args=[user.pk]),
            {
                'usable_password': 'true',
                'password1': 'NewSecurePass456!',
                'password2': 'DifferentPass789!',
            },
        )

        self.assertEqual(response.status_code, 200)
        user.refresh_from_db()
        self.assertTrue(user.check_password('OldPass123!'))
        self.assertTrue(TemporaryCredential.objects.filter(user=user).exists())

    def test_teacher_admin_creates_user_and_temporary_credentials_on_manual_add(self):
        request = type('Request', (), {'user': self.admin_user})()
        model_admin = django_admin.site._registry[Teacher]
        teacher = Teacher(
            full_name='Иванов Иван Иванович',
            email='new-teacher@example.com',
        )

        model_admin.save_model(request, teacher, form=None, change=False)

        teacher.refresh_from_db()
        self.assertIsNotNone(teacher.user)
        self.assertEqual(teacher.user.username, 'Иванов Иван')
        self.assertEqual(teacher.user.first_name, 'Иван')
        self.assertEqual(teacher.user.last_name, 'Иванов')
        self.assertTrue(teacher.user.groups.filter(name='Преподаватель').exists())
        credential = TemporaryCredential.objects.get(login=teacher.user.username)
        self.assertEqual(credential.user, teacher.user)
        self.assertTrue(credential.temporary_password)
        self.assertTrue(teacher.user.check_password(credential.temporary_password))

    def test_teacher_admin_update_preserves_password_and_temporary_password(self):
        request = type('Request', (), {'user': self.admin_user})()
        model_admin = django_admin.site._registry[Teacher]
        teacher = Teacher(full_name='Иванов Иван Иванович')
        model_admin.save_model(request, teacher, form=None, change=False)

        user = teacher.user
        credential = TemporaryCredential.objects.get(user=user)
        original_password_hash = user.password
        original_temporary_password = credential.temporary_password

        teacher.full_name = 'Петров Пётр Петрович'
        teacher.phone = '+7 (999) 555-44-33'
        model_admin.save_model(request, teacher, form=None, change=True)

        user.refresh_from_db()
        credential.refresh_from_db()
        self.assertEqual(user.password, original_password_hash)
        self.assertEqual(credential.temporary_password, original_temporary_password)
        self.assertTrue(user.check_password(original_temporary_password))

    def test_teacher_admin_does_not_reset_password_of_selected_existing_user(self):
        request = type('Request', (), {'user': self.admin_user})()
        model_admin = django_admin.site._registry[Teacher]
        existing_user = User.objects.create_user(
            username='existing teacher account',
            password='ExistingPass123!',
        )
        original_password_hash = existing_user.password
        teacher = Teacher(
            full_name='Преподаватель с аккаунтом',
            user=existing_user,
        )

        model_admin.save_model(request, teacher, form=None, change=False)
        existing_user.refresh_from_db()

        self.assertEqual(existing_user.password, original_password_hash)
        self.assertFalse(TemporaryCredential.objects.filter(user=existing_user).exists())

    def test_student_admin_creates_user_and_temporary_credentials_on_manual_add(self):
        request = type('Request', (), {'user': self.admin_user})()
        model_admin = django_admin.site._registry[Student]
        student = Student(
            full_name='Сидоров Семён Сергеевич',
            group=self.create_group(),
            instrument=self.create_instrument(),
            student_phone='+7 (999) 111-22-33',
        )

        model_admin.save_model(request, student, form=None, change=False)

        student.refresh_from_db()
        self.assertIsNotNone(student.user)
        self.assertEqual(student.user.username, 'Сидоров Семён')
        self.assertEqual(student.user.first_name, 'Семён')
        self.assertEqual(student.user.last_name, 'Сидоров')
        self.assertTrue(student.user.groups.filter(name='Ученик').exists())
        credential = TemporaryCredential.objects.get(login=student.user.username)
        self.assertEqual(credential.user, student.user)
        self.assertEqual(credential.student_phone, '+7 (999) 111-22-33')
        self.assertTrue(credential.temporary_password)
        self.assertTrue(student.user.check_password(credential.temporary_password))

    def test_student_admin_update_preserves_password_and_temporary_password(self):
        request = type('Request', (), {'user': self.admin_user})()
        model_admin = django_admin.site._registry[Student]
        student = Student(
            full_name='Новый Ученик',
            group=self.create_group(),
            instrument=self.create_instrument(),
            student_phone='+7 (999) 111-22-33',
        )
        model_admin.save_model(request, student, form=None, change=False)

        user = student.user
        credential = TemporaryCredential.objects.get(user=user)
        original_password_hash = user.password
        original_temporary_password = credential.temporary_password

        student.full_name = 'Кузнецов Кирилл Олегович'
        student.student_phone = '+7 (999) 111-22-44'
        model_admin.save_model(request, student, form=None, change=True)

        user.refresh_from_db()
        credential.refresh_from_db()
        self.assertEqual(user.password, original_password_hash)
        self.assertEqual(credential.temporary_password, original_temporary_password)
        self.assertEqual(credential.student_phone, '+7 (999) 111-22-44')
        self.assertTrue(user.check_password(original_temporary_password))

    def test_student_admin_does_not_reset_password_of_selected_existing_user(self):
        request = type('Request', (), {'user': self.admin_user})()
        model_admin = django_admin.site._registry[Student]
        existing_user = User.objects.create_user(
            username='existing student account',
            password='ExistingPass123!',
        )
        original_password_hash = existing_user.password
        student = Student(
            full_name='Ученик с аккаунтом',
            group=self.create_group(),
            instrument=self.create_instrument(),
            user=existing_user,
        )

        model_admin.save_model(request, student, form=None, change=False)
        existing_user.refresh_from_db()

        self.assertEqual(existing_user.password, original_password_hash)
        self.assertTrue(existing_user.groups.filter(name='Ученик').exists())
        self.assertFalse(TemporaryCredential.objects.filter(user=existing_user).exists())

    def test_teacher_admin_allows_adding_group_and_individual_subjects_inline(self):
        year = self.create_academic_year()
        group = StudyGroup.objects.create(name='Группа преподавателя', academic_year=year)
        instrument = self.create_instrument()
        student = Student.objects.create(
            full_name='Индивидуальный Ученик',
            group=group,
            instrument=instrument,
        )
        teacher = self.create_teacher(
            full_name='Преподаватель Назначений',
            username='teacher_assignments',
        )
        group_subject = self.create_subject(name='Групповой предмет')
        individual_subject = self.create_subject(name='Специальность назначений', is_specialty=True)
        self.client.login(username='dashboard_admin', password='Pass12345!')

        get_response = self.client.get(reverse('admin:journal_teacher_change', args=[teacher.pk]))
        self.assertNotContains(get_response, 'name="group_subjects-0-group"')
        self.assertNotContains(get_response, 'name="individual_subjects-0-student"')
        self.assertContains(get_response, 'name="group_subjects-__prefix__-group"')
        self.assertContains(get_response, 'name="individual_subjects-__prefix__-student"')

        response = self.client.post(
            reverse('admin:journal_teacher_change', args=[teacher.pk]),
            data={
                'full_name': teacher.full_name,
                'birth_date': '',
                'phone': '',
                'email': '',
                'comments': '',
                'user': teacher.user_id,
                'is_active': 'on',
                'group_subjects-TOTAL_FORMS': '1',
                'group_subjects-INITIAL_FORMS': '0',
                'group_subjects-MIN_NUM_FORMS': '0',
                'group_subjects-MAX_NUM_FORMS': '1000',
                'group_subjects-0-id': '',
                'group_subjects-0-teacher': teacher.pk,
                'group_subjects-0-group': group.pk,
                'group_subjects-0-subject': group_subject.pk,
                'group_subjects-0-sort_order': '10',
                'group_subjects-0-is_active': 'on',
                'individual_subjects-TOTAL_FORMS': '1',
                'individual_subjects-INITIAL_FORMS': '0',
                'individual_subjects-MIN_NUM_FORMS': '0',
                'individual_subjects-MAX_NUM_FORMS': '1000',
                'individual_subjects-0-id': '',
                'individual_subjects-0-teacher': teacher.pk,
                'individual_subjects-0-student': student.pk,
                'individual_subjects-0-subject': individual_subject.pk,
                'individual_subjects-0-is_active': 'on',
                'responsible_assessment_items-TOTAL_FORMS': '0',
                'responsible_assessment_items-INITIAL_FORMS': '0',
                'responsible_assessment_items-MIN_NUM_FORMS': '0',
                'responsible_assessment_items-MAX_NUM_FORMS': '1000',
                '_save': 'Save',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(
            GroupSubject.objects.filter(
                group=group,
                subject=group_subject,
                teacher=teacher,
            ).exists(),
        )
        self.assertTrue(
            StudentSubject.objects.filter(
                student=student,
                subject=individual_subject,
                teacher=teacher,
            ).exists(),
        )

    def test_changing_group_assignment_in_admin_cascades_to_existing_grades(self):
        data = self.create_base_journal()
        assignment = GroupSubject.objects.get(
            group=data['group'],
            subject=data['solfeggio'],
        )
        grade = Grade.objects.create(
            student=data['student'],
            subject=data['solfeggio'],
            teacher=data['teacher'],
            date=date(2025, 10, 8),
            value='5',
        )
        self.client.login(username='dashboard_admin', password='Pass12345!')

        response = self.client.post(
            reverse('admin:journal_groupsubject_change', args=[assignment.pk]),
            data={
                'group': data['group'].pk,
                'subject': data['solfeggio'].pk,
                'teacher': data['other_teacher'].pk,
                'sort_order': assignment.sort_order,
                'is_active': 'on',
                '_save': 'Save',
            },
        )

        self.assertEqual(response.status_code, 302)
        grade.refresh_from_db()
        self.assertEqual(grade.teacher, data['other_teacher'])

    def test_changing_individual_assignment_in_admin_cascades_to_existing_grades(self):
        data = self.create_base_journal()
        assignment = StudentSubject.objects.get(
            student=data['student'],
            subject=data['specialty'],
        )
        grade = Grade.objects.create(
            student=data['student'],
            subject=data['specialty'],
            teacher=data['other_teacher'],
            date=date(2025, 10, 9),
            value='4',
        )
        self.client.login(username='dashboard_admin', password='Pass12345!')

        response = self.client.post(
            reverse('admin:journal_studentsubject_change', args=[assignment.pk]),
            data={
                'student': data['student'].pk,
                'subject': data['specialty'].pk,
                'teacher': data['teacher'].pk,
                'is_active': 'on',
                '_save': 'Save',
            },
        )

        self.assertEqual(response.status_code, 302)
        grade.refresh_from_db()
        self.assertEqual(grade.teacher, data['teacher'])

    def test_student_birth_date_change_ignores_unchanged_historical_subject_results(self):
        data = self.create_base_journal()
        assignment = GroupSubject.objects.get(
            group=data['group'],
            subject=data['solfeggio'],
        )
        individual_assignment = StudentSubject.objects.get(
            student=data['student'],
            subject=data['specialty'],
        )
        result = SubjectResult.objects.create(
            student=data['student'],
            subject=data['solfeggio'],
            academic_year=data['year'],
            exam_grade='5',
            final_grade='5',
        )
        assignment.is_active = False
        assignment.save()
        self.client.login(username='dashboard_admin', password='Pass12345!')

        response = self.client.post(
            reverse('admin:journal_student_change', args=[data['student'].pk]),
            data={
                'full_name': data['student'].full_name,
                'gender': data['student'].gender,
                'birth_date': '2011-02-02',
                'group': data['group'].pk,
                'instrument': data['instrument'].pk,
                'is_active': 'on',
                'student_phone': data['student'].student_phone,
                'parent_contacts': data['student'].parent_contacts,
                'city_church': data['student'].city_church,
                'music_education': data['student'].music_education,
                'comments': data['student'].comments,
                'user': data['student'].user_id,
                'individual_subjects-TOTAL_FORMS': '1',
                'individual_subjects-INITIAL_FORMS': '1',
                'individual_subjects-MIN_NUM_FORMS': '0',
                'individual_subjects-MAX_NUM_FORMS': '1000',
                'individual_subjects-0-id': individual_assignment.pk,
                'individual_subjects-0-student': data['student'].pk,
                'individual_subjects-0-subject': data['specialty'].pk,
                'individual_subjects-0-teacher': data['other_teacher'].pk,
                'individual_subjects-0-is_active': 'on',
                'subject_results-TOTAL_FORMS': '1',
                'subject_results-INITIAL_FORMS': '1',
                'subject_results-MIN_NUM_FORMS': '0',
                'subject_results-MAX_NUM_FORMS': '1000',
                'subject_results-0-id': result.pk,
                'subject_results-0-student': data['student'].pk,
                'subject_results-0-academic_year': data['year'].pk,
                'subject_results-0-subject': data['solfeggio'].pk,
                'subject_results-0-exam_grade': '5',
                'subject_results-0-final_grade': '5',
                'assessment_group_assignments-TOTAL_FORMS': '0',
                'assessment_group_assignments-INITIAL_FORMS': '0',
                'assessment_group_assignments-MIN_NUM_FORMS': '0',
                'assessment_group_assignments-MAX_NUM_FORMS': '1000',
                'grades-TOTAL_FORMS': '0',
                'grades-INITIAL_FORMS': '0',
                'grades-MIN_NUM_FORMS': '0',
                'grades-MAX_NUM_FORMS': '1000',
                '_save': 'Save',
            },
        )

        self.assertEqual(response.status_code, 302)
        data['student'].refresh_from_db()
        self.assertEqual(data['student'].birth_date, date(2011, 2, 2))


class CaseInsensitiveAuthenticationTests(TestCase):
    def test_login_ignores_ascii_username_case(self):
        user = User.objects.create_user(
            username='Teacher.Mixed',
            password='Pass12345!',
        )

        login_page = self.client.get(reverse('login'))

        response = self.client.post(
            reverse('login'),
            data={
                'username': 'teacher.mIXED',
                'password': 'Pass12345!',
            },
        )

        self.assertContains(login_page, 'Регистр букв в логине не важен.')
        self.assertEqual(response.status_code, 302)
        self.assertEqual(int(self.client.session['_auth_user_id']), user.pk)

    def test_login_ignores_cyrillic_username_case(self):
        user = User.objects.create_user(
            username='Иванов Иван',
            password='Pass12345!',
        )

        authenticated = self.client.login(
            username='ивАНОВ иВАН',
            password='Pass12345!',
        )

        self.assertTrue(authenticated)
        self.assertEqual(int(self.client.session['_auth_user_id']), user.pk)

    def test_login_still_rejects_an_incorrect_password(self):
        User.objects.create_user(
            username='CaseSensitivePassword',
            password='Pass12345!',
        )

        self.assertFalse(
            self.client.login(
                username='casesensitivepassword',
                password='pass12345!',
            ),
        )


class PasswordRecoveryViewTests(TestCase):
    def test_login_page_contains_password_help_link(self):
        response = self.client.get(reverse('login'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Забыли пароль?')
        self.assertContains(response, reverse('password_help'))

    def test_password_help_lists_only_active_contacts_in_configured_order(self):
        second = PasswordRecoveryContact.objects.create(
            name='Второй администратор',
            phone='8 999 222 33 44',
            messengers='WhatsApp',
            display_order=20,
        )
        first = PasswordRecoveryContact.objects.create(
            name='Первый администратор',
            phone='+7 (999) 111-22-33',
            messengers='Telegram, MAX',
            messenger_username='first_admin',
            display_order=10,
        )
        PasswordRecoveryContact.objects.create(
            name='Скрытый администратор',
            phone='+7 (999) 000-00-00',
            messengers='Telegram',
            is_active=False,
        )

        response = self.client.get(reverse('password_help'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context['contacts']), [first, second])
        self.assertContains(response, 'Первый администратор')
        self.assertContains(response, '+7 (999) 111-22-33')
        self.assertContains(response, 'Telegram')
        self.assertContains(response, 'href="https://t.me/first_admin"')
        self.assertContains(response, '@first_admin')
        self.assertContains(response, 'Второй администратор')
        self.assertNotContains(response, 'Скрытый администратор')

    def test_telegram_username_is_used_only_for_telegram(self):
        contact = PasswordRecoveryContact.objects.create(
            name='Администратор',
            phone='+7 (999) 123-45-67',
            messengers='Telegram, WhatsApp, Viber, MAX',
            messenger_username='journal_admin',
        )

        links = {item['name']: item['url'] for item in contact.messenger_links}

        self.assertTrue(contact.has_telegram_messenger)
        self.assertEqual(links['Telegram'], 'https://t.me/journal_admin')
        self.assertEqual(links['WhatsApp'], 'https://wa.me/79991234567')
        self.assertEqual(links['Viber'], 'viber://chat?number=%2B79991234567')
        self.assertEqual(links['MAX'], 'tel:+79991234567')

    def test_telegram_uses_phone_when_username_is_empty(self):
        contact = PasswordRecoveryContact.objects.create(
            name='Администратор',
            phone='+7 (999) 123-45-67',
            messengers='Telegram',
            messenger_username='',
        )

        self.assertEqual(
            contact.messenger_links[0]['url'],
            'tg://resolve?phone=%2B79991234567',
        )

    def test_non_telegram_contact_does_not_expose_telegram_username(self):
        contact = PasswordRecoveryContact.objects.create(
            name='Администратор',
            phone='+7 (999) 123-45-67',
            messengers='WhatsApp',
            messenger_username='unused_username',
        )

        response = self.client.get(reverse('password_help'))

        self.assertFalse(contact.has_telegram_messenger)
        self.assertNotContains(response, '@unused_username')
        self.assertContains(response, 'href="https://wa.me/79991234567"')

    def test_password_help_has_empty_state_without_configured_contacts(self):
        response = self.client.get(reverse('password_help'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Контакты администраторов пока не опубликованы')

    def test_recovery_contact_normalizes_values_and_builds_phone_link(self):
        contact = PasswordRecoveryContact.objects.create(
            name='  Администратор  ',
            phone='8 999 123 45 67',
            messengers='  Telegram, WhatsApp  ',
            messenger_username='  @journal_admin  ',
        )

        self.assertEqual(contact.name, 'Администратор')
        self.assertEqual(contact.phone, '+7 (999) 123-45-67')
        self.assertEqual(contact.messengers, 'Telegram, WhatsApp')
        self.assertEqual(contact.messenger_username, 'journal_admin')
        self.assertEqual(contact.phone_uri, 'tel:+79991234567')
        self.assertEqual(contact.messenger_links[0]['url'], 'https://t.me/journal_admin')
        self.assertEqual(contact.messenger_links[1]['url'], 'https://wa.me/79991234567')


class CourseRegistrationViewTests(JournalTestDataMixin, TestCase):
    def setUp(self):
        academic_year = self.create_academic_year(name='2025/2026')
        registration_settings = CourseRegistrationSettings.load(academic_year)
        registration_settings.telegram_group_url = 'https://t.me/test_group'
        registration_settings.minimum_registration_age = 14
        registration_settings.save()

    def test_registration_page_creates_confirmed_application_and_shows_credentials(
        self,
    ):
        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='Temp12345!',
        ):
            response = self.client.post(
                reverse('course_registration'),
                data=self.application_form_payload(),
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(CourseApplication.objects.count(), 1)
        self.assertEqual(Student.objects.count(), 1)
        self.assertEqual(TemporaryCredential.objects.count(), 1)
        self.assertContains(response, 'Иванов Иван')
        self.assertContains(response, 'Temp12345!')

    def test_registration_page_has_no_initial_age_calculation_prompt(self):
        response = self.client.get(reverse('course_registration'))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(
            response,
            'Возраст будет рассчитан после выбора даты рождения.',
        )

    def test_manual_registration_close_hides_form_and_rejects_new_application(self):
        registration_settings = CourseRegistrationSettings.load()
        registration_settings.registration_mode = (
            CourseRegistrationSettings.REGISTRATION_MODE_CLOSED
        )
        registration_settings.save()

        get_response = self.client.get(reverse('course_registration'))
        post_response = self.client.post(
            reverse('course_registration'),
            data=self.application_form_payload(),
        )

        self.assertEqual(get_response.status_code, 200)
        self.assertContains(get_response, 'Регистрация завершена')
        self.assertNotContains(get_response, 'id="course-registration-form"')
        self.assertEqual(post_response.status_code, 409)
        self.assertContains(post_response, 'Новые заявки сейчас не принимаются', status_code=409)
        self.assertFalse(CourseApplication.objects.exists())

    def test_automatic_limit_ignores_rejected_applications(self):
        registration_settings = CourseRegistrationSettings.load()
        registration_settings.registration_mode = (
            CourseRegistrationSettings.REGISTRATION_MODE_AUTOMATIC
        )
        registration_settings.application_limit = 1
        registration_settings.save()
        self.assertTrue(registration_settings.registration_is_open())

        application = CourseApplication.objects.create(**self.application_payload())
        self.assertFalse(registration_settings.registration_is_open())

        application.status = CourseApplication.STATUS_REJECTED
        application.save()

        self.assertEqual(registration_settings.registered_applications_count(), 0)
        self.assertTrue(registration_settings.registration_is_open())

    def test_manual_open_overrides_reached_application_limit(self):
        CourseApplication.objects.create(**self.application_payload())
        registration_settings = CourseRegistrationSettings.load()
        registration_settings.application_limit = 1
        registration_settings.registration_mode = (
            CourseRegistrationSettings.REGISTRATION_MODE_OPEN
        )
        registration_settings.save()

        response = self.client.get(reverse('course_registration'))

        self.assertTrue(registration_settings.registration_is_open())
        self.assertContains(response, 'id="course-registration-form"')

    def test_registration_page_rejects_duplicate_phone_without_second_application(
        self,
    ):
        CourseApplication.objects.create(**self.application_payload())

        response = self.client.post(
            reverse('course_registration'),
            data=self.application_form_payload(
                last_name='Петров',
                first_name='Пётр',
                student_phone='8 999 123 45 67',
            ),
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(CourseApplication.objects.count(), 1)
        self.assertIn(
            'Ученик с таким номером телефона уже зарегистрирован.',
            response.content.decode('utf-8'),
        )

    def test_registration_page_shows_friendly_csrf_error(self):
        csrf_client = Client(enforce_csrf_checks=True, HTTP_HOST='127.0.0.1')

        response = csrf_client.post(
            reverse('course_registration'),
            data=self.application_form_payload(),
        )

        self.assertEqual(response.status_code, 403)
        self.assertContains(response, 'Срок действия формы истёк', status_code=403)
        self.assertContains(
            response,
            'Обновите страницу и повторите действие',
            status_code=403,
        )
        self.assertContains(response, 'Обновить страницу', status_code=403)
        self.assertNotContains(response, 'CSRF', status_code=403)
        self.assertFalse(CourseApplication.objects.exists())

    def test_duplicate_phone_is_checked_inside_academic_year_only(self):
        first_application = CourseApplication.objects.create(**self.application_payload())

        with self.assertRaisesMessage(ValidationError, 'Ученик с таким номером телефона уже зарегистрирован.'):
            CourseApplication.objects.create(
                **self.application_payload(
                    last_name='Петров',
                    first_name='Пётр',
                    student_phone='8 999 123 45 67',
                ),
            )

        AcademicYear.objects.create(
            name='2026/2027',
            starts_on=date(2026, 9, 1),
            ends_on=date(2027, 8, 31),
        )
        second_application = CourseApplication.objects.create(
            **self.application_payload(
                last_name='Петров',
                first_name='Пётр',
                student_phone='8 999 123 45 67',
            ),
        )

        first_application.refresh_from_db()
        second_application.refresh_from_db()

        self.assertNotEqual(first_application.academic_year_id, second_application.academic_year_id)
        self.assertEqual(CourseApplication.objects.count(), 2)

    def test_registration_api_creates_credentials_without_returning_password(self):
        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='Temp12345!',
        ):
            response = self.client.post(
                reverse('course_registration_api'),
                data=self.application_form_payload(),
            )

        self.assertEqual(response.status_code, 201)

        payload = response.json()

        self.assertTrue(payload['success'])
        self.assertEqual(payload['status'], CourseApplication.STATUS_CONFIRMED)
        self.assertEqual(payload['status_display'], 'Подтверждена')
        self.assertTrue(payload['credentials_created'])
        self.assertNotIn('login', payload)
        self.assertNotIn('temporary_password', payload)

    def test_registration_api_stops_accepting_applications_at_automatic_limit(self):
        registration_settings = CourseRegistrationSettings.load()
        registration_settings.registration_mode = (
            CourseRegistrationSettings.REGISTRATION_MODE_AUTOMATIC
        )
        registration_settings.application_limit = 1
        registration_settings.save()

        first_response = self.client.post(
            reverse('course_registration_api'),
            data=self.application_form_payload(),
        )
        second_response = self.client.post(
            reverse('course_registration_api'),
            data=self.application_form_payload(
                last_name='Петров',
                student_phone='+7 (999) 765-43-21',
            ),
        )

        self.assertEqual(first_response.status_code, 201)
        self.assertEqual(second_response.status_code, 409)
        self.assertEqual(
            second_response.json()['message'],
            'Регистрация завершена. Новые заявки сейчас не принимаются.',
        )
        self.assertEqual(CourseApplication.objects.count(), 1)

    def test_registration_api_rejects_non_object_json_payload(self):
        response = self.client.post(
            reverse('course_registration_api'),
            data='[]',
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.json(),
            {'success': False, 'message': 'Неверный формат запроса.'},
        )
        self.assertFalse(CourseApplication.objects.exists())

    def test_registration_api_rejects_invalid_utf8_json_payload(self):
        response = self.client.post(
            reverse('course_registration_api'),
            data=b'{"student_phone": "\xff"}',
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.json(),
            {'success': False, 'message': 'Неверный формат запроса.'},
        )
        self.assertFalse(CourseApplication.objects.exists())

    def test_registration_api_requires_csrf_cookie(self):
        csrf_client = Client(enforce_csrf_checks=True, HTTP_HOST='127.0.0.1')

        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='Temp12345!',
        ):
            response = csrf_client.post(
                reverse('course_registration_api'),
                data=self.application_form_payload(),
            )

        self.assertEqual(response.status_code, 403)
        self.assertEqual(response['Content-Type'], 'application/json')
        self.assertEqual(
            response.json(),
            {
                'success': False,
                'code': 'csrf_failed',
                'message': (
                    'Не удалось подтвердить отправку формы. '
                    'Обновите страницу и повторите действие.'
                ),
            },
        )
        self.assertEqual(CourseApplication.objects.count(), 0)

    def test_registration_api_rejects_duplicate_phone(self):
        CourseApplication.objects.create(**self.application_payload())

        response = self.client.post(
            reverse('course_registration_api'),
            data=self.application_form_payload(
                last_name='Петров',
                first_name='Пётр',
                student_phone='8 999 123 45 67',
            ),
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.json()['success'])
        self.assertIn('student_phone', response.json()['errors'])
        self.assertEqual(CourseApplication.objects.count(), 1)

    @override_settings(
        CACHES={
            'default': {
                'BACKEND': 'django.core.cache.backends.locmem.LocMemCache',
                'LOCATION': 'course-registration-throttle-test',
            },
        },
    )
    def test_registration_api_limits_repeated_requests_from_same_ip(self):
        url = reverse('course_registration_api')

        for _ in range(10):
            response = self.client.post(
                url,
                data='{}',
                content_type='application/json',
                REMOTE_ADDR='203.0.113.10',
            )
            self.assertEqual(response.status_code, 400)

        response = self.client.post(
            url,
            data='{}',
            content_type='application/json',
            REMOTE_ADDR='203.0.113.10',
        )

        self.assertEqual(response.status_code, 429)
        self.assertFalse(response.json()['success'])

    @override_settings(TRUST_X_FORWARDED_FOR=True, TRUSTED_PROXY_COUNT=1)
    def test_registration_rate_limit_uses_ip_appended_by_trusted_proxy(self):
        url = reverse('course_registration_api')

        for attempt in range(10):
            response = self.client.post(
                url,
                data='{}',
                content_type='application/json',
                REMOTE_ADDR='172.18.0.1',
                HTTP_X_FORWARDED_FOR=f'198.51.100.{attempt}, 203.0.113.10',
            )
            self.assertEqual(response.status_code, 400)

        response = self.client.post(
            url,
            data='{}',
            content_type='application/json',
            REMOTE_ADDR='172.18.0.1',
            HTTP_X_FORWARDED_FOR='198.51.100.250, 203.0.113.10',
        )

        self.assertEqual(response.status_code, 429)
        self.assertFalse(response.json()['success'])


class DockerMigrationBootstrapTests(SimpleTestCase):
    project_root = Path(__file__).resolve().parents[1]

    def test_entrypoint_runs_makemigrations_before_runtime_migrate(self):
        entrypoint = (self.project_root / 'docker' / 'entrypoint.sh').read_text(encoding='utf-8')

        create_position = entrypoint.index('python manage.py makemigrations --noinput')
        check_position = entrypoint.index('python manage.py makemigrations --check --dry-run')
        migrate_position = entrypoint.index('python manage.py migrate --noinput')

        self.assertLess(create_position, migrate_position)
        self.assertLess(check_position, migrate_position)
        self.assertIn('MIGRATION_MODE="${MIGRATION_MODE:-check}"', entrypoint)
        self.assertIn('create)', entrypoint)
        self.assertIn('check)', entrypoint)

    def test_docker_build_validates_and_applies_migration_chain(self):
        dockerfile = (self.project_root / 'Dockerfile').read_text(encoding='utf-8')

        self.assertIn('python manage.py makemigrations --check --dry-run', dockerfile)
        self.assertIn('python manage.py migrate --noinput', dockerfile)
        self.assertIn('/tmp/cadet-journal-build-check.sqlite3', dockerfile)
        self.assertLess(
            dockerfile.index('python manage.py makemigrations --check --dry-run'),
            dockerfile.index('python manage.py migrate --noinput'),
        )

    def test_docker_build_collects_static_and_runtime_is_read_only(self):
        dockerfile = (self.project_root / 'Dockerfile').read_text(encoding='utf-8')
        entrypoint = (self.project_root / 'docker' / 'entrypoint.sh').read_text(encoding='utf-8')
        production_compose = (
            self.project_root / 'docker-compose.prod.yml'
        ).read_text(encoding='utf-8')

        self.assertIn('python manage.py collectstatic --noinput --clear', dockerfile)
        self.assertNotIn('python manage.py collectstatic', entrypoint)
        self.assertIn('read_only: true', production_compose)
        self.assertIn('no-new-privileges:true', production_compose)
        self.assertIn('cap_drop:', production_compose)

    def test_compose_uses_safe_migration_modes(self):
        development_compose = (
            self.project_root / 'docker-compose.dev.yml'
        ).read_text(encoding='utf-8')
        production_compose = (
            self.project_root / 'docker-compose.prod.yml'
        ).read_text(encoding='utf-8')

        self.assertIn('MIGRATION_MODE: create', development_compose)
        self.assertIn('MIGRATION_MODE: check', production_compose)
        production_environment = (
            self.project_root / '.env.prod.example'
        ).read_text(encoding='utf-8')
        self.assertNotIn('MIGRATION_MODE=', production_environment)

    def test_environment_examples_do_not_duplicate_compose_database_settings(self):
        local_environment = (
            self.project_root / '.env.example'
        ).read_text(encoding='utf-8')
        self.assertIn(
            'DB_ENGINE=django.db.backends.sqlite3',
            local_environment,
        )
        self.assertIn('DB_NAME=journal_db', local_environment)

        for env_path in (
            self.project_root / '.env.dev.example',
            self.project_root / '.env.prod.example',
        ):
            with self.subTest(env_path=env_path.name):
                environment = env_path.read_text(encoding='utf-8')
                self.assertIn('POSTGRES_DB=journal_db', environment)
                self.assertIn('POSTGRES_USER=journal_user', environment)
                self.assertNotIn('DB_ENGINE=', environment)
                self.assertNotIn('DB_NAME=', environment)
                self.assertNotIn('DB_USER=', environment)
                self.assertNotIn('DB_PASSWORD=', environment)

        test_environment = (self.project_root / '.env.test').read_text(encoding='utf-8')
        self.assertIn('DJANGO_ENV=test', test_environment)
        self.assertIn('DB_ENGINE=django.db.backends.postgresql', test_environment)
        self.assertEqual(
            test_environment,
            (self.project_root / '.env.test.example').read_text(encoding='utf-8'),
        )


class DeploymentPipelineTests(SimpleTestCase):
    project_root = Path(__file__).resolve().parents[1]

    def test_ci_has_real_fast_and_slow_test_scripts(self):
        ci = (self.project_root / '.github' / 'workflows' / 'ci.yml').read_text(
            encoding='utf-8'
        )

        self.assertTrue((self.project_root / 'scripts' / 'test-fast.sh').exists())
        self.assertTrue((self.project_root / 'scripts' / 'test-slow.sh').exists())
        self.assertIn('./scripts/test-fast.sh -v 2', ci)
        self.assertIn('./scripts/test-slow.sh -v 2', ci)
        fast_script = (self.project_root / 'scripts' / 'test-fast.sh').read_text(
            encoding='utf-8'
        )
        slow_script = (self.project_root / 'scripts' / 'test-slow.sh').read_text(
            encoding='utf-8'
        )
        self.assertIn('--exclude-tag=slow', fast_script)
        self.assertIn('--tag=slow', slow_script)

    def test_deploy_is_called_only_after_required_ci_job(self):
        ci = (self.project_root / '.github' / 'workflows' / 'ci.yml').read_text(
            encoding='utf-8'
        )
        cd = (self.project_root / '.github' / 'workflows' / 'cd.yml').read_text(
            encoding='utf-8'
        )

        self.assertIn("needs:\n      - ci-passed", ci)
        self.assertIn("github.event_name == 'push'", ci)
        self.assertIn("github.ref == 'refs/heads/main'", ci)
        self.assertIn('uses: ./.github/workflows/cd.yml', ci)
        self.assertIn('workflow_call:', cd)
        self.assertNotIn('workflow_dispatch:', cd)
        self.assertNotIn('workflow_run:', cd)
        self.assertNotIn('create:', cd)

    def test_ci_covers_pull_requests_and_merge_queue_without_deploying_them(self):
        ci = (self.project_root / '.github' / 'workflows' / 'ci.yml').read_text(
            encoding='utf-8'
        )

        self.assertIn('pull_request:', ci)
        self.assertIn('merge_group:', ci)
        self.assertIn('types: [checks_requested]', ci)
        self.assertIn("github.ref == 'refs/heads/main'", ci)

    def test_cd_uses_verified_sha_and_key_based_ssh(self):
        cd = (self.project_root / '.github' / 'workflows' / 'cd.yml').read_text(
            encoding='utf-8'
        )
        remote = (self.project_root / 'scripts' / 'remote-deploy.sh').read_text(
            encoding='utf-8'
        )

        self.assertIn('SSH_PRIVATE_KEY', cd)
        self.assertIn('SSH_KNOWN_HOSTS', cd)
        self.assertNotIn('SSH_PASSWORD', cd)
        self.assertIn('DEPLOY_SHA: ${{ github.sha }}', cd)
        self.assertIn('git rev-parse origin/main', remote)
        self.assertIn('git reset --hard "$DEPLOY_SHA"', remote)

    def test_production_env_is_rendered_and_validated_before_upload(self):
        cd = (self.project_root / '.github' / 'workflows' / 'cd.yml').read_text(
            encoding='utf-8'
        )
        run_prod = (self.project_root / 'scripts' / 'run-prod.sh').read_text(
            encoding='utf-8'
        )

        self.assertIn('scripts/render_prod_env.py', cd)
        self.assertIn('scripts/validate_env.py --file .env.prod', run_prod)
        self.assertIn('config --quiet', run_prod)
        self.assertIn('--wait-timeout', run_prod)


class AsyncDatabaseViewTests(TestCase):
    def test_healthcheck_verifies_database_connection(self):
        response = self.client.get(reverse('healthcheck'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {'status': 'ok'})

    def test_database_backed_url_views_are_async(self):
        from asgiref.sync import iscoroutinefunction

        from journal import admin_tools, views

        async_views = (
            views.password_help_view,
            views.grade_options_api,
            views.assignment_options_api,
            views.journal_view,
            views.course_registration_view,
            views.course_registration_api,
            views.export_student_credentials_xlsx,
            views.export_all_data_excel,
            admin_tools.admin_data_tools_view,
            admin_tools.admin_guide_view,
            admin_tools.admin_seed_test_data_view,
            admin_tools.admin_delete_database_view,
            admin_tools.admin_export_test_credentials_excel_view,
        )

        for view_func in async_views:
            with self.subTest(view=view_func.__name__):
                self.assertTrue(iscoroutinefunction(view_func))


class ExportTemporaryCredentialsAdminXlsxTests(JournalTestDataMixin, TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            username='admin_xlsx',
            password='Pass12345!',
            email='admin-xlsx@example.com',
        )
        self.regular_user = User.objects.create_user(
            username='regular_xlsx',
            password='Pass12345!',
        )
        self.staff_user = User.objects.create_user(
            username='staff_xlsx',
            password='Pass12345!',
            is_staff=True,
        )

        self.teacher_group = Group.objects.create(name='Преподаватель')
        self.student_group = Group.objects.create(name='Ученик')

        self.teacher_user = User.objects.create_user(
            username='teacher_export',
            password='Pass12345!',
        )
        self.teacher_user.groups.add(self.teacher_group)

        self.student_user = User.objects.create_user(
            username='student_export',
            password='Pass12345!',
        )
        self.student_user.groups.add(self.student_group)

        TemporaryCredential.objects.create(
            login='teacher_export',
            temporary_password='TeacherTemp123!',
        )
        TemporaryCredential.objects.create(
            login='student_export',
            temporary_password='StudentTemp123!',
        )

    def test_superuser_can_download_temporary_credentials_xlsx(self):
        self.client.login(username='admin_xlsx', password='Pass12345!')

        response = self.client.get(reverse('admin_export_test_credentials_excel'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))

        self.assertEqual(rows[0], ('Логин', 'Пароль', 'Роль'))
        self.assertIn(('teacher_export', 'TeacherTemp123!', 'Преподаватель'), rows)
        self.assertIn(('student_export', 'StudentTemp123!', 'Ученик'), rows)
        self.assertNotIn('Телефон ученика', rows[0])
        self.assertNotIn('Заявка', rows[0])

    def test_admin_temporary_credentials_export_uses_selected_year(self):
        old_year = self.create_academic_year(name='2025/2026')
        old_application = CourseApplication.objects.create(**self.application_payload())
        self.create_academic_year(name='2026/2027')
        active_application = CourseApplication.objects.create(
            **self.application_payload(
                last_name='Петров',
                birth_date=date(2001, 2, 2),
                student_phone='+7 (999) 765-43-21',
            ),
        )
        session = self.client.session
        session['journal_admin_academic_year_id'] = old_year.pk
        session.save()
        self.client.login(username='admin_xlsx', password='Pass12345!')

        response = self.client.get(reverse('admin_export_test_credentials_excel'))
        workbook = load_workbook(BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))

        self.assertIn(old_application.generated_login, {row[0] for row in rows[1:]})
        self.assertNotIn(active_application.generated_login, {row[0] for row in rows[1:]})
        self.assertNotIn('teacher_export', {row[0] for row in rows[1:]})

    def test_admin_temporary_credentials_export_rejects_post(self):
        self.client.login(username='admin_xlsx', password='Pass12345!')

        response = self.client.post(reverse('admin_export_test_credentials_excel'))

        self.assertEqual(response.status_code, 405)

    def test_temporary_credentials_export_escapes_excel_formulas(self):
        TemporaryCredential.objects.create(
            login='=HYPERLINK("https://example.invalid")',
            temporary_password='+1+1',
        )
        self.client.login(username='admin_xlsx', password='Pass12345!')

        response = self.client.get(reverse('admin_export_test_credentials_excel'))
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        rows = list(workbook.active.iter_rows(values_only=True))

        self.assertIn(
            ("'=HYPERLINK(\"https://example.invalid\")", "'+1+1", None),
            rows,
        )

    def test_legacy_temporary_credentials_export_rejects_post(self):
        self.client.login(username='admin_xlsx', password='Pass12345!')

        response = self.client.post(reverse('export_student_credentials_xlsx'))

        self.assertEqual(response.status_code, 405)

    def test_regular_user_cannot_download_temporary_credentials_xlsx(self):
        self.client.login(username='regular_xlsx', password='Pass12345!')

        response = self.client.get(reverse('admin_export_test_credentials_excel'))

        self.assertEqual(response.status_code, 302)

    def test_staff_user_cannot_download_temporary_credentials_xlsx(self):
        self.client.login(username='staff_xlsx', password='Pass12345!')

        response = self.client.get(reverse('admin_export_test_credentials_excel'))

        self.assertEqual(response.status_code, 302)

    def test_staff_user_cannot_open_data_tools(self):
        self.client.login(username='staff_xlsx', password='Pass12345!')

        response = self.client.get(reverse('admin_data_tools'))

        self.assertEqual(response.status_code, 302)

    def test_superuser_can_open_data_tools_with_delete_database_button(self):
        self.client.login(username='admin_xlsx', password='Pass12345!')

        response = self.client.get(reverse('admin_data_tools'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Удалить базу данных')
        self.assertContains(response, 'name="pas_key_data"')
        self.assertContains(response, reverse('admin_guide'))

    def test_staff_user_cannot_open_seed_test_data_tool(self):
        self.client.login(username='staff_xlsx', password='Pass12345!')

        response = self.client.get(reverse('admin_seed_test_data'))

        self.assertEqual(response.status_code, 302)

    def test_staff_user_cannot_delete_database(self):
        self.client.login(username='staff_xlsx', password='Pass12345!')

        response = self.client.post(
            reverse('admin_delete_database'),
            data={
                'confirm_delete': 'yes',
                'pas_key_data': 'rtycds28',
            },
        )

        self.assertEqual(response.status_code, 302)

    @patch('journal.admin_tools.call_command')
    def test_superuser_can_open_seed_test_data_tool_without_running_it(self, mocked_call_command):
        self.client.login(username='admin_xlsx', password='Pass12345!')

        response = self.client.get(reverse('admin_seed_test_data'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Запуск тестовых данных')
        self.assertContains(response, 'Подтверждаю пересоздание тестовых данных')
        self.assertContains(response, 'name="pas_key_data"')
        mocked_call_command.assert_not_called()

    @override_settings(DATA_TOOLS_PASSWORD='rtycds28')
    @patch('journal.admin_tools.call_command')
    def test_superuser_can_run_seed_test_data_tool(self, mocked_call_command):
        self.client.login(username='admin_xlsx', password='Pass12345!')

        response = self.client.post(
            reverse('admin_seed_test_data'),
            data={
                'confirm': 'yes',
                'pas_key_data': 'rtycds28',
            },
        )

        self.assertEqual(response.status_code, 302)
        mocked_call_command.assert_called_once_with('seed_data')

    @override_settings(DATA_TOOLS_PASSWORD='rtycds28')
    @patch('journal.admin_tools.call_command')
    def test_seed_test_data_tool_rejects_wrong_password(self, mocked_call_command):
        self.client.login(username='admin_xlsx', password='Pass12345!')

        response = self.client.post(
            reverse('admin_seed_test_data'),
            data={
                'confirm': 'yes',
                'pas_key_data': 'wrong',
            },
        )

        self.assertEqual(response.status_code, 302)
        mocked_call_command.assert_not_called()

    @override_settings(DATA_TOOLS_PASSWORD='rtycds28')
    def test_superuser_can_delete_database_with_confirmation_password(self):
        self.create_base_journal()
        registration_settings = CourseRegistrationSettings.load()
        registration_settings.telegram_group_url = 'https://t.me/test_group'
        registration_settings.save()
        PasswordRecoveryContact.objects.create(
            name='Администратор',
            phone='+7 (999) 123-45-67',
            messengers='Telegram',
        )
        self.client.login(username='admin_xlsx', password='Pass12345!')

        response = self.client.post(
            reverse('admin_delete_database'),
            data={
                'confirm_delete': 'yes',
                'pas_key_data': 'rtycds28',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(AcademicYear.objects.exists())
        self.assertFalse(Instrument.objects.exists())
        self.assertFalse(Subject.objects.exists())
        self.assertFalse(Teacher.objects.exists())
        self.assertFalse(TeacherEnrollment.objects.exists())
        self.assertFalse(Student.objects.exists())
        self.assertFalse(GroupSubject.objects.exists())
        self.assertFalse(StudentSubject.objects.exists())
        self.assertFalse(Grade.objects.exists())
        self.assertFalse(SubjectResult.objects.exists())
        self.assertFalse(CourseApplication.objects.exists())
        self.assertFalse(CourseRegistrationSettings.objects.exists())
        self.assertFalse(PasswordRecoveryContact.objects.exists())
        self.assertFalse(TemporaryCredential.objects.exists())
        self.assertTrue(User.objects.filter(pk=self.admin_user.pk).exists())
        self.assertTrue(User.objects.filter(pk=self.staff_user.pk).exists())
        self.assertFalse(User.objects.filter(pk=self.regular_user.pk).exists())

    @override_settings(DATA_TOOLS_PASSWORD='rtycds28')
    def test_delete_database_rejects_wrong_password(self):
        self.create_base_journal()
        self.client.login(username='admin_xlsx', password='Pass12345!')

        response = self.client.post(
            reverse('admin_delete_database'),
            data={
                'confirm_delete': 'yes',
                'pas_key_data': 'wrong',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(Student.objects.exists())

    def test_staff_user_cannot_download_full_export(self):
        self.client.login(username='staff_xlsx', password='Pass12345!')

        response = self.client.get(reverse('admin_export_all_data_excel'))

        self.assertEqual(response.status_code, 302)

    def test_superuser_can_download_full_export(self):
        self.client.login(username='admin_xlsx', password='Pass12345!')

        response = self.client.get(reverse('admin_export_all_data_excel'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def test_full_export_rejects_post(self):
        self.client.login(username='admin_xlsx', password='Pass12345!')

        response = self.client.post(reverse('admin_export_all_data_excel'))

        self.assertEqual(response.status_code, 405)

    def test_full_export_escapes_excel_formulas(self):
        Instrument.objects.create(name='=1+1')
        self.client.login(username='admin_xlsx', password='Pass12345!')

        response = self.client.get(reverse('admin_export_all_data_excel'))
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        instrument_values = [
            cell.value
            for row in workbook['Инструменты'].iter_rows()
            for cell in row
        ]

        self.assertIn("'=1+1", instrument_values)

    @override_settings(ENABLE_DESTRUCTIVE_DATA_TOOLS=False)
    def test_destructive_data_tools_are_hidden_and_forbidden_when_disabled(self):
        self.client.login(username='admin_xlsx', password='Pass12345!')

        tools_response = self.client.get(reverse('admin_data_tools'))
        seed_response = self.client.get(reverse('admin_seed_test_data'))
        delete_response = self.client.post(
            reverse('admin_delete_database'),
            data={
                'confirm_delete': 'yes',
                'pas_key_data': 'rtycds28',
            },
        )

        self.assertNotContains(tools_response, 'Запуск тестовых данных')
        self.assertNotContains(tools_response, 'Удалить базу данных')
        self.assertEqual(seed_response.status_code, 403)
        self.assertEqual(delete_response.status_code, 403)


class AccountUtilityTests(JournalTestDataMixin, TestCase):
    def test_build_username_helpers_use_name_and_surname(self):
        self.assertEqual(
            build_display_name_from_full_name('Иванов Иван Иванович'),
            'Иванов Иван',
        )
        self.assertEqual(
            build_username_from_full_name('Иванов Иван Иванович'),
            'Иванов Иван',
        )
        self.assertEqual(
            split_user_name('Иванов Иван Иванович'),
            ('Иван', 'Иванов'),
        )
        self.assertEqual(build_course_application_login('Иванов', 'Иван'), 'Иванов Иван')

    def test_temporary_passwords_are_short_and_easy_to_type(self):
        password = generate_temporary_password()

        self.assertEqual(len(password), 8)
        self.assertLessEqual(
            set(password),
            set('abcdefghjkmnpqrstuvwxyz23456789'),
        )

    def test_display_name_for_user_prefers_student_profile(self):
        group = self.create_group()
        instrument = self.create_instrument()
        user = User.objects.create_user(
            username='tempuser',
            password='Pass12345!',
            first_name='Иван',
            last_name='Иванов',
        )

        Student.objects.create(
            full_name='Иванов Иван Иванович',
            group=group,
            instrument=instrument,
            user=user,
        )

        self.assertEqual(display_name_for_user(user), 'Иванов Иван')

    def test_display_name_for_user_prefers_teacher_profile(self):
        user = User.objects.create_user(
            username='teacher_user',
            password='Pass12345!',
            first_name='Иван',
            last_name='Иванов',
        )

        Teacher.objects.create(full_name='Петров Пётр Петрович', user=user)

        self.assertEqual(display_name_for_user(user), 'Петров Пётр')

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_user_creation_form_accepts_username_with_space(self):
        form = UserCreationForm(
            data={
                'username': 'Админ Тест',
                'password1': 'Pass12345!',
                'password2': 'Pass12345!',
            },
        )

        self.assertTrue(form.is_valid(), form.errors)
        user = form.save()
        user.full_clean()
        self.assertEqual(user.username, 'Админ Тест')

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_user_creation_form_rejects_control_characters_in_username(self):
        form = UserCreationForm(
            data={
                'username': 'Админ\tТест',
                'password1': 'Pass12345!',
                'password2': 'Pass12345!',
            },
        )

        self.assertFalse(form.is_valid())
        self.assertIn('username', form.errors)


class AccountCommandTests(JournalTestDataMixin, TestCase):
    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_create_student_accounts_stores_actual_unique_usernames(self):
        group = self.create_group()
        instrument = self.create_instrument()
        Student.objects.create(
            full_name='Иван Иванов',
            group=group,
            instrument=instrument,
        )
        Student.objects.create(
            full_name='Иван Иванов',
            group=group,
            instrument=instrument,
        )

        call_command('create_student_accounts', stdout=StringIO())

        self.assertEqual(
            list(User.objects.order_by('username').values_list('username', flat=True)),
            ['Иванов Иван', 'Иванов Иван 2'],
        )
        self.assertEqual(
            list(TemporaryCredential.objects.order_by('login').values_list('login', flat=True)),
            ['Иванов Иван', 'Иванов Иван 2'],
        )
        self.assertFalse(TemporaryCredential.objects.filter(user__isnull=True).exists())

        call_command('create_student_accounts', stdout=StringIO())

        self.assertEqual(TemporaryCredential.objects.count(), 2)

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_create_student_accounts_preserves_existing_password_without_credential(self):
        user = User.objects.create_user(
            username='existing student',
            password='ExistingPass123!',
        )
        original_password_hash = user.password
        Student.objects.create(
            full_name='Существующий Ученик',
            group=self.create_group(),
            instrument=self.create_instrument(),
            user=user,
        )

        call_command('create_student_accounts', stdout=StringIO())
        user.refresh_from_db()

        self.assertEqual(user.password, original_password_hash)
        self.assertFalse(TemporaryCredential.objects.filter(user=user).exists())

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_create_teacher_accounts_stores_actual_unique_usernames(self):
        Teacher.objects.create(full_name='Иван Иванов')
        Teacher.objects.create(full_name='Иван Иванов')

        call_command('create_teacher_accounts', stdout=StringIO())

        self.assertEqual(
            list(User.objects.order_by('username').values_list('username', flat=True)),
            ['Иванов Иван', 'Иванов Иван 2'],
        )
        self.assertEqual(
            list(TemporaryCredential.objects.order_by('login').values_list('login', flat=True)),
            ['Иванов Иван', 'Иванов Иван 2'],
        )
        self.assertFalse(TemporaryCredential.objects.filter(user__isnull=True).exists())

        call_command('create_teacher_accounts', stdout=StringIO())

        self.assertEqual(TemporaryCredential.objects.count(), 2)

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_create_teacher_accounts_preserves_existing_password_without_credential(self):
        user = User.objects.create_user(
            username='existing teacher',
            password='ExistingPass123!',
        )
        original_password_hash = user.password
        Teacher.objects.create(
            full_name='Существующий Преподаватель',
            user=user,
        )

        call_command('create_teacher_accounts', stdout=StringIO())
        user.refresh_from_db()

        self.assertEqual(user.password, original_password_hash)
        self.assertFalse(TemporaryCredential.objects.filter(user=user).exists())

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_temporary_password_cannot_be_reassigned_to_existing_user(self):
        user = User.objects.create_user(
            username='immutable temporary password',
            password='InitialPass123!',
        )
        credential = TemporaryCredential.objects.create(
            user=user,
            login=user.username,
            temporary_password='InitialPass123!',
        )
        original_password_hash = user.password

        with self.assertRaisesRegex(
            ValueError,
            'only be stored when a new user is created',
        ):
            ensure_temporary_credential_for_user(
                user,
                password='ReplacementPass123!',
            )

        user.refresh_from_db()
        credential.refresh_from_db()
        self.assertEqual(user.password, original_password_hash)
        self.assertEqual(credential.temporary_password, 'InitialPass123!')
        self.assertTrue(user.check_password('InitialPass123!'))

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_createsuperuser_rolls_back_if_credentials_cannot_be_stored(self):
        with (
            patch.dict(
                'os.environ',
                {'DJANGO_SUPERUSER_PASSWORD': 'AdminTemp123!'},
            ),
            patch(
                'journal.command_overrides.management.commands.createsuperuser.'
                'ensure_temporary_credential_for_user',
                side_effect=RuntimeError('credential failure'),
            ),
        ):
            with self.assertRaisesRegex(RuntimeError, 'credential failure'):
                call_command(
                    'createsuperuser',
                    interactive=False,
                    username='rolled back admin',
                    email='rollback@example.com',
                    stdout=StringIO(),
                )

        self.assertFalse(User.objects.filter(username='rolled back admin').exists())
        self.assertFalse(TemporaryCredential.objects.filter(login='rolled back admin').exists())

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_createsuperuser_stores_temporary_credentials(self):
        with patch.dict(
            'os.environ',
            {'DJANGO_SUPERUSER_PASSWORD': 'AdminTemp123!'},
        ):
            call_command(
                'createsuperuser',
                interactive=False,
                username='created admin',
                email='created-admin@example.com',
                stdout=StringIO(),
            )

        user = User.objects.get(username='created admin')
        credential = TemporaryCredential.objects.get(login='created admin')

        self.assertTrue(user.is_superuser)
        self.assertTrue(user.groups.filter(name='Администратор').exists())
        self.assertEqual(credential.user, user)
        self.assertEqual(credential.temporary_password, 'AdminTemp123!')
        self.assertTrue(user.check_password(credential.temporary_password))


class UserCreationCredentialTests(JournalTestDataMixin, TestCase):
    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_user_admin_creation_stores_temporary_credentials(self):
        admin_user = User.objects.create_superuser(
            username='admin creator',
            password='Pass12345!',
        )
        request = type('Request', (), {'user': admin_user})()
        model_admin = django_admin.site._registry[User]
        form = model_admin.add_form(data={
            'username': 'created in admin',
            'password1': 'AdminCreated123!',
            'password2': 'AdminCreated123!',
        })
        self.assertTrue(form.is_valid(), form.errors)
        user = form.save(commit=False)

        model_admin.save_model(request, user, form, change=False)

        credential = TemporaryCredential.objects.get(user=user)
        self.assertEqual(credential.login, 'created in admin')
        self.assertEqual(credential.temporary_password, 'AdminCreated123!')
        self.assertTrue(user.check_password(credential.temporary_password))

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_user_admin_update_preserves_password_and_syncs_existing_credential(self):
        admin_user = User.objects.create_superuser(
            username='admin editor',
            password='Pass12345!',
        )
        request = type('Request', (), {'user': admin_user})()
        model_admin = django_admin.site._registry[User]
        user = User.objects.create_user(
            username='before edit',
            password='ExistingPass123!',
        )
        credential = TemporaryCredential.objects.create(
            user=user,
            login=user.username,
            temporary_password='ExistingPass123!',
        )
        original_password_hash = user.password

        user.username = 'after edit'
        user.email = 'updated@example.com'
        model_admin.save_model(request, user, form=None, change=True)

        user.refresh_from_db()
        credential.refresh_from_db()
        self.assertEqual(user.password, original_password_hash)
        self.assertEqual(credential.login, 'after edit')
        self.assertEqual(credential.temporary_password, 'ExistingPass123!')
        self.assertTrue(user.check_password('ExistingPass123!'))

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_user_admin_update_does_not_create_temporary_password(self):
        admin_user = User.objects.create_superuser(
            username='admin editor without credential',
            password='Pass12345!',
        )
        request = type('Request', (), {'user': admin_user})()
        model_admin = django_admin.site._registry[User]
        user = User.objects.create_user(
            username='regular account',
            password='ExistingPass123!',
        )
        original_password_hash = user.password

        user.email = 'regular-updated@example.com'
        model_admin.save_model(request, user, form=None, change=True)

        user.refresh_from_db()
        self.assertEqual(user.password, original_password_hash)
        self.assertFalse(TemporaryCredential.objects.filter(user=user).exists())
        self.assertTrue(user.check_password('ExistingPass123!'))

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_ensure_superuser_creation_stores_temporary_credentials(self):
        env = {
            'DJANGO_SUPERUSER_USERNAME': 'container admin',
            'DJANGO_SUPERUSER_EMAIL': 'container-admin@example.com',
            'DJANGO_SUPERUSER_PASSWORD': 'ContainerAdmin123!',
        }
        with patch.dict('os.environ', env, clear=False):
            call_command('ensure_superuser', stdout=StringIO())

        user = User.objects.get(username='container admin')
        credential = TemporaryCredential.objects.get(user=user)
        self.assertTrue(user.is_superuser)
        self.assertEqual(credential.temporary_password, 'ContainerAdmin123!')
        self.assertTrue(user.check_password(credential.temporary_password))

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_ensure_superuser_update_preserves_password_without_rotation(self):
        user = User.objects.create_superuser(
            username='existing managed admin',
            email='old@example.com',
            password='ActualAdmin123!',
        )
        credential = TemporaryCredential.objects.create(
            user=user,
            login=user.username,
            temporary_password='ActualAdmin123!',
        )
        original_password_hash = user.password
        env = {
            'DJANGO_SUPERUSER_USERNAME': user.username,
            'DJANGO_SUPERUSER_EMAIL': 'new@example.com',
            'DJANGO_SUPERUSER_PASSWORD': 'DifferentConfiguredPassword123!',
        }

        with patch.dict('os.environ', env, clear=False):
            call_command('ensure_superuser', stdout=StringIO())

        user.refresh_from_db()
        credential.refresh_from_db()
        self.assertEqual(user.email, 'new@example.com')
        self.assertEqual(user.password, original_password_hash)
        self.assertEqual(credential.temporary_password, 'ActualAdmin123!')
        self.assertTrue(user.check_password('ActualAdmin123!'))
        self.assertFalse(user.check_password('DifferentConfiguredPassword123!'))

    @override_settings(AUTH_PASSWORD_VALIDATORS=[])
    def test_ensure_superuser_does_not_create_credentials_for_existing_user(self):
        user = User.objects.create_superuser(
            username='existing container admin',
            password='ActualAdmin123!',
        )
        original_password_hash = user.password
        env = {
            'DJANGO_SUPERUSER_USERNAME': user.username,
            'DJANGO_SUPERUSER_EMAIL': 'updated-admin@example.com',
            'DJANGO_SUPERUSER_PASSWORD': 'DifferentConfiguredPassword123!',
        }

        with patch.dict('os.environ', env, clear=False):
            call_command('ensure_superuser', stdout=StringIO())

        user.refresh_from_db()
        self.assertEqual(user.email, 'updated-admin@example.com')
        self.assertEqual(user.password, original_password_hash)
        self.assertFalse(TemporaryCredential.objects.filter(user=user).exists())
        self.assertTrue(user.check_password('ActualAdmin123!'))



class CacheConfigurationTests(SimpleTestCase):
    def test_cache_is_disabled_in_non_production_settings(self):
        self.assertFalse(settings.CACHE_ENABLED)
        self.assertEqual(
            settings.CACHES['default']['BACKEND'],
            'django.core.cache.backends.dummy.DummyCache',
        )

    def test_production_cache_source_uses_redis(self):
        source = Path('config/settings.py').read_text(encoding='utf-8')

        self.assertIn('CACHE_ENABLED = IS_PRODUCTION_ENV', source)
        self.assertIn('django.core.cache.backends.redis.RedisCache', source)
        self.assertIn("os.getenv('REDIS_URL', 'redis://redis:6379/1')", source)
        self.assertIn('django.core.cache.backends.dummy.DummyCache', source)


class PerformanceConfigurationTests(SimpleTestCase):
    def test_frequent_queries_have_composite_indexes(self):
        expected_indexes = {
            StudentEnrollment: 'enroll_year_active_group_idx',
            StudentSubject: 'stud_subj_year_active_idx',
            Grade: 'grade_enroll_subj_date_idx',
            SubjectResult: 'result_enroll_subject_idx',
            AssessmentGroup: 'assess_group_year_active_idx',
            PasswordRecoveryContact: 'recovery_active_order_idx',
            CourseApplication: 'course_app_year_status_idx',
        }

        for model, expected_name in expected_indexes.items():
            with self.subTest(model=model.__name__):
                self.assertIn(expected_name, {index.name for index in model._meta.indexes})

    def test_production_compose_provides_healthy_redis(self):
        compose = Path('docker-compose.prod.yml').read_text(encoding='utf-8')
        requirements = Path('requirements.txt').read_text(encoding='utf-8')

        self.assertIn('redis:7.4.10-alpine', compose)
        self.assertIn('["CMD", "redis-cli", "ping"]', compose)
        self.assertIn('allkeys-lru', compose)
        self.assertIn('redis==7.4.0', requirements)

    def test_database_connection_reuse_is_production_only(self):
        source = Path('config/settings.py').read_text(encoding='utf-8')

        self.assertIn("'DB_CONN_MAX_AGE'", source)
        self.assertIn('60 if IS_PRODUCTION_ENV else 0', source)
        self.assertIn("'CONN_HEALTH_CHECKS': IS_PRODUCTION_ENV", source)


@tag('slow', 'seed')
class SeedDataCommandTests(TestCase):
    @staticmethod
    def run_seed_data():
        with TemporaryDirectory() as tmp_dir:
            credentials_path = Path(tmp_dir) / 'secrets.csv'
            call_command(
                'seed_data',
                credentials_output=str(credentials_path),
                stdout=StringIO(),
            )

    @classmethod
    def setUpTestData(cls):
        cls.run_seed_data()

    def test_seed_data_creates_new_architecture_records(self):
        self.assertTrue(CourseRegistrationSettings.objects.filter(academic_year__is_active=True).exists())
        self.assertEqual(PasswordRecoveryContact.objects.count(), 2)
        self.assertEqual(AcademicYear.objects.count(), 2)
        years = list(AcademicYear.objects.order_by('starts_on'))
        archived_year, active_year = years
        self.assertFalse(archived_year.is_active)
        self.assertTrue(active_year.is_active)
        self.assertEqual((archived_year.ends_on - archived_year.starts_on).days, 13)
        self.assertEqual((active_year.ends_on - active_year.starts_on).days, 13)
        self.assertLess(archived_year.ends_on, active_year.starts_on)
        self.assertTrue(archived_year.name.startswith('Арх '))
        self.assertTrue(active_year.name.startswith('Акт '))
        self.assertTrue(Instrument.objects.exists())
        self.assertEqual(OrchestraPart.objects.count(), 10)
        self.assertTrue(
            OrchestraPart.objects.filter(
                instrument__name='Домра',
                name='Малая первая',
            ).exists(),
        )
        self.assertTrue(StudyGroup.objects.exists())
        self.assertTrue(Subject.objects.exists())
        self.assertTrue(Teacher.objects.exists())
        self.assertTrue(Student.objects.exists())
        self.assertTrue(Student.objects.filter(orchestra_part__isnull=False).exists())
        self.assertTrue(Student.objects.filter(custom_instrument='Гусли').exists())
        self.assertTrue(GroupSubject.objects.exists())
        self.assertTrue(StudentSubject.objects.exists())
        self.assertTrue(Grade.objects.exists())
        self.assertTrue(SubjectResult.objects.exists())
        self.assertTrue(CourseApplication.objects.exists())
        self.assertGreaterEqual(AssessmentGroup.objects.count(), 4)
        self.assertGreaterEqual(AssessmentItem.objects.count(), 11)
        self.assertTrue(AssessmentItem.objects.filter(is_required=False).exists())
        self.assertTrue(
            FinalGradeRule.objects.filter(
                rule_type=FinalGradeRule.RULE_ALL_REQUIRED,
            ).exists(),
        )
        self.assertTrue(
            Student.objects.annotate(
                assessment_group_count=Count(
                    'assessment_group_assignments',
                    filter=Q(assessment_group_assignments__is_active=True),
                    distinct=True,
                ),
            ).filter(assessment_group_count__gte=2).exists(),
        )

    def test_seed_data_reuses_same_compact_student_cohort_for_two_years(self):
        archived_year, active_year = AcademicYear.objects.order_by('starts_on')
        main_group_names = {
            'Подготовительная группа',
            '1 класс (начинающие)',
            '2 класс (средний уровень)',
            '3 класс (продвинутые)',
            'Старший ансамбль',
        }
        archived_students = set(
            StudentEnrollment.objects.filter(
                academic_year=archived_year,
                group__name__in=main_group_names,
            ).values_list('student_id', flat=True)
        )
        active_students = set(
            StudentEnrollment.objects.filter(
                academic_year=active_year,
                group__name__in=main_group_names,
            ).values_list('student_id', flat=True)
        )

        self.assertEqual(len(archived_students), 15)
        self.assertEqual(active_students, archived_students)
        self.assertEqual(
            TeacherEnrollment.objects.values('teacher_id')
            .annotate(years=Count('academic_year_id'))
            .filter(years=2)
            .count(),
            Teacher.objects.count(),
        )

    def test_seed_data_creates_course_applications_with_instruments(self):
        applications = CourseApplication.objects.select_related(
            'instrument_reference',
        )

        self.assertTrue(applications.exists())

        for application in applications:
            self.assertTrue(application.instrument)

            if application.instrument_reference_id:
                self.assertEqual(
                    application.instrument,
                    application.instrument_reference.name,
                )
                self.assertEqual(application.custom_instrument, '')
            else:
                self.assertEqual(
                    application.instrument,
                    application.custom_instrument,
                )
                self.assertTrue(application.custom_instrument)

    def test_seed_data_assigns_user_roles(self):
        self.assertTrue(Group.objects.filter(name='Администратор').exists())
        self.assertTrue(Group.objects.filter(name='Преподаватель').exists())
        self.assertTrue(Group.objects.filter(name='Ученик').exists())

        teacher = Teacher.objects.select_related('user').first()
        student = Student.objects.select_related('user').first()

        self.assertIsNotNone(teacher)
        self.assertIsNotNone(student)

        self.assertTrue(
            teacher.user.groups.filter(name='Преподаватель').exists(),
        )
        self.assertTrue(
            student.user.groups.filter(name='Ученик').exists(),
        )

    def test_seed_data_preserves_existing_admin_password_without_fabricating_credential(self):
        admin_user = User.objects.create_superuser(
            username='existing_admin',
            password='OriginalPass123!',
            email='existing-admin@example.com',
        )

        self.run_seed_data()

        admin_user.refresh_from_db()

        self.assertTrue(User.objects.filter(username='existing_admin').exists())
        self.assertTrue(admin_user.is_staff)
        self.assertTrue(admin_user.is_superuser)
        self.assertTrue(
            admin_user.groups.filter(name='Администратор').exists(),
        )
        self.assertTrue(admin_user.check_password('OriginalPass123!'))
        self.assertFalse(TemporaryCredential.objects.filter(user=admin_user).exists())

    @override_settings(IS_PRODUCTION_ENV=True)
    def test_seed_data_is_blocked_in_production_without_explicit_override(self):
        students_before = Student.objects.count()

        with TemporaryDirectory() as tmp_dir:
            with self.assertRaisesMessage(CommandError, 'запрещена в production'):
                call_command(
                    'seed_data',
                    credentials_output=str(Path(tmp_dir) / 'secrets.csv'),
                    stdout=StringIO(),
                )

        self.assertEqual(Student.objects.count(), students_before)

    def test_seed_data_creates_temporary_credentials_for_every_user(self):
        user_logins = set(User.objects.values_list('username', flat=True))
        credential_logins = set(TemporaryCredential.objects.values_list('login', flat=True))
        credential_user_ids = set(TemporaryCredential.objects.values_list('user_id', flat=True))

        self.assertEqual(credential_logins, user_logins)
        self.assertEqual(credential_user_ids, set(User.objects.values_list('id', flat=True)))

    def test_seed_data_has_no_assignment_or_grade_contradictions(self):
        active_year = AcademicYear.objects.get(is_active=True)
        years = list(AcademicYear.objects.order_by('starts_on'))

        self.assertEqual(len(years), 2)
        self.assertFalse(CourseApplication.objects.exclude(academic_year=active_year).exists())
        for academic_year in years:
            with self.subTest(academic_year=academic_year.name):
                self.assertTrue(StudyGroup.objects.filter(academic_year=academic_year).exists())
                self.assertTrue(Grade.objects.filter(academic_year=academic_year).exists())
                self.assertTrue(SubjectResult.objects.filter(academic_year=academic_year).exists())
                self.assertFalse(
                    Grade.objects.filter(academic_year=academic_year).filter(
                        Q(date__lt=academic_year.starts_on)
                        | Q(date__gt=academic_year.ends_on),
                    ).exists(),
                )
                self.assertFalse(
                    StudentEnrollment.objects
                    .filter(academic_year=academic_year, is_active=True, group__isnull=False)
                    .values('group_id')
                    .annotate(total=Count('id'))
                    .filter(total__gt=3)
                    .exists(),
                )

        self.assertFalse(GroupSubject.objects.filter(subject__is_specialty=True).exists())
        self.assertFalse(StudentSubject.objects.filter(subject__is_specialty=False).exists())
        self.assertFalse(
            Student.objects
            .filter(is_active=True)
            .annotate(
                active_individual_subjects=Count(
                    'individual_subjects',
                    filter=Q(
                        individual_subjects__is_active=True,
                        individual_subjects__subject__is_specialty=True,
                    ),
                ),
            )
            .filter(active_individual_subjects=0)
            .exists(),
        )

        group_grade_keys = set(
            GroupSubject.objects
            .filter(is_active=True)
            .values_list('group_id', 'subject_id', 'teacher_id')
        )
        individual_grade_keys = set(
            StudentSubject.objects
            .filter(is_active=True)
            .values_list('student_id', 'subject_id', 'teacher_id')
        )
        invalid_grade_ids = [
            grade_id
            for grade_id, student_id, group_id, subject_id, teacher_id in Grade.objects.values_list(
                'pk',
                'student_id',
                'enrollment__group_id',
                'subject_id',
                'teacher_id',
            )
            if (
                (group_id, subject_id, teacher_id) not in group_grade_keys
                and (student_id, subject_id, teacher_id) not in individual_grade_keys
            )
        ]
        self.assertEqual(invalid_grade_ids, [])

        group_result_keys = set(
            GroupSubject.objects
            .filter(is_active=True)
            .values_list('group_id', 'subject_id')
        )
        individual_result_keys = set(
            StudentSubject.objects
            .filter(is_active=True)
            .values_list('student_id', 'subject_id')
        )
        invalid_result_ids = [
            result_id
            for result_id, student_id, group_id, subject_id in SubjectResult.objects.values_list(
                'pk',
                'student_id',
                'enrollment__group_id',
                'subject_id',
            )
            if (group_id, subject_id) not in group_result_keys and (student_id, subject_id) not in individual_result_keys
        ]
        self.assertEqual(invalid_result_ids, [])

    def test_seed_data_can_be_run_twice_without_duplicate_settings_error(self):
        self.run_seed_data()

        self.assertEqual(
            CourseRegistrationSettings.objects.filter(academic_year__is_active=True).count(),
            1,
        )
        self.assertTrue(Student.objects.exists())
        self.assertTrue(Teacher.objects.exists())

    def test_seed_data_populates_compact_demo_profiles(self):
        self.assertGreaterEqual(Instrument.objects.count(), 14)
        self.assertGreaterEqual(Subject.objects.count(), 21)
        self.assertGreaterEqual(StudyGroup.objects.count(), 13)
        self.assertGreaterEqual(Teacher.objects.count(), 9)
        self.assertGreaterEqual(Student.objects.count(), 17)
        self.assertLessEqual(Student.objects.count(), 20)
        self.assertGreaterEqual(GroupSubject.objects.count(), 60)
        self.assertGreaterEqual(StudentSubject.objects.count(), 60)
        self.assertLessEqual(StudentSubject.objects.count(), 75)
        self.assertGreaterEqual(StudentSubject.objects.filter(subject__is_specialty=True).count(), 60)
        self.assertFalse(GroupSubject.objects.filter(subject__is_specialty=True).exists())
        self.assertFalse(StudentSubject.objects.filter(subject__is_specialty=False).exists())
        expected_grade_count = 0
        enrollments = StudentEnrollment.objects.filter(
            is_active=True,
            group__isnull=False,
        ).select_related('student', 'group', 'academic_year')
        for enrollment in enrollments:
            subject_ids = set(
                GroupSubject.objects.filter(
                    group=enrollment.group,
                    is_active=True,
                ).exclude(
                    subject__assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
                ).values_list('subject_id', flat=True)
            )
            subject_ids.update(
                StudentSubject.objects.filter(
                    student=enrollment.student,
                    academic_year=enrollment.academic_year,
                    is_active=True,
                ).exclude(
                    subject__assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
                ).values_list('subject_id', flat=True)
            )
            expected_grade_count += len(subject_ids) * 3

        self.assertEqual(Grade.objects.count(), expected_grade_count)

        self.assertEqual(
            set(Grade.objects.values_list('value', flat=True)),
            {Grade.GRADE_1, Grade.GRADE_2, Grade.GRADE_3, Grade.GRADE_4, Grade.GRADE_5, Grade.GRADE_ABSENT},
        )
        self.assertFalse(
            Student.objects
            .filter(is_active=True)
            .annotate(grades_count=Count('grades'))
            .filter(grades_count=0)
            .exists(),
        )
        self.assertEqual(
            set(Grade.objects.values_list('academic_year_id', flat=True)),
            set(AcademicYear.objects.values_list('id', flat=True)),
        )
        self.assertFalse(CourseApplication.objects.exclude(academic_year__is_active=True).exists())
        for academic_year in AcademicYear.objects.all():
            self.assertFalse(
                Grade.objects.filter(academic_year=academic_year).filter(
                    Q(date__lt=academic_year.starts_on)
                    | Q(date__gt=academic_year.ends_on),
                ).exists(),
            )
        self.assertEqual(CourseApplication.objects.count(), 3)

        for model in apps.get_app_config('journal').get_models():
            for field in model._meta.concrete_fields:
                if (
                    field.primary_key
                    or field.null
                    or field.blank
                    or field.get_internal_type() not in {
                        'CharField', 'TextField', 'EmailField', 'SlugField', 'URLField',
                    }
                ):
                    continue
                with self.subTest(model=model._meta.label, field=field.name):
                    self.assertFalse(model.objects.filter(**{field.name: ''}).exists())

        registration_settings = CourseRegistrationSettings.objects.get(academic_year__is_active=True)
        self.assertEqual(
            registration_settings.telegram_group_url,
            'https://t.me/cadet_journal_demo',
        )
        self.assertEqual(registration_settings.minimum_registration_age, 14)
        self.assertFalse(hasattr(registration_settings, 'course_starts_on'))
        self.assertFalse(hasattr(registration_settings, 'course_ends_on'))

        self.assertFalse(Teacher.objects.filter(birth_date__isnull=True).exists())
        for field_name in ('phone', 'email', 'comments'):
            self.assertFalse(Teacher.objects.filter(**{field_name: ''}).exists())

        self.assertFalse(Student.objects.filter(birth_date__isnull=True).exists())
        for field_name in (
            'gender',
            'city_church',
            'music_education',
            'student_phone',
            'parent_contacts',
            'comments',
        ):
            self.assertFalse(Student.objects.filter(**{field_name: ''}).exists())

        self.assertTrue(StudyGroup.objects.filter(is_active=False).exists())
        self.assertTrue(GroupSubject.objects.filter(is_active=False).exists())
        self.assertTrue(StudentSubject.objects.filter(is_active=False).exists())
        self.assertFalse(Grade.objects.filter(comment='').exists())
        self.assertGreaterEqual(AssessmentGroup.objects.count(), 4)
        self.assertGreaterEqual(AssessmentItem.objects.count(), 11)
        self.assertTrue(AssessmentItem.objects.filter(is_required=False).exists())
        self.assertTrue(AssessmentResult.objects.filter(status=AssessmentResult.STATUS_PASSED).exists())
        self.assertTrue(AssessmentResult.objects.filter(status=AssessmentResult.STATUS_FAILED).exists())

        self.assertGreaterEqual(
            CourseApplication.objects.filter(
                status=CourseApplication.STATUS_CONFIRMED,
                student__isnull=False,
                user__isnull=False,
            ).count(),
            2,
        )
        self.assertGreaterEqual(
            CourseApplication.objects.filter(
                status=CourseApplication.STATUS_REJECTED,
                student__isnull=True,
                user__isnull=True,
            ).count(),
            1,
        )

        for student in Student.objects.select_related('user'):
            credential = TemporaryCredential.objects.get(login=student.user.username)
            self.assertEqual(credential.user, student.user)
            self.assertEqual(credential.student_phone, student.student_phone)


class ElementAssessmentWorkflowTests(JournalTestDataMixin, TestCase):
    def setUp(self):
        self.year = self.create_academic_year()
        self.group = self.create_group(academic_year=self.year)
        self.subject = Subject.objects.create(
            name='Сдача оркестровых произведений',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )
        self.teacher = self.create_teacher(
            full_name='Дирижёр Первый',
            username='assessment_teacher',
        )
        self.student = self.create_student(
            full_name='Ученик Оркестра',
            group=self.group,
            username='assessment_student',
        )
        GroupSubject.objects.create(
            group=self.group,
            subject=self.subject,
            teacher=self.teacher,
        )
        self.assessment_group = AssessmentGroup.objects.create(
            name='Старший состав',
            subject=self.subject,
            academic_year=self.year,
        )
        self.item = AssessmentItem.objects.create(
            title='Произведение №1',
            subject=self.subject,
            academic_year=self.year,
            group=self.assessment_group,
            responsible_teacher=self.teacher,
        )
        self.assignment = StudentAssessmentGroup.objects.create(
            student=self.student,
            assessment_group=self.assessment_group,
            academic_year=self.year,
        )
        FinalGradeRule.objects.create(
            subject=self.subject,
            academic_year=self.year,
            rule_type=FinalGradeRule.RULE_COUNT,
            passed_count=0,
            grade='N',
            priority=10,
        )
        FinalGradeRule.objects.create(
            subject=self.subject,
            academic_year=self.year,
            rule_type=FinalGradeRule.RULE_COUNT,
            passed_count=1,
            grade='Зачёт',
            priority=10,
        )

    def test_subject_inline_group_name_is_a_strict_database_dropdown(self):
        another_group = AssessmentGroup.objects.create(
            name='Камерный состав',
            subject=self.subject,
            academic_year=self.year,
        )
        form_class = type(
            'SubjectScopedAssessmentGroupForm',
            (AssessmentGroupForSubjectAdminForm,),
            {
                'parent_subject': self.subject,
                'parent_academic_year': self.year,
            },
        )

        form = form_class(instance=another_group)

        self.assertIsInstance(form.fields['name'].widget, forms.Select)
        self.assertEqual(
            {value for value, _label in form.fields['name'].choices if value},
            {'Старший состав', 'Камерный состав'},
        )
        self.assertNotIsInstance(form.fields['name'].widget, forms.TextInput)

    def test_group_item_uses_catalog_element_and_snapshots_its_text(self):
        catalog_item = AssessmentElement.objects.create(
            subject=self.subject,
            title='Каталожное произведение',
            description='Описание из справочника',
        )

        placement = AssessmentItem.objects.create(
            element=catalog_item,
            subject=self.subject,
            academic_year=self.year,
            group=self.assessment_group,
            responsible_teacher=self.teacher,
        )

        self.assertEqual(placement.title, catalog_item.title)
        self.assertEqual(placement.description, catalog_item.description)
        self.assertEqual(placement.element, catalog_item)

    def test_item_admin_form_limits_elements_to_group_subject(self):
        other_subject = Subject.objects.create(
            name='Другой оркестровый предмет',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
        )
        allowed = AssessmentElement.objects.create(
            subject=self.subject,
            title='Допустимое произведение',
        )
        AssessmentElement.objects.create(
            subject=other_subject,
            title='Чужое произведение',
        )

        form_class = type(
            'GroupScopedAssessmentItemAdminForm',
            (AssessmentItemAdminForm,),
            {'parent_assessment_group': self.assessment_group},
        )
        form = form_class()

        self.assertEqual(list(form.fields['element'].queryset), [allowed])
        self.assertTrue(form.fields['element'].required)

    def test_student_sees_only_items_from_assigned_assessment_groups(self):
        other_group = AssessmentGroup.objects.create(
            name='Другой состав',
            subject=self.subject,
            academic_year=self.year,
        )
        AssessmentItem.objects.create(
            title='Чужое произведение',
            subject=self.subject,
            academic_year=self.year,
            group=other_group,
            responsible_teacher=self.teacher,
        )

        items = list(available_assessment_items_for_student(self.student, self.year))

        self.assertEqual(items, [self.item])

    def test_work_group_assignment_does_not_require_a_separate_subject_assignment(self):
        orchestra_subject = Subject.objects.create(
            name='Отдельная оркестровая программа',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )
        orchestra_group = AssessmentGroup.objects.create(
            name='Сводный оркестр',
            subject=orchestra_subject,
            academic_year=self.year,
        )
        orchestra_item = AssessmentItem.objects.create(
            title='Произведение без учебного назначения',
            subject=orchestra_subject,
            academic_year=self.year,
            group=orchestra_group,
            responsible_teacher=self.teacher,
        )
        other_study_group = self.create_group(
            name='Учебная группа без оркестрового предмета',
            academic_year=self.year,
        )
        orchestra_student = self.create_student(
            full_name='Ученик сводного оркестра',
            group=other_study_group,
            instrument=self.student.instrument,
            username='independent_orchestra_student',
        )

        assignment = StudentAssessmentGroup.objects.create(
            student=orchestra_student,
            assessment_group=orchestra_group,
            academic_year=self.year,
        )

        self.assertEqual(assignment.enrollment.student, orchestra_student)
        self.assertEqual(
            list(available_assessment_items_for_student(orchestra_student, self.year)),
            [orchestra_item],
        )
        self.assertTrue(SubjectResult.objects.filter(
            student=orchestra_student,
            subject=orchestra_subject,
            academic_year=self.year,
            is_auto_calculated=True,
        ).exists())

        self.client.force_login(orchestra_student.user)
        response = self.client.get(
            reverse('journal'),
            {'academic_year': self.year.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [section['subject'] for section in response.context['assessment_subject_sections']],
            [orchestra_subject],
        )
        self.assertContains(response, orchestra_item.title)
        self.assertContains(response, orchestra_group.name)

        assignment.delete()
        self.assertFalse(
            available_assessment_items_for_student(orchestra_student, self.year).exists()
        )
        self.assertTrue(SubjectResult.objects.filter(
            student=orchestra_student,
            subject=orchestra_subject,
            academic_year=self.year,
            is_auto_calculated=True,
        ).exists())

    def test_assessment_item_can_be_moved_to_another_subject_group(self):
        other_subject = Subject.objects.create(
            name='Новая оркестровая программа',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )
        other_group = AssessmentGroup.objects.create(
            name='Новая группа произведений',
            subject=other_subject,
            academic_year=self.year,
        )
        other_element = AssessmentElement.objects.create(
            subject=other_subject,
            title=self.item.title,
            description=self.item.description,
        )
        form = AssessmentItemAdminForm(
            data={
                'element': other_element.pk,
                'subject': self.subject.pk,
                'academic_year': self.year.pk,
                'group': other_group.pk,
                'responsible_teacher': self.teacher.pk,
                'sort_order': self.item.sort_order,
                'is_required': 'on',
                'is_active': 'on',
            },
            instance=self.item,
        )

        self.assertTrue(form.is_valid(), form.errors)
        moved_item = form.save()
        self.assertEqual(moved_item.group, other_group)
        self.assertEqual(moved_item.subject, other_subject)
        self.assertEqual(moved_item.academic_year, self.year)

    def test_admin_options_include_all_active_year_students_and_teachers(self):
        unrelated_group = self.create_group(
            name='Группа без предмета оркестра',
            academic_year=self.year,
        )
        unrelated_student = self.create_student(
            full_name='Ученик без предмета оркестра',
            group=unrelated_group,
            instrument=self.student.instrument,
            username='assessment_options_unrelated_student',
        )
        unrelated_teacher = self.create_teacher(
            full_name='Преподаватель без связи с предметом',
            username='assessment_options_unrelated_teacher',
        )
        TeacherEnrollment.objects.update_or_create(
            teacher=unrelated_teacher,
            academic_year=self.year,
            defaults={'is_active': True},
        )
        superuser = User.objects.create_superuser(
            username='all_assessment_options_admin',
            password='AdminPass123!',
        )
        self.client.force_login(superuser)

        student_response = self.client.get(reverse('assessment_options_api'), {
            'type': 'student_group',
            'assessment_group': self.assessment_group.pk,
            'academic_year': self.year.pk,
        })
        teacher_response = self.client.get(reverse('assessment_options_api'), {
            'type': 'item',
            'group': self.assessment_group.pk,
            'academic_year': self.year.pk,
        })

        self.assertEqual(student_response.status_code, 200)
        self.assertEqual(teacher_response.status_code, 200)
        self.assertIn(
            unrelated_student.pk,
            {item['id'] for item in student_response.json()['students']},
        )
        self.assertIn(
            unrelated_teacher.pk,
            {item['id'] for item in teacher_response.json()['teachers']},
        )

    def test_teacher_result_updates_automatic_string_final(self):
        result = set_assessment_result(
            item=self.item,
            student=self.student,
            acting_teacher=self.teacher,
            status=AssessmentResult.STATUS_PASSED,
            comment='Сдано уверенно',
        )

        final = SubjectResult.objects.get(
            student=self.student,
            subject=self.subject,
            academic_year=self.year,
        )
        self.assertEqual(result.comment, 'Сдано уверенно')
        self.assertEqual(final.final_grade, 'Зачёт')
        self.assertTrue(final.is_auto_calculated)
        self.assertEqual(final.calculation_details['passed_count'], 1)

    def test_old_teacher_loses_edit_access_but_result_author_is_preserved(self):
        result = set_assessment_result(
            item=self.item,
            student=self.student,
            acting_teacher=self.teacher,
            status=AssessmentResult.STATUS_PASSED,
        )
        new_teacher = self.create_teacher(
            full_name='Дирижёр Второй',
            username='assessment_teacher_2',
        )
        TeacherEnrollment.objects.update_or_create(
            teacher=new_teacher,
            academic_year=self.year,
            defaults={'is_active': True},
        )
        TeacherSubject.objects.create(teacher=new_teacher, subject=self.subject)
        self.item.responsible_teacher = new_teacher
        self.item.save()

        with self.assertRaises(PermissionDenied):
            set_assessment_result(
                item=self.item,
                student=self.student,
                acting_teacher=self.teacher,
                status=AssessmentResult.STATUS_FAILED,
            )

        result.refresh_from_db()
        self.assertEqual(result.assessed_by, self.teacher)

    def test_removing_group_keeps_result_and_recalculates_current_final(self):
        result = set_assessment_result(
            item=self.item,
            student=self.student,
            acting_teacher=self.teacher,
            status=AssessmentResult.STATUS_PASSED,
        )

        self.assignment.delete()

        self.assertTrue(AssessmentResult.objects.filter(pk=result.pk).exists())
        final = SubjectResult.objects.get(
            student=self.student,
            subject=self.subject,
            academic_year=self.year,
        )
        self.assertEqual(final.final_grade, 'N')
        self.assertEqual(final.calculation_details['total_count'], 0)

    def test_manual_final_for_element_mode_is_rejected(self):
        with self.assertRaises(ValidationError):
            SubjectResult.objects.create(
                student=self.student,
                subject=self.subject,
                academic_year=self.year,
                final_grade='5',
            )

    def test_arbitrary_grade_string_is_preserved_for_standard_subject(self):
        standard_subject = self.create_subject(name='Теория музыки')
        GroupSubject.objects.create(
            group=self.group,
            subject=standard_subject,
            teacher=self.teacher,
        )

        grade = Grade.objects.create(
            student=self.student,
            subject=standard_subject,
            teacher=self.teacher,
            academic_year=self.year,
            date=date(2025, 10, 10),
            value='5+',
        )

        self.assertEqual(grade.value, '5+')

    def test_assessment_group_admin_exposes_cascading_workspaces(self):
        superuser = User.objects.create_superuser(
            username='assessment_admin',
            email='assessment-admin@example.com',
            password='AdminPass123!',
        )
        request = RequestFactory().get('/admin/', {'academic_year': self.year.pk})
        request.user = superuser
        request.session = {}
        model_admin = django_admin.site._registry[AssessmentGroup]

        inline_instances = model_admin.get_inline_instances(request, self.assessment_group)
        inline_names = {type(inline).__name__ for inline in inline_instances}

        self.assertEqual(
            inline_names,
            {
                'AssessmentItemForGroupInline',
                'StudentAssessmentGroupForGroupInline',
                'FinalGradeRuleForGroupInline',
            },
        )

        student_inline = next(
            inline for inline in inline_instances
            if type(inline).__name__ == 'StudentAssessmentGroupForGroupInline'
        )
        formset = student_inline.get_formset(request, self.assessment_group)(
            instance=self.assessment_group,
        )
        self.assertIn(self.student, formset.forms[-1].fields['student'].queryset)
        self.assertEqual(student_inline.extra, 0)

    def test_assessment_workspace_fields_offer_values_from_related_tables(self):
        other_subject = Subject.objects.create(
            name='Другая сдача',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )
        other_group = AssessmentGroup.objects.create(
            name='Группа другого предмета',
            subject=other_subject,
            academic_year=self.year,
        )
        superuser = User.objects.create_superuser(
            username='assessment_dropdown_admin',
            email='assessment-dropdown@example.com',
            password='AdminPass123!',
        )
        request = RequestFactory().get('/admin/', {'academic_year': self.year.pk})
        request.user = superuser
        request.session = {}

        subject_admin = django_admin.site._registry[Subject]
        subject_inlines = subject_admin.get_inline_instances(request, self.subject)
        group_inline = next(
            inline for inline in subject_inlines
            if type(inline).__name__ == 'AssessmentGroupForSubjectInline'
        )
        group_formset = group_inline.get_formset(request, self.subject)(instance=self.subject)
        group_name_html = str(group_formset.empty_form['name'])
        self.assertIn('<datalist', group_name_html)
        self.assertIn(self.assessment_group.name, group_name_html)
        self.assertNotIn(other_group.name, group_name_html)

        rule_inline = next(
            inline for inline in subject_inlines
            if type(inline).__name__ == 'FinalGradeRuleForSubjectInline'
        )
        rule_formset = rule_inline.get_formset(request, self.subject)(instance=self.subject)
        rule_groups = rule_formset.empty_form.fields['assessment_group'].queryset
        self.assertIn(self.assessment_group, rule_groups)
        self.assertNotIn(other_group, rule_groups)

        group_admin = django_admin.site._registry[AssessmentGroup]
        item_inline = next(
            inline for inline in group_admin.get_inline_instances(request, self.assessment_group)
            if type(inline).__name__ == 'AssessmentItemForGroupInline'
        )
        item_formset = item_inline.get_formset(request, self.assessment_group)(
            instance=self.assessment_group,
        )
        item_title_html = str(item_formset.empty_form['title'])
        self.assertIn('<datalist', item_title_html)
        self.assertIn(self.item.title, item_title_html)

    def test_academic_year_admin_exposes_cascading_workspaces(self):
        superuser = User.objects.create_superuser(
            username='academic_year_workspace_admin',
            email='year-workspace@example.com',
            password='AdminPass123!',
        )
        request = RequestFactory().get('/admin/')
        request.user = superuser
        request.session = {}
        model_admin = django_admin.site._registry[AcademicYear]

        inline_instances = model_admin.get_inline_instances(request, self.year)
        inline_names = {type(inline).__name__ for inline in inline_instances}

        self.assertEqual(
            inline_names,
            {
                'StudyGroupForAcademicYearInline',
                'TeacherEnrollmentForAcademicYearInline',
                'UserAcademicYearMembershipForAcademicYearInline',
                'AssessmentGroupForAcademicYearInline',
            },
        )
        group_inline = next(
            inline for inline in inline_instances
            if type(inline).__name__ == 'AssessmentGroupForAcademicYearInline'
        )
        formset = group_inline.get_formset(request, self.year)(instance=self.year)
        self.assertIn(self.subject, formset.forms[-1].fields['subject'].queryset)

    def test_assignment_admin_links_to_filtered_items_and_results(self):
        superuser = User.objects.create_superuser(
            username='assignment_workspace_admin',
            email='assignment-workspace@example.com',
            password='AdminPass123!',
        )
        request = RequestFactory().get('/admin/', {'academic_year': self.year.pk})
        request.user = superuser
        request.session = {}
        model_admin = django_admin.site._registry[StudentAssessmentGroup]
        assignment = model_admin.get_queryset(request).get(pk=self.assignment.pk)

        items_link = str(model_admin.items_count_display(assignment))
        results_link = str(model_admin.results_workspace_link(assignment))

        self.assertIn('assessmentitem', items_link)
        self.assertIn(f'group__id__exact={self.assessment_group.pk}', items_link)
        self.assertIn('assessmentresult', results_link)
        self.assertIn(f'enrollment__student__id__exact={self.student.pk}', results_link)

    def test_assessment_item_admin_exposes_filtered_result_inline(self):
        superuser = User.objects.create_superuser(
            username='assessment_result_admin',
            email='result-admin@example.com',
            password='AdminPass123!',
        )
        request = RequestFactory().get('/admin/', {'academic_year': self.year.pk})
        request.user = superuser
        request.session = {}
        model_admin = django_admin.site._registry[AssessmentItem]
        inline = model_admin.get_inline_instances(request, self.item)[0]
        formset = inline.get_formset(request, self.item)(instance=self.item)
        blank_form = formset.empty_form

        self.assertEqual(type(inline).__name__, 'AssessmentResultForItemInline')
        self.assertEqual(inline.extra, 0)
        self.assertIn(self.assignment.enrollment, blank_form.fields['enrollment'].queryset)
        self.assertEqual(blank_form.fields['assessed_by'].initial, self.teacher.pk)

    def test_assessment_result_options_api_returns_only_valid_related_values(self):
        superuser = User.objects.create_superuser(
            username='assessment_options_admin',
            email='options-admin@example.com',
            password='AdminPass123!',
        )
        self.client.force_login(superuser)

        response = self.client.get(
            reverse('assessment_options_api'),
            {'type': 'result', 'item': self.item.pk},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(
            [row['id'] for row in payload['enrollments']],
            [self.assignment.enrollment_id],
        )
        self.assertEqual([row['id'] for row in payload['teachers']], [self.teacher.pk])
        self.assertEqual(payload['defaults']['assessed_by_id'], self.teacher.pk)

    def test_assessment_item_options_do_not_hide_groups_by_teacher_qualification(self):
        other_subject = Subject.objects.create(
            name='Другая оркестровая дисциплина',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
        )
        other_teacher = self.create_teacher(
            full_name='Дирижёр Другой',
            username='assessment_options_other_teacher',
        )
        GroupSubject.objects.create(
            group=self.group,
            subject=other_subject,
            teacher=other_teacher,
        )
        self.assertEqual(list(self.teacher.qualified_subjects.all()), [self.subject])
        self.assertEqual(
            list(Subject.objects.filter(qualified_teachers=self.teacher)),
            [self.subject],
        )
        superuser = User.objects.create_superuser(
            username='assessment_item_options_admin',
            email='item-options-admin@example.com',
            password='AdminPass123!',
        )
        self.client.force_login(superuser)

        response = self.client.get(
            reverse('assessment_options_api'),
            {
                'type': 'item',
                'academic_year': self.year.pk,
                'responsible_teacher': self.teacher.pk,
                'changed': 'responsible_teacher',
                'strict': '1',
            },
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(
            {row['id'] for row in payload['subjects']},
            {self.subject.pk, other_subject.pk},
        )
        self.assertIn(
            self.teacher.pk,
            [row['id'] for row in payload['teachers']],
        )

    def test_teacher_assessment_filters_include_group_and_individual_subjects(self):
        individual_subject = Subject.objects.create(
            name='Индивидуальная сдача партий',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
            is_specialty=True,
        )
        StudentSubject.objects.create(
            student=self.student,
            subject=individual_subject,
            teacher=self.teacher,
        )
        individual_group = AssessmentGroup.objects.create(
            name='Индивидуальная программа',
            subject=individual_subject,
            academic_year=self.year,
        )
        individual_item = AssessmentItem.objects.create(
            title='Индивидуальное произведение',
            subject=individual_subject,
            academic_year=self.year,
            group=individual_group,
            responsible_teacher=self.teacher,
        )
        StudentAssessmentGroup.objects.create(
            student=self.student,
            assessment_group=individual_group,
            academic_year=self.year,
        )
        self.client.force_login(self.teacher.user)

        response = self.client.get(
            reverse('assessment_filter_options_api'),
            {'academic_year': self.year.pk},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertCountEqual(
            [row['id'] for row in payload['subjects']],
            [self.subject.pk, individual_subject.pk],
        )
        self.assertCountEqual(
            [row['id'] for row in payload['items']],
            [self.item.pk, individual_item.pk],
        )
        self.assertNotIn('study_groups', payload)
        self.assertEqual(
            [row['id'] for row in payload['students']],
            [self.student.pk],
        )

        filtered_response = self.client.get(
            reverse('assessment_filter_options_api'),
            {
                'academic_year': self.year.pk,
                'assessment_group': individual_group.pk,
            },
        )
        filtered_payload = filtered_response.json()
        self.assertEqual(
            [row['id'] for row in filtered_payload['subjects']],
            [individual_subject.pk],
        )
        self.assertEqual(
            [row['id'] for row in filtered_payload['items']],
            [individual_item.pk],
        )

    def test_assessment_group_is_the_only_primary_dependency_filter(self):
        other_study_group = self.create_group(
            name='Другая учебная группа',
            academic_year=self.year,
        )
        other_subject = Subject.objects.create(
            name='Другая оркестровая программа',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )
        other_student = self.create_student(
            full_name='Ученик Другого Оркестра',
            group=other_study_group,
            instrument=self.student.instrument,
            username='other_assessment_student',
        )
        GroupSubject.objects.create(
            group=other_study_group,
            subject=other_subject,
            teacher=self.teacher,
        )
        other_assessment_group = AssessmentGroup.objects.create(
            name='Другой оркестр',
            subject=other_subject,
            academic_year=self.year,
        )
        other_item = AssessmentItem.objects.create(
            title='Чужая программа',
            subject=other_subject,
            academic_year=self.year,
            group=other_assessment_group,
            responsible_teacher=self.teacher,
        )
        StudentAssessmentGroup.objects.create(
            student=other_student,
            assessment_group=other_assessment_group,
            academic_year=self.year,
        )
        self.client.force_login(self.teacher.user)

        all_response = self.client.get(
            reverse('assessment_filter_options_api'),
            {
                'academic_year': self.year.pk,
                'assessment_subject': other_subject.pk,
                'assessment_item': other_item.pk,
                'assessment_student': other_student.pk,
            },
        )
        all_payload = all_response.json()
        self.assertEqual(
            {row['id'] for row in all_payload['assessment_groups']},
            {self.assessment_group.pk, other_assessment_group.pk},
        )
        self.assertEqual(
            {row['id'] for row in all_payload['subjects']},
            {self.subject.pk, other_subject.pk},
        )
        self.assertEqual(
            {row['id'] for row in all_payload['items']},
            {self.item.pk, other_item.pk},
        )
        self.assertEqual(
            {row['id'] for row in all_payload['students']},
            {self.student.pk, other_student.pk},
        )

        group_response = self.client.get(
            reverse('assessment_filter_options_api'),
            {
                'academic_year': self.year.pk,
                'assessment_group': self.assessment_group.pk,
                'assessment_subject': other_subject.pk,
                'assessment_item': other_item.pk,
                'assessment_student': other_student.pk,
                'changed': 'assessment_group',
            },
        )
        group_payload = group_response.json()
        self.assertEqual([row['id'] for row in group_payload['subjects']], [self.subject.pk])
        self.assertEqual([row['id'] for row in group_payload['items']], [self.item.pk])
        self.assertEqual([row['id'] for row in group_payload['students']], [self.student.pk])

    def test_teacher_assessment_item_filter_limits_rendered_sections(self):
        other_item = AssessmentItem.objects.create(
            title='Произведение №2',
            subject=self.subject,
            academic_year=self.year,
            group=self.assessment_group,
            responsible_teacher=self.teacher,
            sort_order=20,
        )
        self.client.force_login(self.teacher.user)

        response = self.client.get(
            reverse('journal'),
            {
                'academic_year': self.year.pk,
                'assessment_item': other_item.pk,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [section['item'] for section in response.context['assessment_sections']],
            [other_item],
        )
        self.assertNotContains(response, 'name="assessment_study_group"')
        self.assertContains(response, 'name="assessment_group"')
        self.assertContains(response, 'name="assessment_student"')

    def test_admin_can_filter_and_save_assessment_result_from_journal(self):
        superuser = User.objects.create_superuser(
            username='assessment_journal_admin',
            email='assessment-journal-admin@example.com',
            password='AdminPass123!',
        )
        self.client.force_login(superuser)

        get_response = self.client.get(
            reverse('journal'),
            {
                'academic_year': self.year.pk,
                'assessment_teacher': self.teacher.pk,
                'assessment_subject': self.subject.pk,
            },
        )

        self.assertEqual(get_response.status_code, 200)
        self.assertEqual(len(get_response.context['assessment_sections']), 1)
        self.assertContains(get_response, 'name="assessment_teacher"')
        self.assertContains(
            get_response,
            (
                '<input type="hidden" name="assessment_teacher" '
                f'value="{self.teacher.pk}">'
            ),
            html=True,
        )
        self.assertContains(
            get_response,
            (
                '<input type="hidden" name="assessment_subject" '
                f'value="{self.subject.pk}">'
            ),
            html=True,
        )

        post_response = self.client.post(
            (
                f'{reverse("journal")}?academic_year={self.year.pk}'
                f'&assessment_teacher={self.teacher.pk}'
            ),
            {
                'action': 'assessment_result',
                'item_id': self.item.pk,
                'student_id': self.student.pk,
                'status': AssessmentResult.STATUS_PASSED,
                'comment': 'Проверено администратором',
            },
        )

        self.assertEqual(post_response.status_code, 302)
        self.assertTrue(
            AssessmentResult.objects.filter(
                item=self.item,
                enrollment=self.assignment.enrollment,
                assessed_by=self.teacher,
                status=AssessmentResult.STATUS_PASSED,
            ).exists(),
        )

    def test_teacher_can_save_result_with_quick_assessment_form(self):
        self.client.force_login(self.teacher.user)
        query = (
            f'?academic_year={self.year.pk}'
            f'&assessment_group={self.assessment_group.pk}'
        )

        get_response = self.client.get(f'{reverse("journal")}{query}')

        self.assertEqual(get_response.status_code, 200)
        self.assertContains(get_response, 'id="quick-assessment-title"')
        self.assertContains(get_response, 'data-save-context="quick-assessment"')
        self.assertContains(get_response, 'data-save-context="assessment-')
        self.assertContains(get_response, 'name="assessment_group"')
        self.assertContains(get_response, 'name="assessment_item"')
        self.assertContains(get_response, 'name="assessment_student"')
        self.assertContains(get_response, 'name="status"')
        self.assertContains(get_response, 'id="quick_assessment_comment"')
        self.assertContains(get_response, 'name="comment"')
        self.assertContains(get_response, 'Результат и комментарий')
        self.assertNotContains(get_response, 'Текущий результат')
        self.assertNotContains(get_response, 'name="assessment_study_group"')

        post_response = self.client.post(
            f'{reverse("journal")}{query}',
            {
                'action': 'assessment_result',
                'assessment_group': self.assessment_group.pk,
                'assessment_item': self.item.pk,
                'assessment_student': self.student.pk,
                'status': AssessmentResult.STATUS_FAILED,
                'comment': 'Нужно доработать вступление',
            },
        )

        self.assertRedirects(
            post_response,
            f'{reverse("journal")}{query}',
            fetch_redirect_response=False,
        )
        result = AssessmentResult.objects.get(
            item=self.item,
            enrollment=self.assignment.enrollment,
        )
        self.assertEqual(result.status, AssessmentResult.STATUS_FAILED)
        self.assertEqual(result.comment, 'Нужно доработать вступление')

        refreshed_response = self.client.get(f'{reverse("journal")}{query}')
        self.assertContains(refreshed_response, 'Незачёт: 1')
        self.assertContains(refreshed_response, 'data-flash-message')
        self.assertContains(refreshed_response, 'data-message-level="success"')

    def test_quick_assessment_error_is_rendered_for_local_red_toast(self):
        self.client.force_login(self.teacher.user)
        response = self.client.post(
            f'{reverse("journal")}?academic_year={self.year.pk}',
            {
                'action': 'assessment_result',
                'assessment_group': self.assessment_group.pk,
                'assessment_item': 999999,
                'assessment_student': self.student.pk,
                'status': AssessmentResult.STATUS_PASSED,
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-flash-message')
        self.assertContains(response, 'data-message-level="error"')
        self.assertContains(response, 'Не удалось определить произведение или ученика.')
        self.assertFalse(AssessmentResult.objects.exists())


    def test_teacher_journal_shows_compact_assessment_summary(self):
        set_assessment_result(
            item=self.item,
            student=self.student,
            acting_teacher=self.teacher,
            status=AssessmentResult.STATUS_PASSED,
        )
        self.client.force_login(self.teacher.user)

        response = self.client.get(
            reverse('journal'),
            {'academic_year': self.year.pk, 'group': self.group.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['assessment_summary']['item_count'], 1)
        self.assertEqual(response.context['assessment_summary']['student_count'], 1)
        self.assertEqual(response.context['assessment_summary']['passed_count'], 1)
        self.assertContains(response, 'Сводка по сдаче произведений')
        self.assertContains(response, 'Не оценено: 0')
        self.assertContains(response, 'Свернуть все таблицы сдачи произведений')
        self.assertContains(response, 'Развернуть все таблицы сдачи произведений')
        self.assertContains(response, 'data-collapse-target="teacher-assessment-blocks"', count=1)
        self.assertContains(response, 'id="teacher-assessment-blocks"')
        self.assertContains(
            response,
            (
                '<details class="table-card collapsible-card" '
                f'data-collapse-key="assessment-item-{self.teacher.user_id}-'
                f'{self.year.pk}-{self.item.pk}" open>'
            ),
            html=False,
        )
        self.assertContains(response, 'Произведение / элемент:')

    def test_student_journal_groups_assessment_progress_by_subject(self):
        set_assessment_result(
            item=self.item,
            student=self.student,
            acting_teacher=self.teacher,
            status=AssessmentResult.STATUS_PASSED,
        )
        optional_item = AssessmentItem.objects.create(
            title='Дополнительное произведение',
            subject=self.subject,
            academic_year=self.year,
            group=self.assessment_group,
            responsible_teacher=self.teacher,
            is_required=False,
            sort_order=20,
        )
        self.client.force_login(self.student.user)

        response = self.client.get(
            reverse('journal'),
            {'academic_year': self.year.pk},
        )

        self.assertEqual(response.status_code, 200)
        sections = response.context['assessment_subject_sections']
        self.assertEqual(len(sections), 1)
        self.assertEqual(sections[0]['subject'], self.subject)
        self.assertEqual(sections[0]['item_count'], 2)
        self.assertEqual(sections[0]['passed_count'], 1)
        self.assertEqual(sections[0]['not_evaluated_count'], 1)
        self.assertEqual(sections[0]['progress_percent'], 50)
        self.assertContains(response, optional_item.title)
        self.assertContains(response, 'Обязательных сдано: 1 / 1')
        self.assertContains(
            response,
            (
                '<details class="table-card assessment-subject-card '
                'collapsible-card" '
                f'data-collapse-key="student-assessment-{self.student.user_id}-'
                f'{self.year.pk}-{self.subject.pk}" open>'
            ),
            html=False,
        )

    def test_student_collapse_all_uses_reliable_single_button_script(self):
        self.client.force_login(self.student.user)

        response = self.client.get(
            reverse('journal'),
            {'academic_year': self.year.pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'journal/collapsible_tables.js')
        self.assertNotContains(response, 'journal/collapse_controls.js')
        self.assertContains(
            response,
            'data-collapse-target="student-assessment-blocks"',
            count=1,
        )
        self.assertNotContains(response, 'data-collapse-action="collapse"')

        script = Path(
            'journal/static/journal/collapsible_tables.js'
        ).read_text(encoding='utf-8')
        self.assertIn(
            'var open = details.length > 0 && details.every',
            script,
        )
        self.assertIn("control.setAttribute('aria-expanded'", script)
        self.assertNotIn('target.hidden', script)


class SelectedAcademicYearExportTests(JournalTestDataMixin, TestCase):
    def test_export_uses_requested_archive_year_and_exact_results_columns(self):
        old_year = self.create_academic_year(name='2025/2026')
        group = self.create_group(name='Архивная группа', academic_year=old_year)
        student = self.create_student(
            full_name='Архивный Ученик',
            group=group,
            username='archive_export_student',
        )
        subject = self.create_subject(name='Архивный предмет')
        teacher = self.create_teacher(
            full_name='Архивный Преподаватель',
            username='archive_export_teacher',
        )
        GroupSubject.objects.create(group=group, subject=subject, teacher=teacher)
        SubjectResult.objects.create(
            student=student,
            subject=subject,
            academic_year=old_year,
            exam_grade='4+',
            final_grade='5-',
        )
        self.create_academic_year(name='2026/2027')

        workbook = build_full_export_workbook(old_year)
        results_sheet = workbook['Итоги']

        self.assertEqual(
            [cell.value for cell in results_sheet[1]],
            ['Ученик', 'Предмет', 'Экзамен', 'Итоговая оценка', 'Группа', 'Учебный год'],
        )
        self.assertEqual(results_sheet['A2'].value, 'Архивный Ученик')
        self.assertEqual(results_sheet['C2'].value, '4+')
        self.assertEqual(results_sheet['D2'].value, '5-')
        self.assertEqual(results_sheet['F2'].value, '2025/2026')
        self.assertEqual(results_sheet['C2'].number_format, '@')
        self.assertEqual(results_sheet['D2'].number_format, '@')

    def test_export_uses_student_snapshot_from_selected_year(self):
        old_year = self.create_academic_year(name='2025/2026')
        group = self.create_group(name='Архивная группа', academic_year=old_year)
        student = self.create_student(
            full_name='Имя в архиве',
            group=group,
            username='student_snapshot_export',
        )
        Student.objects.filter(pk=student.pk).update(
            full_name='Текущее имя',
            student_phone='+79990000000',
            comments='Текущий комментарий',
        )

        workbook = build_full_export_workbook(old_year)
        students_sheet = workbook['Ученики']
        headers = {
            cell.value: cell.column
            for cell in students_sheet[1]
        }

        self.assertEqual(students_sheet.cell(2, headers['Ученик']).value, 'Имя в архиве')
        self.assertEqual(students_sheet.cell(2, headers['Телефон ученика']).value, '')
        self.assertEqual(students_sheet.cell(2, headers['Комментарий']).value, '')

    def test_export_includes_orchestra_part_from_selected_year_snapshot(self):
        old_year = self.create_academic_year(name='2025/2026')
        group = self.create_group(name='Архивная группа', academic_year=old_year)
        student = self.create_student(
            full_name='Оркестровый Ученик',
            group=group,
            username='orchestra_export_student',
        )
        archive_part = OrchestraPart.objects.create(
            instrument=student.instrument,
            name='Партия первого тенора',
        )
        current_part = OrchestraPart.objects.create(
            instrument=student.instrument,
            name='Текущая партия',
        )
        student.orchestra_part = archive_part
        student.save()
        self.create_academic_year(name='2026/2027')
        Student.objects.filter(pk=student.pk).update(orchestra_part=current_part)

        workbook = build_full_export_workbook(old_year)
        students_sheet = workbook['Ученики']
        headers = {cell.value: cell.column for cell in students_sheet[1]}

        self.assertEqual(
            students_sheet.cell(2, headers['Партия в оркестре']).value,
            'Партия первого тенора',
        )

    def test_export_includes_teacher_temporary_credentials_with_owner_details(self):
        old_year = self.create_academic_year(name='2025/2026')
        teacher = self.create_teacher(
            full_name='Преподаватель Экспорта',
            username='teacher_credentials_export',
        )
        TemporaryCredential.objects.create(
            user=teacher.user,
            login=teacher.user.username,
            temporary_password='TeacherExport123!',
        )

        workbook = build_full_export_workbook(old_year)
        credentials_sheet = workbook['Временные доступы']

        self.assertEqual(
            [cell.value for cell in credentials_sheet[1]],
            [
                'ФИО',
                'Роль',
                'Логин',
                'Временный пароль',
                'Телефон ученика',
                'Дата выдачи',
                'Учебный год',
            ],
        )
        self.assertEqual(credentials_sheet['A2'].value, 'Преподаватель Экспорта')
        self.assertEqual(credentials_sheet['B2'].value, 'Преподаватель')
        self.assertEqual(credentials_sheet['C2'].value, 'teacher_credentials_export')
        self.assertEqual(credentials_sheet['G2'].value, '2025/2026')

    def test_export_includes_new_catalog_and_administration_sheets(self):
        year = self.create_academic_year(name='2025/2026')
        group = self.create_group(name='Оркестровая группа', academic_year=year)
        subject = Subject.objects.create(
            name='Сдача партий',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
            final_grade_type=Subject.FINAL_GRADE_TYPE_NUMERIC,
        )
        teacher = self.create_teacher(
            full_name='Дирижёр Экспорта',
            username='export_conductor',
        )
        TeacherSubject.objects.create(teacher=teacher, subject=subject)
        GroupSubject.objects.create(group=group, subject=subject, teacher=teacher)
        element = AssessmentElement.objects.create(
            subject=subject,
            title='Марш для проверки',
            description='Каталожное описание',
        )
        settings_obj = CourseRegistrationSettings.objects.get(academic_year=year)
        settings_obj.telegram_group_url = 'https://t.me/example_group'
        settings_obj.minimum_registration_age = 14
        settings_obj.registration_mode = CourseRegistrationSettings.REGISTRATION_MODE_OPEN
        settings_obj.save()
        PasswordRecoveryContact.objects.create(
            name='Администратор Экспорта',
            phone='+7 (900) 000-00-00',
            messengers='Telegram',
            messenger_username='export_admin',
        )

        workbook = build_full_export_workbook(year)

        self.assertIn('Каталог произведений', workbook.sheetnames)
        self.assertIn('Квалификации преподавателей', workbook.sheetnames)
        self.assertIn('Доступ к учебному году', workbook.sheetnames)
        self.assertIn('Настройки регистрации', workbook.sheetnames)
        self.assertIn('Контакты восстановления', workbook.sheetnames)
        self.assertEqual(workbook['Каталог произведений']['A2'].value, element.title)
        self.assertEqual(
            workbook['Квалификации преподавателей']['A2'].value,
            teacher.full_name,
        )
        self.assertEqual(
            workbook['Настройки регистрации']['A2'].value,
            settings_obj.academic_year.name,
        )
        self.assertEqual(
            workbook['Контакты восстановления']['D1'].value,
            'Имя пользователя в Telegram',
        )
        self.assertEqual(
            workbook['Контакты восстановления']['D2'].value,
            'export_admin',
        )
        membership_rows = list(
            workbook['Доступ к учебному году'].iter_rows(min_row=2, values_only=True)
        )
        self.assertIn(
            ('export_conductor', '', 'Да', '2025/2026'),
            membership_rows,
        )

    def test_assessment_item_export_preserves_catalog_reference_and_snapshot(self):
        year = self.create_academic_year(name='2025/2026')
        subject = Subject.objects.create(
            name='Оркестровая практика',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
            final_grade_type=Subject.FINAL_GRADE_TYPE_NUMERIC,
        )
        group = AssessmentGroup.objects.create(
            name='Основная программа',
            subject=subject,
            academic_year=year,
        )
        element = AssessmentElement.objects.create(
            subject=subject,
            title='Симфония № 1',
            description='Исходное описание',
        )
        item = AssessmentItem.objects.create(
            element=element,
            group=group,
            subject=subject,
            academic_year=year,
            sort_order=20,
        )
        AssessmentElement.objects.filter(pk=element.pk).update(title='Новое название')

        workbook = build_full_export_workbook(year)
        sheet = workbook['Произведения']
        headers = {cell.value: cell.column for cell in sheet[1]}

        self.assertEqual(sheet.cell(2, headers['Произведение']).value, item.title)
        self.assertEqual(
            sheet.cell(2, headers['Запись справочника']).value,
            'Новое название',
        )
        self.assertEqual(sheet.cell(2, headers['Порядок отображения']).value, 20)

    def test_active_year_export_includes_administrator_account(self):
        active_year = self.create_academic_year(name='2025/2026')
        administrator = User.objects.create_superuser(
            username='export_administrator',
            password='Pass12345!',
            first_name='Администратор',
        )
        UserAcademicYearMembership.objects.create(
            user=administrator,
            academic_year=active_year,
        )

        workbook = build_full_export_workbook(active_year)
        users_sheet = workbook['Пользователи']
        rows = list(users_sheet.iter_rows(min_row=2, values_only=True))

        self.assertIn(
            ('export_administrator', 'Администратор', '', 'Администратор'),
            rows,
        )

class ExportCommandsCompatibilityTests(JournalTestDataMixin, TestCase):
    """
    Эти тесты можно оставить, если в проекте сохранены management-команды
    export_temporary_credentials и export_student_credentials_with_phone.

    Если команды удалены, удалите этот класс из tests.py.
    """

    def test_export_temporary_credentials_command_outputs_csv_if_command_exists(
        self,
    ):
        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='Temp12345!',
        ):
            CourseApplication.objects.create(**self.application_payload())

        output = StringIO()

        try:
            call_command('export_temporary_credentials', stdout=output)
        except CommandError:  # pragma: no cover
            self.skipTest(
                'Команда export_temporary_credentials не найдена в проекте.',
            )

        csv_output = output.getvalue()

        self.assertIn('role,name,login,temporary_password,created_at,phone', csv_output)
        self.assertIn('student', csv_output)
        self.assertIn('login', csv_output)
        self.assertIn('temporary_password', csv_output)
        self.assertIn('Иванов Иван', csv_output)
        self.assertIn('Temp12345!', csv_output)

    def test_export_student_credentials_with_phone_command_outputs_csv_if_command_exists(
        self,
    ):
        with patch(
            'journal.account_utils.generate_temporary_password',
            return_value='Temp12345!',
        ):
            CourseApplication.objects.create(**self.application_payload())

        with TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / 'export.csv'

            try:
                call_command(
                    'export_student_credentials_with_phone',
                    output=str(output_path),
                )
            except CommandError:  # pragma: no cover
                self.skipTest(
                    'Команда export_student_credentials_with_phone '
                    'не найдена в проекте.',
                )

            csv_output = output_path.read_text(encoding='utf-8')

        self.assertIn('login', csv_output)
        self.assertIn('temporary_password', csv_output)
        self.assertIn('student_phone', csv_output)
        self.assertIn('Иванов Иван', csv_output)
        self.assertIn('Temp12345!', csv_output)
        self.assertIn('+7 (999) 123-45-67', csv_output)


class RelatedResultAndErrorHandlingTests(JournalTestDataMixin, TestCase):
    def setUp(self):
        self.year = self.create_academic_year()
        self.group = self.create_group(academic_year=self.year)
        self.subject = self.create_subject()
        self.teacher = self.create_teacher()
        self.student = self.create_student(group=self.group)
        self.create_group_assignment(
            group=self.group,
            subject=self.subject,
            teacher=self.teacher,
        )
        self.result = SubjectResult.objects.create(
            student=self.student,
            subject=self.subject,
            academic_year=self.year,
            exam_grade='4+',
            final_grade='5-',
        )

    def test_related_subject_result_label_contains_exam_and_final_grades(self):
        model_admin = django_admin.site._registry[SubjectResult]
        label = model_admin.get_related_record_label(self.result)
        self.assertIn('Экзамен: 4+', label)
        self.assertIn('Итоговая оценка: 5-', label)

    @override_settings(DEBUG=False)
    def test_custom_404_page_contains_actionable_message_and_request_id(self):
        response = self.client.get('/definitely-missing-page/')
        self.assertEqual(response.status_code, 404)
        self.assertContains(response, 'Страница не найдена', status_code=404)
        self.assertContains(response, 'Код ошибки:', status_code=404)
        self.assertTrue(response.headers.get('X-Request-ID'))

    @override_settings(DEBUG=False)
    def test_custom_404_json_has_stable_error_shape(self):
        response = self.client.get(
            '/definitely-missing-api/',
            HTTP_ACCEPT='application/json',
        )
        self.assertEqual(response.status_code, 404)
        payload = response.json()
        self.assertFalse(payload['success'])
        self.assertEqual(payload['error']['status'], 404)
        self.assertEqual(payload['error']['code'], 'http_404')
        self.assertEqual(payload['error']['request_id'], response.headers['X-Request-ID'])

    def test_error_log_cleanup_keeps_only_newest_1000_records(self):
        ErrorLog.objects.bulk_create(
            [
                ErrorLog(level='ERROR', logger_name='test', message=f'Ошибка {index}')
                for index in range(1005)
            ],
            batch_size=250,
        )
        deleted = ErrorLog.prune_old_entries()
        self.assertEqual(deleted, 5)
        self.assertEqual(ErrorLog.objects.count(), 1000)
        self.assertFalse(ErrorLog.objects.filter(message='Ошибка 0').exists())
        self.assertTrue(ErrorLog.objects.filter(message='Ошибка 1004').exists())

    def test_database_error_handler_stores_request_context(self):
        from journal.error_logging import DatabaseErrorHandler

        request = RequestFactory().get('/journal/test-error/')
        request.request_id = 'request-test-1234'
        request.user = User.objects.create_user(username='error-user')
        record = logging.LogRecord(
            name='journal.test',
            level=logging.ERROR,
            pathname=__file__,
            lineno=1,
            msg='Проверочная ошибка',
            args=(),
            exc_info=None,
        )
        record.request = request
        record.status_code = 500

        DatabaseErrorHandler(max_records=1000).emit(record)

        saved = ErrorLog.objects.get(message='Проверочная ошибка')
        self.assertEqual(saved.request_id, 'request-test-1234')
        self.assertEqual(saved.status_code, 500)
        self.assertEqual(saved.path, '/journal/test-error/')
        self.assertEqual(saved.user_label, 'error-user')


class CabinetAndDependencyRegressionTests(JournalTestDataMixin, TestCase):
    def setUp(self):
        self.year = self.create_academic_year(name='2026/2027')
        self.group = self.create_group(name='Основная группа', academic_year=self.year)
        self.instrument = self.create_instrument(name='Домра')

    def test_student_admin_embeds_parts_for_every_instrument(self):
        domra_part = OrchestraPart.objects.create(
            instrument=self.instrument,
            name='Первая домра',
        )
        bayan = self.create_instrument(name='Баян')
        bayan_part = OrchestraPart.objects.create(
            instrument=bayan,
            name='Первый баян',
        )

        form = StudentAdminForm()
        field = form.fields['orchestra_part']
        parts_map = json.loads(field.widget.attrs['data-orchestra-parts-map'])

        self.assertEqual(
            parts_map[str(self.instrument.pk)],
            [{'id': domra_part.pk, 'name': domra_part.name}],
        )
        self.assertEqual(
            parts_map[str(bayan.pk)],
            [{'id': bayan_part.pk, 'name': bayan_part.name}],
        )
        self.assertFalse(field.disabled)
        self.assertIn(domra_part, field.queryset)
        self.assertIn(bayan_part, field.queryset)
        self.assertIn(
            'journal/orchestra_part_dependencies_v5.js',
            tuple(str(item) for item in form.media._js),
        )

    def test_student_admin_page_contains_local_part_map_and_v5_script(self):
        OrchestraPart.objects.create(
            instrument=self.instrument,
            name='Вторая домра',
        )
        superuser = User.objects.create_superuser(
            username='instrument_admin',
            password='AdminPass123!',
        )
        self.client.force_login(superuser)

        response = self.client.get(reverse('admin:journal_student_add'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-orchestra-parts-map=')
        self.assertContains(response, 'Вторая домра')
        self.assertContains(response, 'journal/orchestra_part_dependencies_v5.js')

    def test_admin_and_registration_share_instrument_dependency_contract(self):
        OrchestraPart.objects.create(
            instrument=self.instrument,
            name='Контрактная партия',
        )

        admin_form = StudentAdminForm()
        registration_form = CourseApplicationPublicForm()

        for form, instrument_name in (
            (admin_form, 'instrument'),
            (registration_form, 'instrument_reference'),
        ):
            instrument_field = form.fields[instrument_name]
            custom_field = form.fields['custom_instrument']
            part_field = form.fields['orchestra_part']

            self.assertEqual(instrument_field.label, 'Инструмент')
            self.assertEqual(instrument_field.empty_label, 'Другой инструмент')
            self.assertEqual(
                instrument_field.widget.attrs['data-instrument-reference'],
                '1',
            )
            self.assertEqual(custom_field.label, 'Собственный инструмент')
            self.assertEqual(custom_field.widget.attrs['data-custom-instrument'], '1')
            self.assertEqual(part_field.label, 'Партия в оркестре')
            self.assertEqual(part_field.widget.attrs['data-orchestra-part'], '1')
            self.assertEqual(
                part_field.widget.attrs['data-native-dependent-select'],
                '1',
            )

        self.assertEqual(
            admin_form.fields['orchestra_part'].widget.attrs['data-orchestra-parts-map'],
            registration_form.fields['orchestra_part'].widget.attrs['data-orchestra-parts-map'],
        )
        self.assertIn(
            'journal/orchestra_part_dependencies_v5.js',
            tuple(str(item) for item in admin_form.media._js),
        )
        self.assertIn(
            'journal/orchestra_part_dependencies_v5.js',
            tuple(str(item) for item in registration_form.media._js),
        )

    def test_student_admin_form_saves_selected_instrument_part(self):
        part = OrchestraPart.objects.create(
            instrument=self.instrument,
            name='Сохраняемая партия',
        )
        form = StudentAdminForm(data={
            'full_name': 'Проверочный Ученик',
            'gender': Student.GENDER_MALE,
            'birth_date': '2010-01-01',
            'city_church': '',
            'music_education': Student.MUSIC_EDUCATION_SELF,
            'student_phone': '',
            'parent_contacts': '',
            'comments': '',
            'group': self.group.pk,
            'instrument': self.instrument.pk,
            'custom_instrument': '',
            'orchestra_part': part.pk,
            'is_active': 'on',
            'user': '',
        })

        self.assertTrue(form.is_valid(), form.errors.as_json())
        student = form.save()
        self.assertEqual(student.instrument, self.instrument)
        self.assertEqual(student.orchestra_part, part)
        self.assertEqual(student.custom_instrument, '')

    def test_v5_orchestra_script_keeps_admin_part_select_native_and_refreshable(self):
        source = Path(
            'journal/static/journal/orchestra_part_dependencies_v5.js'
        ).read_text(encoding='utf-8')

        self.assertIn('ensureNativeDependentSelect', source)
        self.assertIn("wrapped.select2('destroy')", source)
        self.assertIn("field.classList.add('journal-native-dependent-select')", source)
        self.assertIn('data.orchestraPartsMap', source.replace('dataset', 'data'))
        self.assertIn("select2:select.journalOrchestraParts", source)
        self.assertIn("refreshFromServer(instrumentId", source)

    def test_mobile_journal_layout_contains_width_guards_for_cards_and_tabs(self):
        css = Path(
            'journal/static/journal/layout-mobile.css'
        ).read_text(encoding='utf-8')
        template = Path('templates/journal.html').read_text(encoding='utf-8')

        self.assertIn('body.journal-workspace .assessment-subject-card', css)
        self.assertIn('body.journal-workspace .journal-tabs', css)
        self.assertIn('contain: inline-size', css)
        self.assertIn('grid-template-columns: minmax(0, 1fr)', css)
        self.assertEqual(
            template.count(
                '<form method="post" class="table-form" '
                'data-save-context="journal-table-{{ table.academic_year.id }}-'
                '{{ table.group.id }}-{{ table.subject.id }}">'
            ),
            1,
        )

    def test_responsive_table_script_ignores_detached_tables(self):
        source = Path(
            'journal/static/journal/responsive_tables.js'
        ).read_text(encoding='utf-8')

        self.assertIn('!table.isConnected', source)
        self.assertIn('const parent = table.parentElement;', source)
        self.assertIn('if (!parent || !parent.isConnected)', source)
        self.assertNotIn('table.parentNode.insertBefore', source)
        self.assertIn('requestAnimationFrame', source)

    def test_admin_and_student_cabinets_keep_stable_workspace_scope(self):
        subject = Subject.objects.create(
            name='Оркестровая программа',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )
        teacher = self.create_teacher(username='workspace_teacher')
        student = self.create_student(
            full_name='Стабильный Ученик',
            group=self.group,
            instrument=self.instrument,
            username='workspace_student',
        )
        self.create_group_assignment(
            group=self.group,
            subject=subject,
            teacher=teacher,
        )
        assessment_group = AssessmentGroup.objects.create(
            name='Программа ученика',
            subject=subject,
            academic_year=self.year,
        )
        item = AssessmentItem.objects.create(
            title='Проверочное произведение',
            subject=subject,
            academic_year=self.year,
            group=assessment_group,
            responsible_teacher=teacher,
        )
        StudentAssessmentGroup.objects.create(
            student=student,
            assessment_group=assessment_group,
            academic_year=self.year,
        )

        admin_user = User.objects.create_superuser(
            username='workspace_admin',
            password='AdminPass123!',
        )
        self.client.force_login(admin_user)
        admin_response = self.client.get(reverse('journal'), {
            'academic_year': self.year.pk,
        })
        self.assertEqual(admin_response.status_code, 200)
        self.assertContains(
            admin_response,
            'class="journal-workspace journal-workspace--superuser"',
        )
        self.assertContains(admin_response, 'journal/responsive_overflow.css')

        self.client.force_login(student.user)
        student_response = self.client.get(reverse('journal'), {
            'academic_year': self.year.pk,
        })
        self.assertEqual(student_response.status_code, 200)
        self.assertContains(
            student_response,
            'class="journal-workspace journal-workspace--student"',
        )
        self.assertContains(student_response, item.title)
        self.assertContains(student_response, 'journal/responsive_overflow.css')


class TeacherLinkedDataRegressionTests(JournalTestDataMixin, TestCase):
    def setUp(self):
        self.year = self.create_academic_year(name='2026/2027')
        self.group = self.create_group(name='Оркестровый класс', academic_year=self.year)
        self.instrument = self.create_instrument(name='Скрипка')
        self.student = self.create_student(
            full_name='Ученик Связанный Данными',
            group=self.group,
            instrument=self.instrument,
            username='linked_data_student',
        )
        self.linked_teacher = self.create_teacher(
            full_name='Связанный Преподаватель',
            username='linked_data_teacher',
        )
        self.responsible_teacher = self.create_teacher(
            full_name='Ответственный Дирижёр',
            username='responsible_data_teacher',
        )
        self.subject = Subject.objects.create(
            name='Оркестровые произведения',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )
        GroupSubject.objects.create(
            group=self.group,
            subject=self.subject,
            teacher=self.linked_teacher,
        )
        self.assessment_group = AssessmentGroup.objects.create(
            name='Основная программа',
            subject=self.subject,
            academic_year=self.year,
        )
        self.item = AssessmentItem.objects.create(
            title='Связанное произведение',
            subject=self.subject,
            academic_year=self.year,
            group=self.assessment_group,
            responsible_teacher=self.responsible_teacher,
        )
        StudentAssessmentGroup.objects.create(
            student=self.student,
            assessment_group=self.assessment_group,
            academic_year=self.year,
        )

    def test_teacher_can_view_item_linked_through_group_subject(self):
        self.assertTrue(
            assessment_items_for_teacher(
                self.linked_teacher,
                self.year,
            ).filter(pk=self.item.pk).exists()
        )

        sections = assessment_sections_for_teacher(
            self.linked_teacher,
            self.year,
        )
        section = next(item for item in sections if item['item'].pk == self.item.pk)
        self.assertFalse(section['can_edit'])
        self.assertEqual(
            [row['student'].pk for row in section['rows']],
            [self.student.pk],
        )

    def test_linked_teacher_sees_read_only_data_in_personal_cabinet(self):
        self.client.force_login(self.linked_teacher.user)

        response = self.client.get(reverse('journal'), {
            'academic_year': self.year.pk,
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.item.title)
        self.assertContains(response, 'Только просмотр')
        self.assertNotContains(
            response,
            f'data-save-context="assessment-{self.item.pk}-{self.student.pk}"',
        )

    def test_responsible_teacher_retains_edit_access(self):
        sections = assessment_sections_for_teacher(
            self.responsible_teacher,
            self.year,
        )
        section = next(item for item in sections if item['item'].pk == self.item.pk)
        self.assertTrue(section['can_edit'])

    def test_responsible_teacher_sees_direct_group_assignment_without_subject_link(self):
        GroupSubject.objects.filter(
            group=self.group,
            subject=self.subject,
        ).delete()

        sections = assessment_sections_for_teacher(
            self.responsible_teacher,
            self.year,
        )

        section = next(item for item in sections if item['item'].pk == self.item.pk)
        self.assertTrue(section['can_edit'])
        self.assertEqual(
            [row['student'].pk for row in section['rows']],
            [self.student.pk],
        )


class SharedProfilesAcrossYearsRegressionTests(JournalTestDataMixin, TestCase):
    def test_student_and_teacher_profiles_are_reused_across_years(self):
        archived_year = self.create_academic_year(name='2025/2026')
        archived_group = self.create_group(
            name='Архивная группа',
            academic_year=archived_year,
        )
        instrument = self.create_instrument(name='Флейта')
        student = self.create_student(
            full_name='Многолетний Ученик',
            group=archived_group,
            instrument=instrument,
            username='multi_year_student_profile',
        )
        teacher = self.create_teacher(
            full_name='Многолетний Преподаватель',
            username='multi_year_teacher_profile',
        )

        active_year = self.create_academic_year(name='2026/2027')
        active_group = self.create_group(
            name='Активная группа',
            academic_year=active_year,
        )
        student.group = active_group
        student.save()
        TeacherEnrollment.objects.create(
            teacher=teacher,
            academic_year=active_year,
            is_active=True,
        )

        self.assertEqual(Student.objects.filter(pk=student.pk).count(), 1)
        self.assertEqual(Teacher.objects.filter(pk=teacher.pk).count(), 1)
        self.assertEqual(
            set(student.enrollments.values_list('academic_year_id', flat=True)),
            {archived_year.pk, active_year.pk},
        )
        self.assertEqual(
            set(teacher.academic_year_memberships.values_list('academic_year_id', flat=True)),
            {archived_year.pk, active_year.pk},
        )
        self.assertIn(StudentEnrollmentHistoryInline, StudentAdmin.inlines)
        self.assertIn(TeacherEnrollmentHistoryInline, TeacherAdmin.inlines)


class TeacherAccessAndErrorLoggingRegressionTests(JournalTestDataMixin, TestCase):
    def setUp(self):
        self.data = self.create_base_journal()

    def test_teacher_sees_assigned_grade_even_when_legacy_owner_differs(self):
        lesson_date = date(2025, 10, 10)
        grade = Grade.objects.create(
            student=self.data['student'],
            subject=self.data['solfeggio'],
            teacher=self.data['teacher'],
            academic_year=self.data['year'],
            date=lesson_date,
            value='4',
        )
        # Simulate a legacy/admin import where the author field no longer
        # matches the current assignment. Visibility is assignment-based.
        Grade.objects.filter(pk=grade.pk).update(
            teacher=self.data['other_teacher'],
        )

        self.client.force_login(self.data['teacher'].user)
        response = self.client.get(
            reverse('journal'),
            {'academic_year': self.data['year'].pk},
        )

        self.assertEqual(response.status_code, 200)
        table = next(
            item
            for item in response.context['journal_tables']
            if item['subject'].pk == self.data['solfeggio'].pk
        )
        row = next(
            item
            for item in table['rows']
            if item['student'].pk == self.data['student'].pk
        )
        self.assertEqual(row['grades_by_date'][lesson_date], '4')

    def test_teacher_can_edit_only_currently_assigned_grade_cells(self):
        lesson_date = date(2025, 10, 10)
        grade = Grade.objects.create(
            student=self.data['student'],
            subject=self.data['solfeggio'],
            teacher=self.data['teacher'],
            academic_year=self.data['year'],
            date=lesson_date,
            value='4',
        )
        Grade.objects.filter(pk=grade.pk).update(
            teacher=self.data['other_teacher'],
        )

        self.client.force_login(self.data['teacher'].user)
        response = self.client.post(
            f"{reverse('journal')}?academic_year={self.data['year'].pk}",
            {
                'action': 'inline_edit',
                (
                    f'grade__{self.data["solfeggio"].pk}__'
                    f'{self.data["student"].pk}__{lesson_date.isoformat()}'
                ): '5',
            },
        )

        self.assertEqual(response.status_code, 302)
        grade.refresh_from_db()
        self.assertEqual(grade.value, '5')

    def test_readonly_archived_enrollment_form_does_not_raise_missing_field_value_error(self):
        old_enrollment = StudentEnrollment.objects.get(
            student=self.data['student'],
            academic_year=self.data['year'],
        )
        self.create_academic_year(name='2026/2027')
        old_enrollment.refresh_from_db()

        class EnrollmentHistoryForm(AcademicYearHistoryInlineForm):
            class Meta:
                model = StudentEnrollment
                fields = ()

        form = EnrollmentHistoryForm(data={}, instance=old_enrollment)

        self.assertTrue(form.is_valid(), form.errors.as_json())

    def test_unhandled_exception_is_written_to_error_log(self):
        request = RequestFactory().post('/admin/journal/student/419/change/')
        request.user = self.data['teacher'].user
        request.request_id = 'student-change-error'
        middleware = ErrorLoggingMiddleware(lambda current_request: HttpResponse())

        middleware.process_exception(
            request,
            ValueError("'StudentEnrollmentForm' has no field named 'academic_year'."),
        )

        entry = ErrorLog.objects.get(request_id='student-change-error')
        self.assertEqual(entry.status_code, 500)
        self.assertEqual(entry.path, '/admin/journal/student/419/change/')
        self.assertIn('StudentEnrollmentForm', entry.message)
        self.assertIn('ValueError', entry.exception)
        self.assertFalse(entry.metadata['handled'])

    def test_handled_http_error_response_is_written_to_error_log(self):
        request = RequestFactory().post('/journal/handled-error/')
        request.user = self.data['teacher'].user
        request.request_id = 'handled-response-error'
        middleware = ErrorLoggingMiddleware(lambda current_request: HttpResponse())

        response = middleware.process_response(
            request,
            HttpResponse('Ошибка данных', status=409),
        )

        self.assertEqual(response.status_code, 409)
        entry = ErrorLog.objects.get(request_id='handled-response-error')
        self.assertEqual(entry.status_code, 409)
        self.assertTrue(entry.metadata['handled'])

    def test_admin_and_registration_reject_same_mismatched_orchestra_part(self):
        selected_instrument = self.data['instrument']
        another_instrument = self.create_instrument(name='Домра')
        wrong_part = OrchestraPart.objects.create(
            instrument=another_instrument,
            name='Первая домра',
        )

        admin_form = StudentAdminForm(data={
            'full_name': 'Проверочный Ученик',
            'gender': Student.GENDER_MALE,
            'birth_date': '2010-01-01',
            'city_church': '',
            'music_education': Student.MUSIC_EDUCATION_SELF,
            'student_phone': '',
            'parent_contacts': '',
            'comments': '',
            'group': self.data['group'].pk,
            'instrument': selected_instrument.pk,
            'custom_instrument': '',
            'orchestra_part': wrong_part.pk,
            'is_active': 'on',
            'user': '',
        })
        registration_form = CourseApplicationPublicForm(data={
            **self.application_form_payload(
                instrument_reference=selected_instrument,
            ),
            'orchestra_part': wrong_part.pk,
        })

        self.assertFalse(admin_form.is_valid())
        self.assertFalse(registration_form.is_valid())
        self.assertEqual(
            admin_form.errors['orchestra_part'],
            registration_form.errors['orchestra_part'],
        )


class TeacherAssignmentVisibilityAndAssessmentChoiceTests(JournalTestDataMixin, TestCase):
    def setUp(self):
        self.data = self.create_base_journal()

    def test_teacher_keeps_assigned_journal_access_without_membership_helper_rows(self):
        lesson_date = date(2025, 10, 10)
        grade = Grade.objects.create(
            student=self.data['student'],
            subject=self.data['solfeggio'],
            teacher=self.data['teacher'],
            academic_year=self.data['year'],
            date=lesson_date,
            value='4',
        )
        TeacherEnrollment.objects.filter(
            teacher=self.data['teacher'],
            academic_year=self.data['year'],
        ).delete()
        UserAcademicYearMembership.objects.filter(
            user=self.data['teacher'].user,
            academic_year=self.data['year'],
        ).delete()
        self.data['teacher'].refresh_from_db()

        self.client.force_login(self.data['teacher'].user)
        response = self.client.get(
            reverse('journal'),
            {'academic_year': self.data['year'].pk},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.data['solfeggio'].name)
        self.assertTrue(response.context['can_edit_journal'])
        self.assertIn(
            self.data['year'].pk,
            list(academic_year_ids_for_user(self.data['teacher'].user)),
        )

        edit_response = self.client.post(
            f"{reverse('journal')}?academic_year={self.data['year'].pk}",
            {
                'action': 'inline_edit',
                (
                    f'grade__{self.data["solfeggio"].pk}__'
                    f'{self.data["student"].pk}__{lesson_date.isoformat()}'
                ): '5',
            },
        )
        self.assertEqual(edit_response.status_code, 302)
        grade.refresh_from_db()
        self.assertEqual(grade.value, '5')

    def _assessment_catalog(self):
        subject = Subject.objects.create(
            name='Оркестровая практика',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )
        group = AssessmentGroup.objects.create(
            name='Концертная программа',
            subject=subject,
            academic_year=self.data['year'],
        )
        first = AssessmentElement.objects.create(
            subject=subject,
            title='Первое произведение',
        )
        second = AssessmentElement.objects.create(
            subject=subject,
            title='Второе произведение',
        )
        item = AssessmentItem.objects.create(
            element=first,
            subject=subject,
            academic_year=self.data['year'],
            group=group,
            responsible_teacher=self.data['teacher'],
        )
        return subject, group, first, second, item

    def test_assessment_item_field_excludes_element_already_used_in_selected_group(self):
        subject, group, first, second, _item = self._assessment_catalog()
        candidate = AssessmentItem(
            subject=subject,
            academic_year=self.data['year'],
            group=group,
            responsible_teacher=self.data['teacher'],
        )

        form = AssessmentItemAdminForm(instance=candidate)

        self.assertNotIn(first.pk, form.fields['element'].queryset.values_list('pk', flat=True))
        self.assertIn(second.pk, form.fields['element'].queryset.values_list('pk', flat=True))

        reverse_candidate = AssessmentItem(
            element=first,
            subject=subject,
            academic_year=self.data['year'],
            responsible_teacher=self.data['teacher'],
        )
        reverse_form = AssessmentItemAdminForm(instance=reverse_candidate)
        self.assertNotIn(
            group.pk,
            reverse_form.fields['group'].queryset.values_list('pk', flat=True),
        )

    def test_assessment_options_api_excludes_occupied_group_element(self):
        _subject, group, first, second, _item = self._assessment_catalog()
        admin_user = User.objects.create_superuser(
            username='assessment_options_admin',
            password='Pass12345!',
        )
        self.client.force_login(admin_user)

        response = self.client.get(
            reverse('assessment_options_api'),
            {
                'type': 'item',
                'group': group.pk,
                'changed': 'group',
                'strict': '1',
            },
        )

        self.assertEqual(response.status_code, 200)
        element_ids = {item['id'] for item in response.json()['elements']}
        self.assertNotIn(first.pk, element_ids)
        self.assertIn(second.pk, element_ids)

        reverse_response = self.client.get(
            reverse('assessment_options_api'),
            {
                'type': 'item',
                'element': first.pk,
                'changed': 'element',
                'strict': '1',
            },
        )
        self.assertEqual(reverse_response.status_code, 200)
        group_ids = {item['id'] for item in reverse_response.json()['groups']}
        self.assertNotIn(group.pk, group_ids)

    def test_inline_formset_rejects_duplicate_group_element_in_two_new_rows(self):
        subject, group, _first, second, _item = self._assessment_catalog()
        FormSet = inlineformset_factory(
            Teacher,
            AssessmentItem,
            fk_name='responsible_teacher',
            form=AssessmentItemAdminForm,
            formset=AssessmentItemInlineFormSet,
            extra=2,
            can_delete=True,
        )
        prefix = FormSet.get_default_prefix()
        row = {
            'group': str(group.pk),
            'subject': str(subject.pk),
            'academic_year': str(self.data['year'].pk),
            'element': str(second.pk),
            'sort_order': '100',
            'is_required': 'on',
            'is_active': 'on',
        }
        payload = {
            f'{prefix}-TOTAL_FORMS': '2',
            f'{prefix}-INITIAL_FORMS': '0',
            f'{prefix}-MIN_NUM_FORMS': '0',
            f'{prefix}-MAX_NUM_FORMS': '1000',
        }
        for index in range(2):
            for field_name, value in row.items():
                payload[f'{prefix}-{index}-{field_name}'] = value

        formset = FormSet(data=payload, instance=self.data['teacher'], prefix=prefix)

        self.assertFalse(formset.is_valid())
        self.assertTrue(any('element' in form.errors for form in formset.forms))
        self.assertFalse(
            any(
                'unique_assessment_item_group_element' in str(form.errors)
                for form in formset.forms
            )
        )


class AssessmentResultAdminOptionsAndFriendlyErrorsTests(JournalTestDataMixin, TestCase):
    def setUp(self):
        self.year = self.create_academic_year()
        self.study_group = self.create_group(academic_year=self.year)
        self.subject = Subject.objects.create(
            name='Оркестровая сдача',
            assessment_mode=Subject.ASSESSMENT_MODE_ELEMENTS,
            final_grade_type=Subject.FINAL_GRADE_TYPE_PASS_FAIL,
        )
        self.teacher = self.create_teacher(
            full_name='Дирижёр Проверочный',
            username='assessment_result_teacher',
        )
        TeacherEnrollment.objects.get_or_create(
            teacher=self.teacher,
            academic_year=self.year,
            defaults={'is_active': True},
        )
        self.student = self.create_student(
            full_name='Ученик Проверочный',
            group=self.study_group,
            username='assessment_result_student',
        )
        self.assessment_group = AssessmentGroup.objects.create(
            name='Проверочная программа',
            subject=self.subject,
            academic_year=self.year,
        )
        self.item = AssessmentItem.objects.create(
            title='Проверочное произведение',
            subject=self.subject,
            academic_year=self.year,
            group=self.assessment_group,
            responsible_teacher=self.teacher,
        )
        StudentAssessmentGroup.objects.create(
            student=self.student,
            assessment_group=self.assessment_group,
            academic_year=self.year,
        )
        self.enrollment = StudentEnrollment.objects.get(
            student=self.student,
            academic_year=self.year,
        )

    def form_class(self):
        return type(
            'YearScopedAssessmentResultAdminFormForTest',
            (AssessmentResultAdminForm,),
            {'parent_academic_year': self.year},
        )

    def test_result_add_form_has_year_candidates_before_item_selection(self):
        form = self.form_class()()

        self.assertIn(
            self.item.pk,
            form.fields['item'].queryset.values_list('pk', flat=True),
        )
        self.assertIn(
            self.enrollment.pk,
            form.fields['enrollment'].queryset.values_list('pk', flat=True),
        )
        self.assertIn(
            self.teacher.pk,
            form.fields['assessed_by'].queryset.values_list('pk', flat=True),
        )

    def test_result_fields_narrow_to_selected_item(self):
        form = self.form_class()(data={'item': str(self.item.pk)})

        self.assertEqual(
            list(form.fields['enrollment'].queryset.values_list('pk', flat=True)),
            [self.enrollment.pk],
        )
        self.assertEqual(
            list(form.fields['assessed_by'].queryset.values_list('pk', flat=True)),
            [self.teacher.pk],
        )
        self.assertEqual(form.fields['assessed_by'].initial, self.teacher.pk)

    def test_result_options_api_returns_enrollments_and_teachers_without_item(self):
        admin_user = User.objects.create_superuser(
            username='assessment_result_admin',
            password='Pass12345!',
        )
        self.client.force_login(admin_user)

        response = self.client.get(
            reverse('assessment_options_api'),
            {'type': 'result', 'academic_year': self.year.pk},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn(self.enrollment.pk, {row['id'] for row in payload['enrollments']})
        self.assertIn(self.teacher.pk, {row['id'] for row in payload['teachers']})
        self.assertIn(self.item.pk, {row['id'] for row in payload['items']})

    def test_admin_required_errors_have_plain_user_message_and_are_logged(self):
        class RequiredFieldsForm(forms.Form):
            enrollment = forms.ModelChoiceField(
                queryset=StudentEnrollment.objects.all(),
                label='Зачисление ученика',
            )
            assessed_by = forms.ModelChoiceField(
                queryset=Teacher.objects.all(),
                label='Преподаватель, выставивший результат',
            )

        form = RequiredFieldsForm(data={})
        self.assertFalse(form.is_valid())
        message = build_admin_form_user_message(form)

        self.assertEqual(
            message,
            (
                'Не удалось сохранить запись. '
                'Выберите ученика и преподавателя, который выставил результат.'
            ),
        )
        self.assertNotIn('(required)', message)
        self.assertNotIn('Обязательное поле', message)

        request = RequestFactory().post('/admin/journal/assessmentresult/add/')
        request.user = User.objects.create_superuser(
            username='friendly_error_admin',
            password='Pass12345!',
        )
        request.request_id = 'friendly-assessment-error'
        log_handled_error(
            request,
            ValidationError('Форма администратора содержит ошибки.'),
            logger_name='journal.admin.form',
            user_message=message,
        )

        entry = ErrorLog.objects.get(request_id='friendly-assessment-error')
        self.assertEqual(entry.user_message, message)
        self.assertIn('Форма администратора', entry.message)

    def test_default_timezone_is_moscow(self):
        self.assertEqual(settings.TIME_ZONE, 'Europe/Moscow')

    def test_admin_summary_uses_server_friendly_message(self):
        project_root = Path(__file__).resolve().parent.parent
        template = (project_root / 'templates/admin/change_form.html').read_text(encoding='utf-8')
        javascript = (
            project_root / 'journal/static/journal/admin_responsive.js'
        ).read_text(encoding='utf-8')

        self.assertIn('data-user-friendly-error-message', template)
        self.assertIn("form.querySelector('[data-user-friendly-error-message]')", javascript)
