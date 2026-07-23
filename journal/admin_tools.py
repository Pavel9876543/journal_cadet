from __future__ import annotations

import datetime
from secrets import compare_digest
from io import BytesIO
from urllib.parse import quote

from asgiref.sync import sync_to_async
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import user_passes_test
from django.conf import settings
from django.core.management import call_command
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect, render
from django.urls import NoReverseMatch, reverse
from django.utils import timezone
from django.views.decorators.http import require_GET

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .academic_year_context import (
    filter_temporary_credentials_for_year,
    get_selected_admin_academic_year,
)
from .models import (
    AcademicYear,
    CourseApplication,
    CourseRegistrationSettings,
    Grade,
    GroupSubject,
    Instrument,
    PasswordRecoveryContact,
    Student,
    StudentEnrollment,
    StudentSubject,
    StudyGroup,
    Subject,
    SubjectResult,
    Teacher,
    TeacherSubject,
    TemporaryCredential,
)


superuser_required = user_passes_test(lambda user: user.is_active and user.is_superuser)

HEADER_FILL = PatternFill('solid', fgColor='D9EAF7')
HEADER_FONT = Font(bold=True)
DEFAULT_COLUMN_WIDTH = 18
MAX_COLUMN_WIDTH = 60
DATA_TOOLS_PASSWORD_FIELD = 'pas_key_data'


async def _run_db_sync(func, *args, **kwargs):
    database_engine = settings.DATABASES['default']['ENGINE']
    return await sync_to_async(
        func,
        thread_sensitive=database_engine.endswith('sqlite3'),
    )(*args, **kwargs)


@superuser_required
async def admin_data_tools_view(request: HttpRequest) -> HttpResponse:
    return await _run_db_sync(_admin_data_tools_view_sync, request)


@superuser_required
async def admin_guide_view(request: HttpRequest) -> HttpResponse:
    return await _run_db_sync(_admin_guide_view_sync, request)


def _admin_guide_view_sync(request: HttpRequest) -> HttpResponse:
    context = {
        'title': 'Инструкция администратора',
        'journal_url': reverse('journal'),
        'students_url': reverse('admin:journal_student_changelist'),
        'groups_url': reverse('admin:journal_studygroup_changelist'),
        'teachers_url': reverse('admin:journal_teacher_changelist'),
        'subjects_url': reverse('admin:journal_subject_changelist'),
        'grades_url': reverse('admin:journal_grade_changelist'),
        'results_url': reverse('admin:journal_subjectresult_changelist'),
        'applications_url': reverse('admin:journal_courseapplication_changelist'),
        'temporary_credentials_url': reverse('admin:journal_temporarycredential_changelist'),
        'academic_years_url': reverse('admin:journal_academicyear_changelist'),
        'settings_url': reverse('admin:journal_courseregistrationsettings_changelist'),
        'password_contacts_url': reverse('admin:journal_passwordrecoverycontact_changelist'),
        'data_tools_url': reverse('admin_data_tools'),
    }
    return render(request, 'admin/journal/admin_guide.html', context)


def _admin_data_tools_view_sync(request: HttpRequest) -> HttpResponse:
    """
    Страница инструментов данных в Django Admin.

    Здесь администратор может:
    - создать тестовые данные;
    - скачать временные учетные данные;
    - перейти к просмотру временных доступов;
    - скачать полную Excel-выгрузку, если соответствующий URL подключен.
    """
    selected_academic_year = get_selected_admin_academic_year(request)
    temporary_credentials_count = filter_temporary_credentials_for_year(
        TemporaryCredential.objects.all(),
        selected_academic_year,
    ).count()

    context = {
        'title': 'Инструменты данных',
        'temporary_credentials_count': temporary_credentials_count,
        'admin_guide_url': reverse('admin_guide'),
        'temporary_credentials_admin_url': get_admin_url(
            'admin:journal_temporarycredential_changelist',
            fallback='/admin/',
        ),
        'export_credentials_url': reverse('admin_export_test_credentials_excel'),
        'seed_confirm_url': reverse('admin_seed_test_data'),
        'seed_url': reverse('admin_seed_test_data'),
        'delete_database_url': reverse('admin_delete_database'),
        'export_all_data_url': safe_reverse('admin_export_all_data_excel'),
        'data_tools_password_field': DATA_TOOLS_PASSWORD_FIELD,
        'destructive_tools_enabled': settings.ENABLE_DESTRUCTIVE_DATA_TOOLS,
    }

    return render(request, 'admin/journal/data_tools.html', context)


@superuser_required
async def admin_seed_test_data_view(request: HttpRequest) -> HttpResponse:
    return await _run_db_sync(_admin_seed_test_data_view_sync, request)


def _admin_seed_test_data_view_sync(request: HttpRequest) -> HttpResponse:
    """
    Запускает management-команду seed_data из админки.

    Важно:
    - запуск разрешен только суперпользователю;
    - команда вызывается через call_command, то есть без shell-скриптов;
    - для защиты от случайного запуска требуется POST и confirm=yes.
    """
    if not request.user.is_superuser:
        raise PermissionDenied('Создание тестовых данных доступно только суперпользователю.')
    if not settings.ENABLE_DESTRUCTIVE_DATA_TOOLS:
        raise PermissionDenied('Создание тестовых данных отключено в этом окружении.')

    if request.method != 'POST':
        context = {
            'title': 'Запуск тестовых данных',
            'seed_url': reverse('admin_seed_test_data'),
            'data_tools_url': reverse('admin_data_tools'),
            'data_tools_password_field': DATA_TOOLS_PASSWORD_FIELD,
        }
        return render(request, 'admin/journal/seed_data_confirm.html', context)

    if request.POST.get('confirm') != 'yes':
        messages.error(request, 'Подтвердите создание тестовых данных.')
        return redirect('admin_data_tools')

    if not validate_data_tools_password(request):
        return redirect('admin_data_tools')

    try:
        call_command('seed_data')
    except Exception as exc:
        messages.error(request, f'Ошибка при создании тестовых данных: {exc}')
        return redirect('admin_data_tools')

    messages.success(
        request,
        'Тестовые данные успешно созданы. '
        'Временные учетные данные доступны для просмотра в админке.',
    )

    return redirect('admin:journal_temporarycredential_changelist')


@superuser_required
async def admin_delete_database_view(request: HttpRequest) -> HttpResponse:
    return await _run_db_sync(_admin_delete_database_view_sync, request)


def _admin_delete_database_view_sync(request: HttpRequest) -> HttpResponse:
    """
    Очищает данные журнала из админских инструментов.

    Суперпользователи и staff-пользователи сохраняются, чтобы после очистки
    администратор не потерял доступ к панели.
    """
    if not request.user.is_superuser:
        raise PermissionDenied('Удаление базы данных доступно только суперпользователю.')
    if not settings.ENABLE_DESTRUCTIVE_DATA_TOOLS:
        raise PermissionDenied('Удаление данных отключено в этом окружении.')

    if request.method != 'POST':
        return redirect('admin_data_tools')

    if request.POST.get('confirm_delete') != 'yes':
        messages.error(request, 'Подтвердите удаление базы данных.')
        return redirect('admin_data_tools')

    if not validate_data_tools_password(request):
        return redirect('admin_data_tools')

    try:
        deleted_counts = clear_database_data()
    except Exception as exc:
        messages.error(request, f'Ошибка при удалении базы данных: {exc}')
        return redirect('admin_data_tools')

    messages.success(
        request,
        f'База данных очищена. Удалено записей: {sum(deleted_counts.values())}. '
        'Суперпользователи и staff-пользователи сохранены.',
    )

    return redirect('admin_data_tools')


def validate_data_tools_password(request: HttpRequest) -> bool:
    expected_password = str(getattr(settings, 'DATA_TOOLS_PASSWORD', '') or '')
    provided_password = request.POST.get(DATA_TOOLS_PASSWORD_FIELD, '')

    if not expected_password:
        messages.error(
            request,
            'Пароль для инструментов данных не настроен. '
            'Добавьте pas_key_data в env-файл.',
        )
        return False

    if not compare_digest(provided_password, expected_password):
        messages.error(request, 'Неверный пароль подтверждения.')
        return False

    return True


def clear_database_data() -> dict[str, int]:
    UserModel = get_user_model()
    protected_user_ids = set(
        UserModel.objects.filter(is_superuser=True).values_list('id', flat=True),
    )
    protected_user_ids.update(
        UserModel.objects.filter(is_staff=True).values_list('id', flat=True),
    )

    deleted_counts: dict[str, int] = {}

    with transaction.atomic():
        for model in (
            TemporaryCredential,
            CourseApplication,
            SubjectResult,
            Grade,
            StudentSubject,
            GroupSubject,
            TeacherSubject,
            StudentEnrollment,
            Student,
            Teacher,
            StudyGroup,
            Subject,
            Instrument,
            AcademicYear,
            PasswordRecoveryContact,
            CourseRegistrationSettings,
        ):
            deleted_count, _deleted_by_model = model.objects.all().delete()
            deleted_counts[model._meta.label_lower] = deleted_count

        deleted_count, _deleted_by_model = (
            UserModel.objects
            .exclude(id__in=protected_user_ids)
            .delete()
        )
        deleted_counts[UserModel._meta.label_lower] = deleted_count

    return deleted_counts


@superuser_required
@require_GET
async def admin_export_test_credentials_excel_view(request: HttpRequest) -> HttpResponse:
    return await _run_db_sync(_admin_export_test_credentials_excel_view_sync, request)


def _admin_export_test_credentials_excel_view_sync(request: HttpRequest) -> HttpResponse:
    """
    Экспорт временных учетных данных в Excel.

    Файл формируется прямо в памяти и сразу отдается администратору.
    """
    workbook = build_temporary_credentials_workbook()

    now = timezone.localtime()
    filename = f'temporary_credentials_{now:%Y-%m-%d_%H-%M}.xlsx'
    encoded_filename = quote(filename)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f"attachment; filename={filename}; filename*=UTF-8''{encoded_filename}"
    )

    return response


def build_temporary_credentials_workbook() -> Workbook:
    """
    Создает Excel-файл с временными учетными данными.

    В выгрузку попадают только данные, которые нужны для выдачи пользователям:
    - логин;
    - временный пароль;
    - роль пользователя.
    """
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = 'Временные доступы'

    worksheet.append([
        'Логин',
        'Пароль',
        'Роль',
    ])

    queryset = (
        TemporaryCredential.objects
        .select_related('user')
        .prefetch_related('user__groups')
        .order_by('id')
    )

    for credential in queryset:
        worksheet.append([
            clean_excel_text(get_credential_login(credential)),
            clean_excel_text(get_credential_password(credential)),
            clean_excel_text(get_credential_role(credential)),
        ])

    format_sheet(worksheet)

    return workbook

def get_credential_login(credential: TemporaryCredential) -> str:
    """
    Возвращает логин из временных учетных данных.

    Поддерживает разные варианты названия поля,
    чтобы экспорт не ломался при изменениях модели.
    """
    return (
        getattr(credential, 'login', None)
        or getattr(credential, 'username', None)
        or getattr(getattr(credential, 'user', None), 'username', None)
        or ''
    )


def get_credential_password(credential: TemporaryCredential) -> str:
    """
    Возвращает временный пароль из временных учетных данных.
    """
    return (
        getattr(credential, 'temporary_password', None)
        or getattr(credential, 'password', None)
        or ''
    )


def get_credential_role(credential: TemporaryCredential) -> str:
    """
    Возвращает роль пользователя.

    Роль берется:
    1. по логину из связанного User;
    2. из групп Django;
    3. из признаков is_superuser/is_staff;
    4. если роль определить нельзя — возвращается пустая строка.
    """
    login = get_credential_login(credential)

    if not login:
        return ''

    user = getattr(credential, 'user', None)
    if user is None:
        user = (
            get_user_model()
            .objects
            .filter(username=login)
            .first()
        )

    if user is None:
        return ''

    group_names = list(
        user.groups
        .order_by('name')
        .values_list('name', flat=True)
    )

    if group_names:
        return ', '.join(group_names)

    if user.is_superuser:
        return 'Администратор'

    if user.is_staff:
        return 'Преподаватель'

    return 'Ученик'


def get_export_fields(model):
    """
    Возвращает обычные поля модели для Excel-выгрузки.

    ManyToMany и обратные связи не экспортируются, потому что они плохо
    ложатся в одну плоскую таблицу.
    """
    fields = []

    for field in model._meta.get_fields():
        if getattr(field, 'many_to_many', False):
            continue

        if getattr(field, 'one_to_many', False):
            continue

        if getattr(field, 'auto_created', False) and not getattr(field, 'concrete', False):
            continue

        if not getattr(field, 'concrete', False):
            continue

        fields.append(field)

    return fields


def field_value(obj, field):
    """
    Возвращает значение поля для Excel.

    Для ForeignKey и OneToOne показывает строковое представление связанного объекта.
    """
    if getattr(field, 'is_relation', False):
        related_obj = getattr(obj, field.name, None)
        return format_value(related_obj)

    return format_value(getattr(obj, field.name, None))


def format_value(value):
    """
    Приводит значение к формату, который корректно сохраняется в Excel.
    """
    if value is None:
        return ''

    if isinstance(value, bool):
        return 'Да' if value else 'Нет'

    if isinstance(value, datetime.datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)

        return value.replace(tzinfo=None)

    if isinstance(value, datetime.date):
        return value

    if isinstance(value, (int, float)):
        return value

    return clean_excel_text(str(value))


def clean_excel_text(value: str) -> str:
    """
    Удаляет символы, которые Excel не принимает.
    """
    value = ILLEGAL_CHARACTERS_RE.sub('', value)
    if value.lstrip().startswith(('=', '+', '-', '@')):
        return f"'{value}"
    return value


def format_sheet(worksheet) -> None:
    """
    Делает лист Excel удобным для чтения:
    - закрепляет верхнюю строку;
    - включает фильтр;
    - выделяет заголовки;
    - подбирает ширину колонок.
    """
    worksheet.freeze_panes = 'A2'

    if worksheet.max_row >= 1 and worksheet.max_column >= 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True,
        )

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical='top',
                wrap_text=True,
            )

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            value_length = len(str(value))

            if value_length > max_length:
                max_length = value_length

        width = min(
            max(max_length + 2, DEFAULT_COLUMN_WIDTH),
            MAX_COLUMN_WIDTH,
        )
        worksheet.column_dimensions[column_letter].width = width


def safe_reverse(url_name: str) -> str | None:
    """
    Безопасный reverse для необязательных ссылок.

    Например, полная Excel-выгрузка может быть подключена не сразу.
    """
    try:
        return reverse(url_name)
    except NoReverseMatch:
        return None


def get_admin_url(url_name: str, fallback: str = '/admin/') -> str:
    """
    Возвращает ссылку на страницу Django Admin.
    """
    try:
        return reverse(url_name)
    except NoReverseMatch:
        return fallback
